#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import re
import shutil
import sqlite3
import sys
import uuid
from datetime import datetime, date

from openpyxl import load_workbook


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'database.db')

FILES = [
    ('张有德牛肉面', '/Users/donggenyuan/Desktop/7月酒店订单/0726/牛肉面7月最新.xlsx'),
    ('全季酒店（辛家庙店）', '/Users/donggenyuan/Desktop/7月酒店订单/0726/全季酒店7月.xlsx'),
    ('西安汉庭酒店（大明宫万达）', '/Users/donggenyuan/Desktop/7月酒店订单/0726/汉庭酒店7月.xlsx'),
    ('浐灞美居', '/Users/donggenyuan/Desktop/7月酒店订单/0726/浐灞美居酒店7月.xlsx'),
]

ALIASES = {
    '蒜薹': '蒜苔',
    '胡萝卜': '红萝卜',
    '广红': '红萝卜',
    '红萝卜': '红萝卜',
    '有机菜花': '菜花',
    '花菜': '菜花',
    '菜花': '菜花',
    '毛芹': '麦芹',
    '波菜': '菠菜',
    '连菜': '莲菜',
    '手工面条': '手工面',
    '细薄非叶面': '细薄韭叶面',
    '韭叶面': '细薄韭叶面',
    '蕃茄酱': '番茄酱',
    '西红柿酱': '番茄酱',
    '糯米': '江米',
    '火龙果': '白心火龙果',
    '立昂火腿': '里昂火腿',
    '之宝大豆油': '大豆油',
    '小青菜': '青菜（把）',
    '青菜': '青菜（把）',
    '青菜（把）': '青菜（把）',
    '黑豆腐丝': '红豆腐丝',
    '方块儿面筋': '方块面筋',
    '立昂火腿': '里昂火腿',
}


def now_str():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def generate_id():
    return str(uuid.uuid4())


def as_float(value):
    try:
        if value is None:
            return 0.0
        text = str(value).replace(',', '').replace('¥', '').strip()
        return float(text) if text else 0.0
    except Exception:
        return 0.0


def normalize_date(value):
    text = str(value or '').strip()
    match = re.search(r'(\d{4})[-./年](\d{1,2})[-./月](\d{1,2})', text)
    if match:
        y, m, d = match.groups()
        return f'{int(y):04d}-{int(m):02d}-{int(d):02d}'
    return date.today().strftime('%Y-%m-%d')


def normalize_name(name):
    text = str(name or '').strip()
    text = re.sub(r'\s+', '', text)
    return ALIASES.get(text, text)


def normalize_customer_key(customer):
    text = str(customer or '').strip()
    if '美居' in text:
        return '浐灞美居'
    if '汉庭' in text:
        return '西安汉庭酒店（大明宫万达）'
    if '全季' in text:
        return '全季酒店（辛家庙店）'
    if '牛肉面' in text or '张有德' in text:
        return '张有德牛肉面'
    return text or '酒店'


def canonical_customer(conn, customer):
    text = str(customer or '').strip()
    rows = conn.execute('SELECT name FROM customers ORDER BY created_at DESC').fetchall()
    for row in rows:
        name = row['name']
        if text == name or text in name or name in text:
            return name
    compact = re.sub(r'[\s（）()]', '', text)
    for row in rows:
        name = row['name']
        name_compact = re.sub(r'[\s（）()]', '', name)
        if compact and (compact in name_compact or name_compact in compact):
            return name
    return normalize_customer_key(text)


def ensure_customer(conn, customer):
    name = canonical_customer(conn, customer)
    if not name:
        return name
    existing = conn.execute('SELECT id FROM customers WHERE name = ? LIMIT 1', (name,)).fetchone()
    if existing:
        return name
    code = 'C' + datetime.now().strftime('%y%m%d') + str(uuid.uuid4().int % 9000 + 1000)
    conn.execute('''
        INSERT INTO customers (id, code, name, company, phone, address, note, created_at)
        VALUES (?, ?, ?, ?, '', '', '', ?)
    ''', (generate_id(), code, name, name, now_str()))
    return name


def ensure_tables(conn):
    conn.execute('''
        CREATE TABLE IF NOT EXISTS customer_product_prices (
            id TEXT PRIMARY KEY,
            customer TEXT NOT NULL,
            product_id TEXT NOT NULL,
            product_name TEXT,
            product_unit TEXT,
            price REAL NOT NULL,
            source TEXT,
            updated_at TEXT,
            created_at TEXT,
            UNIQUE(customer, product_id, product_unit)
        )
    ''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_customer_product_prices_customer_product ON customer_product_prices(customer, product_id, product_unit)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_customer_product_prices_name ON customer_product_prices(customer, product_name, product_unit)')


def next_code(conn):
    while True:
        code = 'HT' + datetime.now().strftime('%y%m%d') + str(uuid.uuid4().int % 9000 + 1000)
        if not conn.execute('SELECT 1 FROM products WHERE code = ?', (code,)).fetchone():
            return code


def find_or_create_product(conn, name, unit, price):
    row = conn.execute('SELECT * FROM products WHERE name = ? AND unit = ? LIMIT 1', (name, unit)).fetchone()
    if not row:
        row = conn.execute('SELECT * FROM products WHERE name = ? LIMIT 1', (name,)).fetchone()
    if not row:
        row = conn.execute('SELECT * FROM products WHERE name LIKE ? LIMIT 1', (f'%{name}%',)).fetchone()
    if row:
        return dict(row), False
    product_id = generate_id()
    conn.execute('''
        INSERT INTO products (id, code, name, spec, unit, cost, stock, alert_line, departments, last_sale_price, created_at)
        VALUES (?, ?, ?, '', ?, ?, 0, 0, '[]', ?, ?)
    ''', (product_id, next_code(conn), name, unit or '件', as_float(price), as_float(price), now_str()))
    row = conn.execute('SELECT * FROM products WHERE id = ?', (product_id,)).fetchone()
    return dict(row), True


def upsert_price(conn, customer, product, unit, price, source):
    existing = conn.execute('''
        SELECT id FROM customer_product_prices
        WHERE customer = ? AND product_id = ? AND COALESCE(product_unit, '') = ?
        LIMIT 1
    ''', (customer, product['id'], unit or '')).fetchone()
    if existing:
        conn.execute('''
            UPDATE customer_product_prices
            SET product_name = ?, price = ?, source = ?, updated_at = ?
            WHERE id = ?
        ''', (product['name'], price, source, now_str(), existing['id']))
        return 'updated'
    conn.execute('''
        INSERT INTO customer_product_prices
            (id, customer, product_id, product_name, product_unit, price, source, updated_at, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (generate_id(), customer, product['id'], product['name'], unit or '', price, source, now_str(), now_str()))
    return 'created'


def extract_customer_date(values, default_customer):
    joined = ' '.join(str(v) for v in values if v is not None)
    customer = default_customer
    date_text = ''
    for value in values:
        text = str(value or '').strip()
        if '客户' in text:
            customer = re.sub(r'^客户[:：]?', '', text).strip() or customer
        if '日期' in text:
            date_text = text
    return customer, normalize_date(date_text)


def parse_files():
    files = FILES
    if len(sys.argv) > 2:
        files = []
        for arg in sys.argv[2:]:
            if '=' in arg:
                customer, path = arg.split('=', 1)
            else:
                customer, path = '', arg
            inferred = customer or os.path.basename(path).replace('.xlsx', '')
            files.append((inferred, path))
    latest = {}
    total_rows = 0
    for default_customer, path in files:
        if not os.path.exists(path):
            print(f'缺少文件: {path}')
            continue
        wb = load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            current_customer = default_customer
            current_date = date.today().strftime('%Y-%m-%d')
            for row_idx in range(1, ws.max_row + 1):
                values = [ws.cell(row_idx, col).value for col in range(1, 9)]
                first = str(values[0] or '').strip()
                row_text = ' '.join(str(v) for v in values if v is not None)
                if '客户' in row_text and '日期' in row_text:
                    current_customer, current_date = extract_customer_date(values, default_customer)
                    current_customer = normalize_customer_key(current_customer)
                    continue
                if first in ('', '序号') or first.startswith('合计'):
                    continue
                if not isinstance(values[0], (int, float)):
                    continue
                name = normalize_name(values[1])
                unit = str(values[4] or '').strip()
                price = as_float(values[5])
                amount = as_float(values[6])
                quantity = as_float(values[3])
                if price <= 0 and quantity > 0 and amount > 0:
                    price = amount / quantity
                if not name or not unit or price <= 0:
                    continue
                total_rows += 1
                key = (normalize_customer_key(current_customer), name, unit)
                sort_key = (current_date, row_idx)
                if key not in latest or sort_key >= latest[key]['sort_key']:
                    latest[key] = {
                        'customer': current_customer,
                        'date': current_date,
                        'name': name,
                        'unit': unit,
                        'price': price,
                        'source': os.path.basename(path),
                        'sort_key': sort_key,
                    }
    return latest, total_rows


def main():
    db_path = sys.argv[1] if len(sys.argv) > 1 else DB_PATH
    backup = f'{db_path}.bak_customer_prices_{datetime.now().strftime("%Y%m%d%H%M%S")}'
    shutil.copy2(db_path, backup)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        ensure_tables(conn)
        latest, total_rows = parse_files()
        created = updated = created_products = 0
        samples = []
        for row in latest.values():
            customer = ensure_customer(conn, row['customer'])
            product, was_created = find_or_create_product(conn, row['name'], row['unit'], row['price'])
            action = upsert_price(conn, customer, product, row['unit'], row['price'], row['source'])
            created += action == 'created'
            updated += action == 'updated'
            created_products += was_created
            if len(samples) < 10:
                samples.append((customer, product['name'], row['unit'], row['price']))
        conn.commit()
        print({
            'db': db_path,
            'backup': backup,
            'source_rows': total_rows,
            'latest_prices': len(latest),
            'created_prices': created,
            'updated_prices': updated,
            'created_products': created_products,
            'samples': samples,
        })
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == '__main__':
    main()
