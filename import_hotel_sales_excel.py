#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Import hotel sales orders from the May order workbook."""

import os
import random
import re
import shutil
import sqlite3
import sys
import uuid
from datetime import datetime

from openpyxl import load_workbook


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'database.db')
BACKUP_DIR = os.path.join(BASE_DIR, 'migration_backups')
DEFAULT_XLSX = '/Volumes/ORICO/酒店/采购/5月酒店/酒店5月订单.xlsx'


def now_str():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def make_id():
    return str(uuid.uuid4())


def to_float(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def normalize_date(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    text = str(value or '').strip().replace('/', '-')
    match = re.search(r'(\d{4}-\d{1,2}-\d{1,2})', text)
    if match:
        parts = match.group(1).split('-')
        return f'{int(parts[0]):04d}-{int(parts[1]):02d}-{int(parts[2]):02d}'
    return text


def backup_database():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(BACKUP_DIR, f'database_before_hotel_sales_import_{stamp}.db')
    shutil.copy2(DB_PATH, path)
    return path


def parse_workbook(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb['订单明细'] if '订单明细' in wb.sheetnames else wb.worksheets[0]
    orders = []
    current = None

    for row in ws.iter_rows(values_only=True):
        first = row[0]
        date_cell = row[6] if len(row) > 6 else None

        if isinstance(first, str) and first.strip().startswith('客户'):
            if current and current['items']:
                orders.append(current)
            customer = first.split('：', 1)[-1].strip()
            current = {
                'customer': customer,
                'company': '西安禾润佳商贸有限公司',
                'phone': '',
                'date': normalize_date(date_cell),
                'items': [],
                'total': 0.0,
            }
            continue

        if not current:
            continue

        if isinstance(first, str) and first.strip().startswith('合计'):
            current['total'] = to_float(row[6] if len(row) > 6 else 0)
            orders.append(current)
            current = None
            continue

        if isinstance(first, int) and row[1]:
            quantity = to_float(row[3])
            price = to_float(row[5])
            subtotal = to_float(row[6]) or quantity * price
            current['items'].append({
                'name': str(row[1]).strip(),
                'spec': str(row[2] or '').strip(),
                'quantity': quantity,
                'unit': str(row[4] or '').strip(),
                'price': price,
                'subtotal': subtotal,
            })

    if current and current['items']:
        orders.append(current)

    for order in orders:
        if not order['total']:
            order['total'] = sum(item['subtotal'] for item in order['items'])
    return orders


def generate_no(conn, date_str):
    yymmdd = datetime.strptime(date_str, '%Y-%m-%d').strftime('%y%m%d')
    while True:
        no = f'XS{yymmdd}{random.randint(1000, 9999)}'
        if not conn.execute('SELECT 1 FROM sales WHERE no = ?', (no,)).fetchone():
            return no


def generate_product_code(conn, date_str):
    yymmdd = datetime.strptime(date_str, '%Y-%m-%d').strftime('%y%m%d')
    while True:
        code = f'IMP{yymmdd}{random.randint(1000, 9999)}'
        if not conn.execute('SELECT 1 FROM products WHERE code = ?', (code,)).fetchone():
            return code


def find_or_create_product(conn, item, date_str, created):
    row = conn.execute(
        'SELECT * FROM products WHERE name = ? AND unit = ? LIMIT 1',
        (item['name'], item['unit'])
    ).fetchone()
    if not row:
        row = conn.execute('SELECT * FROM products WHERE name = ? LIMIT 1', (item['name'],)).fetchone()
    if row:
        return dict(row)

    product_id = make_id()
    code = generate_product_code(conn, date_str)
    conn.execute('''
        INSERT INTO products (id, code, name, spec, unit, cost, stock, alert_line, departments, last_sale_price, created_at)
        VALUES (?, ?, ?, ?, ?, ?, 0, 10, '[]', ?, ?)
    ''', (
        product_id, code, item['name'], item['spec'], item['unit'],
        item['price'], item['price'], now_str()
    ))
    created.append(item['name'])
    return dict(conn.execute('SELECT * FROM products WHERE id = ?', (product_id,)).fetchone())


def import_orders(path):
    orders = parse_workbook(path)
    if not orders:
        raise RuntimeError('没有在工作簿中找到销售单')

    backup_path = backup_database()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')

    dates = sorted({order['date'] for order in orders})
    customers = sorted({order['customer'] for order in orders})
    created_products = []
    deleted_sales = 0
    inserted_items = 0

    try:
        for customer in customers:
            placeholders = ','.join('?' for _ in dates)
            existing = conn.execute(
                f'SELECT id FROM sales WHERE customer = ? AND date IN ({placeholders})',
                [customer, *dates]
            ).fetchall()
            for sale in existing:
                items = conn.execute(
                    'SELECT product_id, quantity FROM sales_items WHERE sale_id = ?',
                    (sale['id'],)
                ).fetchall()
                for item in items:
                    conn.execute(
                        "UPDATE products SET stock = stock + ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')",
                        (item['quantity'], item['product_id'])
                    )
                conn.execute('DELETE FROM sales_items WHERE sale_id = ?', (sale['id'],))
                conn.execute('DELETE FROM sales WHERE id = ?', (sale['id'],))
                deleted_sales += 1

        for order in orders:
            sale_id = make_id()
            no = generate_no(conn, order['date'])
            conn.execute('''
                INSERT INTO sales (id, no, company, customer, phone, date, total, show_handlers, handler, issuer, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, 0, '', '', ?)
            ''', (
                sale_id, no, order['company'], order['customer'], order['phone'],
                order['date'], order['total'], now_str()
            ))

            for item in order['items']:
                product = find_or_create_product(conn, item, order['date'], created_products)
                conn.execute('''
                    INSERT INTO sales_items (id, sale_id, product_id, product_name, product_spec, product_unit, quantity, price, subtotal)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    make_id(), sale_id, product['id'], product['name'], product['spec'],
                    product['unit'], item['quantity'], item['price'], item['subtotal']
                ))
                conn.execute('''
                    UPDATE products
                    SET stock = CASE
                            WHEN code LIKE 'WP%' OR code LIKE 'IMP%' THEN 0
                            ELSE stock - ?
                        END,
                        last_sale_price = ?
                    WHERE id = ?
                ''', (item['quantity'], item['price'], product['id']))
                inserted_items += 1

        fk_errors = conn.execute('PRAGMA foreign_key_check').fetchall()
        if fk_errors:
            raise RuntimeError(f'foreign key check failed: {fk_errors}')
        integrity = conn.execute('PRAGMA integrity_check').fetchone()[0]
        if integrity != 'ok':
            raise RuntimeError(f'integrity check failed: {integrity}')
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    print(f'来源文件: {path}')
    print(f'导入销售单: {len(orders)}')
    print(f'导入明细: {inserted_items}')
    print(f'替换旧销售单: {deleted_sales}')
    print(f'新增物品: {len(created_products)}')
    print(f'数据库备份: {backup_path}')


if __name__ == '__main__':
    import_orders(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX)
