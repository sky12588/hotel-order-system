#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
酒店订单管理系统 - 后端服务
基于 Flask + SQLite 的本地服务器
"""

import os
import base64
import hashlib
import hmac
import json
import sqlite3
import uuid
import csv
import io
import re
import shutil
import urllib.error
import urllib.request
import webbrowser
import threading
import time
from copy import copy
from datetime import datetime, date
from functools import wraps

from flask import Flask, request, jsonify, send_file, Response, redirect, render_template_string
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import qrcode

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)


PUBLIC_PATHS = (
    '/order',
    '/customer_order.html',
    '/api/public/',
    '/api/order-qr',
    '/api/print-agent/',
    '/5655a9d9c3904b26114c4e524f9021c3.txt',
)


def public_order_token():
    return os.environ.get('PUBLIC_ORDER_TOKEN', '').strip()


def public_order_url():
    host = (request.host or '').split(':', 1)[0]
    is_direct_host = bool(re.match(r'^\d{1,3}(\.\d{1,3}){3}$', host)) or host in ('localhost', '127.0.0.1')
    if is_direct_host:
        base = request.host_url.rstrip('/')
    else:
        base = os.environ.get('PUBLIC_ORDER_BASE_URL', '').strip().rstrip('/') or request.host_url.rstrip('/')
    token = public_order_token()
    customer_id = request.args.get('customerId', '').strip()
    url = f'{base}/order'
    query = []
    if token:
        query.append(f'token={token}')
    if customer_id:
        query.append(f'customerId={customer_id}')
    if query:
        url = f"{url}?{'&'.join(query)}"
    return url


def valid_public_order_request():
    token = public_order_token()
    if not token:
        return True
    body = request.get_json(silent=True) if request.is_json else {}
    provided = request.args.get('token') or request.headers.get('X-Order-Token') or (body or {}).get('token')
    return hmac.compare_digest(str(provided or ''), token)


PRINT_AGENT_TOKEN = os.environ.get('PRINT_AGENT_TOKEN', 'ChangeThisPrintAgentToken')
UPLOADED_EXCEL_BASE_DIR = os.environ.get('UPLOADED_EXCEL_BASE_DIR', '/home/ubuntu/hotel_uploads').strip()
UPLOADED_EXCEL_INCOMING_DIR = os.environ.get('UPLOADED_EXCEL_INCOMING_DIR', os.path.join(UPLOADED_EXCEL_BASE_DIR, 'incoming_excels')).strip()
UPLOADED_EXCEL_PROCESSED_DIR = os.environ.get('UPLOADED_EXCEL_PROCESSED_DIR', os.path.join(UPLOADED_EXCEL_BASE_DIR, 'processed')).strip()
UPLOADED_EXCEL_FAILED_DIR = os.environ.get('UPLOADED_EXCEL_FAILED_DIR', os.path.join(UPLOADED_EXCEL_BASE_DIR, 'failed')).strip()
UPLOADED_EXCEL_MANIFEST_PATH = os.environ.get('UPLOADED_EXCEL_MANIFEST_PATH', os.path.join(UPLOADED_EXCEL_BASE_DIR, '.processed_manifest.json')).strip()
FEISHU_APP_ID = os.environ.get('FEISHU_APP_ID', '').strip()
FEISHU_APP_SECRET = os.environ.get('FEISHU_APP_SECRET', '').strip()
FEISHU_BITABLE_APP_TOKEN = os.environ.get('FEISHU_BITABLE_APP_TOKEN', '').strip()
FEISHU_BITABLE_TABLE_ID = os.environ.get('FEISHU_BITABLE_TABLE_ID', '').strip()
FEISHU_SYNC_ENABLED = os.environ.get('FEISHU_SYNC_ENABLED', '').strip().lower() in ('1', 'true', 'yes', 'on')
FEISHU_SYNC_INTERVAL_SECONDS = int(os.environ.get('FEISHU_SYNC_INTERVAL_SECONDS', '60') or 60)
_feishu_token_cache = {'token': '', 'expires_at': 0}
_feishu_sync_thread_started = False
_print_job_condition = threading.Condition()
BLOG_SAFE_POSTS = [
    {
        'date': '',
        'source': '生活笔记',
        'tag': '散文',
        'title': '把复杂的事情写成清单',
        'summary': '清单不是为了让人更忙，而是让脑子少背一点东西。把下一步写清楚，行动就会变得轻一些。',
        'link': '',
    },
    {
        'date': '',
        'source': '技术摘记',
        'tag': 'AI 技术',
        'title': 'AI 工具更像新的工作台',
        'summary': 'AI 正在变成文档、表格、图片和系统之间的连接层。真正有用的部分，是把重复流程变短。',
        'link': '',
    },
    {
        'date': '',
        'source': '观察',
        'tag': '新闻',
        'title': '关于小城新闻的一点观察',
        'summary': '很多新闻离宏大的叙事很远，却和日常很近：一条道路、一场雨、一个市场，都会影响人的生活节奏。',
        'link': '',
    },
    {
        'date': '',
        'source': '读书摘记',
        'tag': '阅读',
        'title': '慢一点读完一本书',
        'summary': '读书有时候不是为了记住所有观点，而是让某一句话在恰当的时候回来，帮人把心里的线理顺。',
        'link': '',
    },
    {
        'date': '',
        'source': '工具笔记',
        'tag': '效率',
        'title': '少一点切换，多一点完成',
        'summary': '工具越多，切换成本越高。真正适合自己的流程，往往是按钮少、路径短、能反复使用。',
        'link': '',
    },
    {
        'date': '',
        'source': '技术摘记',
        'tag': 'AI 学习',
        'title': '好提示词先说清楚边界',
        'summary': '让 AI 做事时，先说明目标、输入、输出和不能做什么，比堆很多形容词更有效。',
        'link': '',
    },
    {
        'date': '',
        'source': '日常',
        'tag': '散文',
        'title': '夏天的午后',
        'summary': '阳光落在墙面上，像一层慢慢移动的薄毯。这样的时刻适合读几页书，也适合什么都不做。',
        'link': '',
    },
    {
        'date': '',
        'source': '备忘',
        'tag': '计划',
        'title': '给未来自己的备忘',
        'summary': '不要等所有条件都齐了才开始。先搭一个能用的小版本，再慢慢让它变得顺手。',
        'link': '',
    },
    {
        'date': '',
        'source': '技术摘记',
        'tag': '自动化',
        'title': '重复的事情适合交给流程',
        'summary': '当一件事每周都要重复，就值得花一点时间把它做成流程。省下来的不是一分钟，而是注意力。',
        'link': '',
    },
    {
        'date': '',
        'source': '生活笔记',
        'tag': '整理',
        'title': '桌面清爽一点，心也松一点',
        'summary': '整理不是把所有东西藏起来，而是让常用的东西更容易被找到，让不常用的东西安静待着。',
        'link': '',
    },
    {
        'date': '',
        'source': '观察',
        'tag': '城市',
        'title': '菜市场里的时间感',
        'summary': '菜市场最能看见一天的节奏。清晨忙，午后慢，傍晚又热闹起来，生活就在这些声音里流动。',
        'link': '',
    },
    {
        'date': '',
        'source': '技术摘记',
        'tag': '数据',
        'title': '表格最重要的是口径一致',
        'summary': '数据表不怕简单，怕的是同一件事有几种叫法。名称、单位和日期统一，后面的统计才可靠。',
        'link': '',
    },
]


def fetch_blog_posts():
    today = date.today()
    offset = today.toordinal() % len(BLOG_SAFE_POSTS)
    rotated = BLOG_SAFE_POSTS[offset:] + BLOG_SAFE_POSTS[:offset]
    posts = []
    for index, post in enumerate(rotated[:9]):
        item = dict(post)
        item['date'] = (today.fromordinal(today.toordinal() - index)).strftime('%Y-%m-%d')
        item['link'] = ''
        posts.append(item)
    return posts


def valid_print_agent_request():
    token = PRINT_AGENT_TOKEN.strip()
    if not token:
        return True
    body = request.get_json(silent=True) if request.is_json else {}
    provided = request.args.get('token') or request.headers.get('X-Print-Agent-Token') or (body or {}).get('token')
    return hmac.compare_digest(str(provided or ''), token)


def basic_auth_required():
    username = os.environ.get('HOTEL_AUTH_USER', '').strip()
    password = os.environ.get('HOTEL_AUTH_PASSWORD', '').strip()
    return bool(username and password)


def unauthorized_response():
    return Response(
        '需要登录',
        401,
        {'WWW-Authenticate': 'Basic realm="Hotel Order System"'}
    )


@app.before_request
def require_basic_auth():
    if request.path == '/':
        return None
    if any(request.path == path or request.path.startswith(path) for path in PUBLIC_PATHS):
        return None
    if not basic_auth_required():
        return None
    auth = request.headers.get('Authorization', '')
    if not auth.startswith('Basic '):
        return unauthorized_response()
    try:
        decoded = base64.b64decode(auth.split(' ', 1)[1]).decode('utf-8')
        username, password = decoded.split(':', 1)
    except Exception:
        return unauthorized_response()
    expected_user = os.environ.get('HOTEL_AUTH_USER', '')
    expected_password = os.environ.get('HOTEL_AUTH_PASSWORD', '')
    if not (
        hmac.compare_digest(username, expected_user)
        and hmac.compare_digest(password, expected_password)
    ):
        return unauthorized_response()
    return None

# 数据库路径
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.db')
SALES_EXPORT_TEMPLATE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sales_export_template.xlsx')
ARK_BASE_URL = os.environ.get('ARK_BASE_URL', 'https://ark.cn-beijing.volces.com/api/v3').rstrip('/')
ARK_API_KEY = os.environ.get('ARK_API_KEY') or os.environ.get('VOLCENGINE_ARK_API_KEY')
ARK_MODEL = os.environ.get('ARK_MODEL') or os.environ.get('ARK_MODEL_ID') or os.environ.get('VOLCENGINE_ARK_MODEL')
AI_RESPONSE_FORMAT_JSON = str(os.environ.get('AI_RESPONSE_FORMAT_JSON', '1')).strip().lower() not in ('0', 'false', 'no', 'off')


def env_float(name, default):
    try:
        return float(os.environ.get(name) or default)
    except (TypeError, ValueError):
        return float(default)


def ai_provider_name():
    base_url = str(ARK_BASE_URL or '').lower()
    if 'tencentmaas' in base_url or 'tokenhub' in base_url:
        return 'tencent_tokenhub'
    if 'volces' in base_url or 'ark' in base_url:
        return 'volcengine_ark'
    return 'openai_compatible'


AI_DAILY_LIMIT_CNY = env_float('AI_DAILY_LIMIT_CNY', 2)
AI_TEXT_FALLBACK_COST_CNY = env_float('AI_TEXT_FALLBACK_COST_CNY', 0.01)
AI_INPUT_PRICE_PER_1K_CNY = env_float('AI_INPUT_PRICE_PER_1K_CNY', 0.0015)
AI_OUTPUT_PRICE_PER_1K_CNY = env_float('AI_OUTPUT_PRICE_PER_1K_CNY', 0.0045)

# ==================== 数据库操作 ====================

class Database:
    def __init__(self, db_path=DB_PATH):
        self.db_path = db_path
        self.init_tables()

    def _get_conn(self):
        conn = sqlite3.connect(self.db_path, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute("PRAGMA busy_timeout = 5000")
        return conn

    def execute(self, sql, params=()):
        conn = self._get_conn()
        try:
            cursor = conn.execute(sql, params)
            conn.commit()
            return cursor
        finally:
            conn.close()

    def query(self, sql, params=()):
        conn = self._get_conn()
        try:
            cursor = conn.execute(sql, params)
            rows = cursor.fetchall()
            return [dict(row) for row in rows]
        finally:
            conn.close()

    def query_one(self, sql, params=()):
        conn = self._get_conn()
        try:
            cursor = conn.execute(sql, params)
            row = cursor.fetchone()
            return dict(row) if row else None
        finally:
            conn.close()

    def init_tables(self):
        """初始化数据库表结构"""
        conn = self._get_conn()
        try:
            conn.execute("PRAGMA journal_mode = WAL")
            # 物品表
            conn.execute('''
                CREATE TABLE IF NOT EXISTS products (
                    id TEXT PRIMARY KEY,
                    code TEXT UNIQUE NOT NULL,
                    name TEXT NOT NULL,
                    spec TEXT,
                    unit TEXT NOT NULL,
                    cost REAL DEFAULT 0,
                    stock REAL DEFAULT 0,
                    alert_line REAL DEFAULT 10,
                    departments TEXT DEFAULT '[]',
                    last_sale_price REAL DEFAULT 0,
                    created_at TEXT
                )
            ''')

            # 供应商表
            conn.execute('''
                CREATE TABLE IF NOT EXISTS suppliers (
                    id TEXT PRIMARY KEY,
                    code TEXT UNIQUE NOT NULL,
                    name TEXT NOT NULL,
                    contact TEXT,
                    phone TEXT,
                    address TEXT,
                    note TEXT,
                    created_at TEXT
                )
            ''')

            # 客户表
            conn.execute('''
                CREATE TABLE IF NOT EXISTS customers (
                    id TEXT PRIMARY KEY,
                    code TEXT UNIQUE NOT NULL,
                    name TEXT NOT NULL,
                    company TEXT,
                    phone TEXT,
                    address TEXT,
                    note TEXT,
                    created_at TEXT
                )
            ''')

            # 采购入库单
            conn.execute('''
                CREATE TABLE IF NOT EXISTS purchases (
                    id TEXT PRIMARY KEY,
                    no TEXT UNIQUE NOT NULL,
                    supplier_id TEXT,
                    supplier_name TEXT,
                    date TEXT NOT NULL,
                    total REAL DEFAULT 0,
                    created_at TEXT
                )
            ''')

            # 采购入库明细
            conn.execute('''
                CREATE TABLE IF NOT EXISTS purchase_items (
                    id TEXT PRIMARY KEY,
                    purchase_id TEXT NOT NULL,
                    product_id TEXT NOT NULL,
                    product_name TEXT,
                    product_spec TEXT,
                    product_unit TEXT,
                    quantity REAL NOT NULL,
                    price REAL NOT NULL,
                    subtotal REAL NOT NULL,
                    FOREIGN KEY (purchase_id) REFERENCES purchases(id) ON DELETE CASCADE
                )
            ''')

            # 部门领用单
            conn.execute('''
                CREATE TABLE IF NOT EXISTS outbounds (
                    id TEXT PRIMARY KEY,
                    no TEXT UNIQUE NOT NULL,
                    department TEXT NOT NULL,
                    person TEXT,
                    date TEXT NOT NULL,
                    total REAL DEFAULT 0,
                    purchase_id TEXT,
                    created_at TEXT
                )
            ''')

            # 部门领用明细
            conn.execute('''
                CREATE TABLE IF NOT EXISTS outbound_items (
                    id TEXT PRIMARY KEY,
                    outbound_id TEXT NOT NULL,
                    product_id TEXT NOT NULL,
                    product_name TEXT,
                    product_spec TEXT,
                    product_unit TEXT,
                    quantity REAL NOT NULL,
                    price REAL NOT NULL,
                    subtotal REAL NOT NULL,
                    FOREIGN KEY (outbound_id) REFERENCES outbounds(id) ON DELETE CASCADE
                )
            ''')

            # 销售出库单
            conn.execute('''
                CREATE TABLE IF NOT EXISTS sales (
                    id TEXT PRIMARY KEY,
                    no TEXT UNIQUE NOT NULL,
                    company TEXT,
                    customer TEXT,
                    phone TEXT,
                    date TEXT NOT NULL,
                    total REAL DEFAULT 0,
                    show_handlers INTEGER DEFAULT 0,
                    handler TEXT,
                    issuer TEXT,
                    order_key TEXT,
                    purchase_id TEXT,
                    created_at TEXT
                )
            ''')

            # 销售出库明细
            conn.execute('''
                CREATE TABLE IF NOT EXISTS sales_items (
                    id TEXT PRIMARY KEY,
                    sale_id TEXT NOT NULL,
                    product_id TEXT NOT NULL,
                    product_name TEXT,
                    product_spec TEXT,
                    product_unit TEXT,
                    quantity REAL NOT NULL,
                    price REAL NOT NULL,
                    subtotal REAL NOT NULL,
                    note TEXT,
                    FOREIGN KEY (sale_id) REFERENCES sales(id) ON DELETE CASCADE
                )
            ''')

            conn.execute('''
                CREATE TABLE IF NOT EXISTS hotel_stock_deductions (
                    sale_item_id TEXT PRIMARY KEY,
                    sale_id TEXT NOT NULL,
                    product_id TEXT NOT NULL,
                    quantity REAL NOT NULL
                )
            ''')

            # 财务记账
            conn.execute('''
                CREATE TABLE IF NOT EXISTS finances (
                    id TEXT PRIMARY KEY,
                    date TEXT NOT NULL,
                    type TEXT NOT NULL,
                    category TEXT,
                    amount REAL NOT NULL,
                    note TEXT,
                    created_at TEXT
                )
            ''')

            conn.execute('''
                CREATE TABLE IF NOT EXISTS ai_usage (
                    id TEXT PRIMARY KEY,
                    date TEXT NOT NULL,
                    provider TEXT,
                    model TEXT,
                    feature TEXT,
                    estimated_cost REAL DEFAULT 0,
                    prompt_tokens INTEGER DEFAULT 0,
                    completion_tokens INTEGER DEFAULT 0,
                    created_at TEXT
                )
            ''')

            conn.execute('''
                CREATE TABLE IF NOT EXISTS customer_product_prices (
                    id TEXT PRIMARY KEY,
                    customer TEXT NOT NULL,
                    product_id TEXT NOT NULL,
                    product_name TEXT,
                    product_unit TEXT,
                    price REAL NOT NULL,
                    source TEXT,
                    updated_at TEXT,
                    created_at TEXT,
                    UNIQUE(customer, product_id, product_unit)
                )
            ''')

            conn.execute('''
                CREATE TABLE IF NOT EXISTS print_jobs (
                    id TEXT PRIMARY KEY,
                    job_type TEXT NOT NULL,
                    title TEXT,
                    target_printer TEXT,
                    paper TEXT,
                    payload_json TEXT,
                    status TEXT DEFAULT 'pending',
                    error TEXT,
                    created_at TEXT,
                    claimed_at TEXT,
                    completed_at TEXT
                )
            ''')

            conn.execute('''
                CREATE TABLE IF NOT EXISTS feishu_sync_records (
                    id TEXT PRIMARY KEY,
                    sale_id TEXT NOT NULL,
                    sale_item_id TEXT NOT NULL,
                    record_id TEXT NOT NULL,
                    status TEXT DEFAULT '待确认',
                    last_error TEXT,
                    synced_at TEXT,
                    created_at TEXT,
                    UNIQUE(sale_item_id)
                )
            ''')

            # 迁移：为已有表添加 purchase_id 字段
            try:
                conn.execute("ALTER TABLE outbounds ADD COLUMN purchase_id TEXT")
            except Exception:
                pass
            try:
                conn.execute("ALTER TABLE sales ADD COLUMN purchase_id TEXT")
            except Exception:
                pass
            try:
                conn.execute("ALTER TABLE sales ADD COLUMN order_key TEXT")
            except Exception:
                pass
            try:
                conn.execute("ALTER TABLE sales_items ADD COLUMN note TEXT")
            except Exception:
                pass

            indexes = [
                'CREATE INDEX IF NOT EXISTS idx_purchases_date ON purchases(date)',
                'CREATE INDEX IF NOT EXISTS idx_purchases_date_created ON purchases(date, created_at)',
                'CREATE INDEX IF NOT EXISTS idx_outbounds_date ON outbounds(date)',
                'CREATE INDEX IF NOT EXISTS idx_outbounds_date_created ON outbounds(date, created_at)',
                'CREATE INDEX IF NOT EXISTS idx_sales_date ON sales(date)',
                'CREATE INDEX IF NOT EXISTS idx_sales_date_created ON sales(date, created_at)',
                'CREATE INDEX IF NOT EXISTS idx_finances_date ON finances(date)',
                'CREATE INDEX IF NOT EXISTS idx_finances_date_created ON finances(date, created_at)',
                'CREATE INDEX IF NOT EXISTS idx_products_created_at ON products(created_at)',
                'CREATE INDEX IF NOT EXISTS idx_products_stock_alert ON products(stock, alert_line)',
                'CREATE INDEX IF NOT EXISTS idx_purchase_items_purchase_id ON purchase_items(purchase_id)',
                'CREATE INDEX IF NOT EXISTS idx_purchase_items_product_id ON purchase_items(product_id)',
                'CREATE INDEX IF NOT EXISTS idx_outbound_items_outbound_id ON outbound_items(outbound_id)',
                'CREATE INDEX IF NOT EXISTS idx_outbound_items_product_id ON outbound_items(product_id)',
                'CREATE INDEX IF NOT EXISTS idx_sales_items_sale_id ON sales_items(sale_id)',
                'CREATE INDEX IF NOT EXISTS idx_sales_items_product_id ON sales_items(product_id)',
                'CREATE INDEX IF NOT EXISTS idx_hotel_stock_deductions_sale_id ON hotel_stock_deductions(sale_id)',
                'CREATE INDEX IF NOT EXISTS idx_outbounds_purchase_id ON outbounds(purchase_id)',
                'CREATE INDEX IF NOT EXISTS idx_sales_purchase_id ON sales(purchase_id)',
                'CREATE INDEX IF NOT EXISTS idx_ai_usage_date ON ai_usage(date)',
                'CREATE INDEX IF NOT EXISTS idx_customer_product_prices_customer_product ON customer_product_prices(customer, product_id, product_unit)',
                'CREATE INDEX IF NOT EXISTS idx_customer_product_prices_name ON customer_product_prices(customer, product_name, product_unit)',
                'CREATE INDEX IF NOT EXISTS idx_print_jobs_status_created ON print_jobs(status, created_at)',
                'CREATE INDEX IF NOT EXISTS idx_feishu_sync_records_sale_id ON feishu_sync_records(sale_id)',
                'CREATE INDEX IF NOT EXISTS idx_feishu_sync_records_record_id ON feishu_sync_records(record_id)'
            ]
            for sql in indexes:
                conn.execute(sql)

            conn.commit()
        finally:
            conn.close()


db = Database()

# ==================== 工具函数 ====================

def generate_id():
    """生成唯一ID"""
    return str(uuid.uuid4())


def generate_no(prefix):
    """生成单号"""
    now = datetime.now()
    date_str = now.strftime('%y%m%d')
    seq = str(uuid.uuid4().int % 9000 + 1000)
    return f"{prefix}{date_str}{seq}"


def now_str():
    """当前时间字符串"""
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def normalize_date(value):
    """兼容旧版 2026/4/27 等日期格式，统一为 YYYY-MM-DD。"""
    if not value:
        return date.today().strftime('%Y-%m-%d')
    text = str(value).strip().replace('/', '-')
    for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    parts = text.split('-')
    if len(parts) == 3 and all(part.isdigit() for part in parts):
        return f"{int(parts[0]):04d}-{int(parts[1]):02d}-{int(parts[2]):02d}"
    return text


def as_float(value, default=0):
    try:
        return float(value if value not in (None, '') else default)
    except (TypeError, ValueError):
        return float(default)


def today_str():
    return date.today().strftime('%Y-%m-%d')


def ai_usage_today_cost():
    row = db.query_one(
        'SELECT COALESCE(SUM(estimated_cost), 0) AS total FROM ai_usage WHERE date = ?',
        (today_str(),)
    )
    return as_float(row.get('total') if row else 0)


def ai_budget_available(estimated_cost):
    if AI_DAILY_LIMIT_CNY <= 0:
        return True, ''
    used = ai_usage_today_cost()
    if used + as_float(estimated_cost) > AI_DAILY_LIMIT_CNY:
        return False, f'今日AI识别额度已用完（上限{AI_DAILY_LIMIT_CNY:.2f}元，已估算使用{used:.2f}元），请明天再用拍照/智能识别，或改为手动输入。'
    return True, ''


def estimate_ai_cost_from_usage(usage, fallback_cost):
    if not isinstance(usage, dict):
        return fallback_cost, 0, 0
    prompt_tokens = int(usage.get('prompt_tokens') or usage.get('input_tokens') or 0)
    completion_tokens = int(usage.get('completion_tokens') or usage.get('output_tokens') or 0)
    if prompt_tokens <= 0 and completion_tokens <= 0:
        return fallback_cost, 0, 0
    cost = (prompt_tokens / 1000) * AI_INPUT_PRICE_PER_1K_CNY + (completion_tokens / 1000) * AI_OUTPUT_PRICE_PER_1K_CNY
    return max(cost, 0), prompt_tokens, completion_tokens


def record_ai_usage(provider, model, feature, estimated_cost, prompt_tokens=0, completion_tokens=0):
    try:
        db.execute('''
            INSERT INTO ai_usage (id, date, provider, model, feature, estimated_cost, prompt_tokens, completion_tokens, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            generate_id(), today_str(), provider, model or '', feature,
            as_float(estimated_cost), int(prompt_tokens or 0), int(completion_tokens or 0), now_str()
        ))
    except Exception:
        pass


def clean_number(value, digits=4):
    number = as_float(value)
    if abs(number) < 0.000001:
        return 0
    return round(number, digits)


def feishu_configured():
    return bool(FEISHU_APP_ID and FEISHU_APP_SECRET and FEISHU_BITABLE_APP_TOKEN and FEISHU_BITABLE_TABLE_ID)


def feishu_date_millis(value):
    text = normalize_date(value)
    try:
        dt = datetime.strptime(text, '%Y-%m-%d')
        return int(time.mktime(dt.timetuple()) * 1000)
    except ValueError:
        return int(time.time() * 1000)


def feishu_field_value(value):
    if isinstance(value, list):
        parts = []
        for item in value:
            if isinstance(item, dict):
                parts.append(str(item.get('text') or item.get('name') or item.get('value') or ''))
            else:
                parts.append(str(item))
        return ''.join(parts).strip()
    if isinstance(value, dict):
        return str(value.get('text') or value.get('name') or value.get('value') or '').strip()
    return str(value or '').strip()


def feishu_get_token():
    if not feishu_configured():
        raise RuntimeError('飞书同步未配置')
    now_ts = time.time()
    if _feishu_token_cache.get('token') and _feishu_token_cache.get('expires_at', 0) > now_ts + 120:
        return _feishu_token_cache['token']
    req = urllib.request.Request(
        'https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal',
        data=json.dumps({'app_id': FEISHU_APP_ID, 'app_secret': FEISHU_APP_SECRET}).encode(),
        headers={'Content-Type': 'application/json; charset=utf-8'},
        method='POST'
    )
    with urllib.request.urlopen(req, timeout=20) as resp:
        data = json.loads(resp.read())
    if data.get('code') != 0:
        raise RuntimeError(data.get('msg') or '获取飞书 token 失败')
    expire = int(data.get('expire') or 7200)
    _feishu_token_cache['token'] = data.get('tenant_access_token') or ''
    _feishu_token_cache['expires_at'] = now_ts + expire
    return _feishu_token_cache['token']


def feishu_api(method, path, data=None):
    token = feishu_get_token()
    url = 'https://open.feishu.cn/open-apis/' + path.lstrip('/')
    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json; charset=utf-8'}
    body = json.dumps(data, ensure_ascii=False).encode() if data is not None else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    with urllib.request.urlopen(req, timeout=30) as resp:
        result = json.loads(resp.read())
    if result.get('code') != 0:
        raise RuntimeError(result.get('msg') or f'飞书接口失败：{result.get("code")}')
    return result


def feishu_record_path(record_id=''):
    base = f'bitable/v1/apps/{FEISHU_BITABLE_APP_TOKEN}/tables/{FEISHU_BITABLE_TABLE_ID}/records'
    return f'{base}/{record_id}' if record_id else base


def feishu_sale_item_fields(sale, item):
    amount = as_float(item.get('quantity')) * as_float(item.get('price'))
    return {
        '订单号': sale.get('no') or '',
        '同步状态': '待确认',
        '客户': sale.get('customer') or '',
        '日期': feishu_date_millis(sale.get('date')),
        '商品名称': item.get('product_name') or '',
        '数量': clean_number(item.get('quantity'), 4),
        '单位': item.get('product_unit') or '',
        '单价': clean_number(item.get('price'), 4),
        '金额': clean_number(amount, 4),
        '备注': item.get('note') or '',
        '系统销售单ID': sale.get('id') or '',
        '系统明细ID': item.get('id') or '',
        '商品ID': item.get('product_id') or '',
        '同步错误': '',
    }


def push_sale_to_feishu(sale_id):
    if not FEISHU_SYNC_ENABLED or not feishu_configured():
        return None
    conn = db._get_conn()
    try:
        sale_row = conn.execute('SELECT * FROM sales WHERE id = ?', (sale_id,)).fetchone()
        if not sale_row:
            return '销售单不存在'
        sale = dict(sale_row)
        old_records = conn.execute('SELECT record_id FROM feishu_sync_records WHERE sale_id = ?', (sale_id,)).fetchall()
        for row in old_records:
            try:
                feishu_api('DELETE', feishu_record_path(row['record_id']))
            except Exception:
                pass
        conn.execute('DELETE FROM feishu_sync_records WHERE sale_id = ?', (sale_id,))

        items = conn.execute('SELECT * FROM sales_items WHERE sale_id = ? ORDER BY rowid', (sale_id,)).fetchall()
        for item_row in items:
            item = dict(item_row)
            result = feishu_api('POST', feishu_record_path(), {'fields': feishu_sale_item_fields(sale, item)})
            record = result.get('data', {}).get('record', {})
            record_id = record.get('record_id') or record.get('id')
            if record_id:
                conn.execute('''
                    INSERT OR REPLACE INTO feishu_sync_records
                        (id, sale_id, sale_item_id, record_id, status, last_error, synced_at, created_at)
                    VALUES (?, ?, ?, ?, '待确认', '', ?, ?)
                ''', (generate_id(), sale_id, item['id'], record_id, now_str(), now_str()))
        conn.commit()
        return None
    except Exception as e:
        conn.rollback()
        return str(e)
    finally:
        conn.close()


def safe_push_sale_to_feishu(sale_id):
    try:
        return push_sale_to_feishu(sale_id)
    except Exception as e:
        return str(e)


def push_sales_to_feishu_async(sale_ids):
    ids = [sale_id for sale_id in (sale_ids or []) if sale_id]
    if not ids or not FEISHU_SYNC_ENABLED:
        return

    def worker():
        for sale_id in ids:
            safe_push_sale_to_feishu(sale_id)

    threading.Thread(target=worker, daemon=True).start()


def feishu_update_record(record_id, fields):
    return feishu_api('PUT', feishu_record_path(record_id), {'fields': fields})


def list_feishu_records():
    records = []
    page_token = ''
    while True:
        query = 'page_size=500'
        if page_token:
            query += '&page_token=' + page_token
        result = feishu_api('GET', feishu_record_path() + '?' + query)
        data = result.get('data') or {}
        records.extend(data.get('items') or [])
        if not data.get('has_more'):
            break
        page_token = data.get('page_token') or ''
        if not page_token:
            break
    return records


def sync_confirmed_feishu_records():
    if not FEISHU_SYNC_ENABLED or not feishu_configured():
        return {'synced': 0, 'failed': 0, 'message': '飞书同步未启用'}
    records = list_feishu_records()
    synced = 0
    failed = 0
    for record in records:
        record_id = record.get('record_id') or record.get('id')
        fields = record.get('fields') or {}
        status = feishu_field_value(fields.get('同步状态'))
        if status != '已确认':
            continue
        try:
            sale_id = feishu_field_value(fields.get('系统销售单ID'))
            sale_item_id = feishu_field_value(fields.get('系统明细ID'))
            product_name = normalize_hotel_item_name(feishu_field_value(fields.get('商品名称')))
            unit = feishu_field_value(fields.get('单位'))
            quantity = as_float(fields.get('数量'))
            price = as_float(fields.get('单价'))
            note = feishu_field_value(fields.get('备注'))
            if not sale_id or not sale_item_id:
                raise ValueError('缺少系统销售单ID或系统明细ID')
            if not product_name or quantity <= 0:
                raise ValueError('商品名称或数量无效')

            conn = db._get_conn()
            try:
                sale = conn.execute('SELECT * FROM sales WHERE id = ?', (sale_id,)).fetchone()
                sale_item = conn.execute('SELECT * FROM sales_items WHERE id = ? AND sale_id = ?', (sale_item_id, sale_id)).fetchone()
                if not sale or not sale_item:
                    raise ValueError('系统销售单或明细不存在，可能已重新生成')
                product = find_product_with_conn(conn, product_name, unit=unit)
                if not product:
                    product = create_hotel_flow_product(conn, product_name, '', unit or '件', price)
                subtotal = quantity * price
                conn.execute('''
                    UPDATE sales_items
                    SET product_id = ?, product_name = ?, product_spec = ?, product_unit = ?,
                        quantity = ?, price = ?, subtotal = ?, note = ?
                    WHERE id = ?
                ''', (
                    product['id'], product['name'], product.get('spec') or '', unit or product.get('unit') or '',
                    quantity, price, subtotal, note, sale_item_id
                ))
                total_row = conn.execute('SELECT COALESCE(SUM(subtotal), 0) AS total FROM sales_items WHERE sale_id = ?', (sale_id,)).fetchone()
                conn.execute('UPDATE sales SET total = ? WHERE id = ?', (as_float(total_row['total'] if total_row else 0), sale_id))
                if price > 0:
                    conn.execute('UPDATE products SET last_sale_price = ? WHERE id = ?', (price, product['id']))
                    upsert_customer_product_price(
                        conn,
                        sale['customer'],
                        product,
                        unit or product.get('unit') or '',
                        price,
                        source='feishu_confirm'
                    )
                conn.execute('''
                    UPDATE feishu_sync_records
                    SET status = '已同步', last_error = '', synced_at = ?
                    WHERE record_id = ?
                ''', (now_str(), record_id))
                conn.commit()
            except Exception:
                conn.rollback()
                raise
            finally:
                conn.close()

            feishu_update_record(record_id, {
                '同步状态': '已同步',
                '金额': clean_number(quantity * price, 4),
                '同步时间': int(time.time() * 1000),
                '同步错误': '',
            })
            synced += 1
        except Exception as e:
            failed += 1
            try:
                feishu_update_record(record_id, {'同步状态': '同步失败', '同步错误': str(e)[:500]})
            except Exception:
                pass
    return {'synced': synced, 'failed': failed}


def feishu_sync_worker():
    time.sleep(8)
    while True:
        try:
            sync_confirmed_feishu_records()
        except Exception as e:
            print(f'[feishu-sync] {e}')
        time.sleep(max(FEISHU_SYNC_INTERVAL_SECONDS, 15))


def start_feishu_sync_thread():
    global _feishu_sync_thread_started
    if _feishu_sync_thread_started or not FEISHU_SYNC_ENABLED or not feishu_configured():
        return
    _feishu_sync_thread_started = True
    thread = threading.Thread(target=feishu_sync_worker, daemon=True)
    thread.start()


def format_date_dot(value):
    text = normalize_date(value)
    try:
        dt = datetime.strptime(text, '%Y-%m-%d')
        return f'{dt.year}.{dt.month}.{dt.day}'
    except ValueError:
        return str(value or '')


def chinese_integer_upper(number):
    number = int(number)
    cn_nums = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    cn_units = ['', '拾', '佰', '仟']
    cn_big_units = ['', '万', '亿']
    if number == 0:
        return '零'
    result = ''
    unit_pos = 0
    zero = False
    for char in reversed(str(number)):
        digit = int(char)
        if digit == 0:
            if not zero and result:
                result = cn_nums[0] + result
                zero = True
        else:
            result = cn_nums[digit] + cn_units[unit_pos % 4] + result
            zero = False
        unit_pos += 1
        if unit_pos % 4 == 0 and unit_pos // 4 < len(cn_big_units) and unit_pos < len(str(number)):
            result = cn_big_units[unit_pos // 4] + result
    return re.sub(r'零+', '零', result).rstrip('零') or '零'


def rmb_upper(amount):
    cents = int(round(as_float(amount) * 100))
    yuan = cents // 100
    jiao = (cents % 100) // 10
    fen = cents % 10
    cn_nums = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    text = chinese_integer_upper(yuan) + '元'
    if jiao == 0 and fen == 0:
        return text + '整'
    if jiao:
        text += cn_nums[jiao] + '角'
    elif fen:
        text += '零'
    if fen:
        text += cn_nums[fen] + '分'
    else:
        text += '整'
    return text


FLOW_ITEM_SQL = "(code LIKE 'WP%' OR code LIKE 'IMP%')"
HOTEL_FLOW_PURCHASE_ID = '__HOTEL_FLOW__'

HOTEL_FRESH_KEYWORDS = [
    '菜', '豆芽', '莲花白', '花白', '白菜', '青菜', '菠菜', '芹', '麦芹',
    '蒜苔', '豆王', '豇豆', '黄瓜', '西红柿', '番茄', '青椒', '红椒',
    '土豆', '红薯', '胡萝卜', '西葫芦', '南瓜', '冬瓜', '玉米', '莲菜', '莲藕',
    '洋葱', '茄子', '油菜', '生菜', '娃娃菜', '茼蒿', '笋', '平菇',
    '香菇', '蘑菇', '金针菇', '杏鲍菇', '大葱', '小葱', '西兰花',
    '豆腐', '豆干', '豆皮', '豆腐皮', '水果',
    '圣女果', '哈密瓜', '西瓜', '火龙果', '柠檬', '油桃', '苹果', '香蕉', '梨', '橙', '肉片', '肉丝', '肉沫', '鲜肉',
    '猪肉', '牛肉', '羊肉', '鸡肉', '鸡蛋', '鸭蛋', '咸鸭蛋'
]


def is_hotel_fresh_item_name(name):
    text = str(name or '')
    return any(keyword in text for keyword in HOTEL_FRESH_KEYWORDS)


HOTEL_FRESH_SQL = ' OR '.join(
    "si.product_name LIKE '%" + keyword.replace("'", "''") + "%'"
    for keyword in HOTEL_FRESH_KEYWORDS
)
EFFECTIVE_SALE_QUANTITY_SQL = """
    CASE
        WHEN COALESCE(si.quantity, 0) > 0 THEN si.quantity
        WHEN COALESCE(si.price, 0) > 0 AND COALESCE(si.subtotal, 0) > 0 THEN si.subtotal / si.price
        ELSE 0
    END
"""

HOTEL_COST_SQL = f"""
    CASE
        WHEN {HOTEL_FRESH_SQL}
        THEN si.subtotal
        ELSE ({EFFECTIVE_SALE_QUANTITY_SQL}) * COALESCE(p.cost, 0)
    END
"""


def rows_to_dicts(rows):
    return [dict(row) for row in rows]


def query_with_conn(conn, sql, params=()):
    return rows_to_dicts(conn.execute(sql, params).fetchall())


def query_one_with_conn(conn, sql, params=()):
    row = conn.execute(sql, params).fetchone()
    return dict(row) if row else None


def attach_items_to_orders(orders, item_table, order_fk):
    """一次性读取单据明细，避免每张单据再查一次数据库。"""
    if not orders:
        return orders

    order_ids = [order['id'] for order in orders]
    placeholders = ','.join('?' for _ in order_ids)
    rows = db.query(f'''
        SELECT
            i.id,
            i.{order_fk},
            i.product_id,
            COALESCE(NULLIF(i.product_name, ''), p.name, '') AS product_name,
            COALESCE(NULLIF(i.product_spec, ''), p.spec, '') AS product_spec,
            COALESCE(NULLIF(i.product_unit, ''), p.unit, '') AS product_unit,
            i.quantity,
            i.price,
            i.subtotal,
            COALESCE(i.note, '') AS note
        FROM {item_table} i
        LEFT JOIN products p ON i.product_id = p.id
        WHERE i.{order_fk} IN ({placeholders})
        ORDER BY i.rowid
    ''', tuple(order_ids))

    grouped = {order_id: [] for order_id in order_ids}
    for item in rows:
        grouped.setdefault(item[order_fk], []).append(item)
    for order in orders:
        order['items'] = grouped.get(order['id'], [])
    return orders


CN_NUMBERS = {
    '零': 0, '〇': 0, '一': 1, '二': 2, '两': 2, '三': 3, '四': 4,
    '五': 5, '六': 6, '七': 7, '八': 8, '九': 9, '俩': 2
}
SALE_UNITS = [
    '公斤', '千克', '毫升', '斤', '克', 'kg', 'KG', 'g', 'G', '升', 'ml', 'ML', 'L',
    '盘', '桶', '板', '件', '个', '只', '包', '袋', '箱', '瓶', '盒', '把', '根',
    '条', '份', '卷', '块', '张', '套', '筐', '盆', '节', '双', '颗', '提', '捆'
]

HOTEL_CATEGORY_ORDER = [
    '蔬菜类', '豆制品类', '水果类', '肉蛋类', '主食面点类',
    '粮油干货类', '调料咸菜类', '饮品乳品类', '冻货类', '饼干', '其他类'
]
HOTEL_CATEGORY_KEYWORDS = {
    '蔬菜类': [
        '大芹菜', '毛芹', '芹菜', '西葫芦', '菜花', '花菜', '豆王', '土豆', '青菜（把）', '青菜', '青椒', '香葱', '香菜',
        '西红柿', '西兰花', '菠菜', '小葱', '香菇', '花白', '胡萝卜', '白菜', '奶白菜',
        '红椒', '蒜苔', '蒜薹', '荷兰豆', '广红', '韭黄', '白玉菇', '麦芹',
        '红薯', '水果玉米', '螺丝椒', '青红线椒', '红线椒', '绿线椒', '红辣椒', '青辣椒', '鲜小米椒', '小米椒', '黄甜椒', '甜豆', '茄子', '黄瓜',
        '油麦菜', '油菜', '茼蒿', '龙须菜', '生菜', '娃娃菜', '莲菜', '莲藕', '冬瓜', '南瓜', '韭菜', '大葱',
        '洋葱', '蒜苗', '上海青', '包菜', '蘑菇', '金针菇', '平菇', '杏鲍菇',
        '波菜', '生姜', '姜', '白萝卜', '红心萝卜', '樱桃小萝卜', '萝卜苗', '象牙萝卜', '象牙白萝卜', '花心萝卜', '净笋', '线椒', '黄芽菜', '广东菜心', '菜心', '净蒜', '去皮蒜', '扒皮蒜', '新蒜', '毛蒜', '白葱',
        '苦菊', '西蓝花', '青笋', '圆生菜', '罗马生菜', '叶生菜', '红萝卜', '圆茄',
        '水洗铁棍山药', '铁棍山药', '山药', '黄土豆', '贝贝南瓜', '长豆角', '芥菜', '广茄',
        '园茄子', '圆茄子', '白豆角', '扁豆角', '龙豆', '长江豆', '长豇豆', '豇豆', '紫甘蓝', '豆苗', '磨菇', '海鲜菇', '藕片', '苦瓜', '西芹',
    ],
    '豆制品类': ['玉米搅团', '凉粉', '牛皮干', '红豆腐丝', '黑豆腐丝', '素毛肚丝', '羊肚丝', '黑豆皮', '鱼酸菜', '酸豆角', '方块面筋', '方块儿面筋', '小面筋', '面筋', '酸菜(黑)', '酸菜（黑）', '豆芽菜', '大豆芽', '小豆芽', '黄豆芽', '豆芽', '豆腐干', '豆腐皮', '豆腐', '豆干', '豆皮', '腐竹', '云丝'],
    '水果类': ['乳瓜', '圣女果', '哈密瓜', '西瓜', '火龙果', '柠檬', '油桃', '小油桃', '桃子', '苹果', '香梨', '香蕉', '青提', '李子', '梨', '橙', '水果'],
    '肉蛋类': ['牛柳片', '五花肉块', '前腿肉块', '猪耳朵', '腊肠', '排骨', '鸡脯肉', '鸡胸肉', '鸡肉丁', '里脊肉片', '后腿肉片', '后腿肉', '五花肉', '五花肉片', '黑鱼片', '肉丝', '肉片', '肉块', '肉沫', '肉末', '鸡蛋', '咸鸭蛋', '鸡腿', '鸡肉', '猪肉', '牛肉', '羊肉', '鸭血', '鱼'],
    '主食面点类': ['黑吐司面包', '白吐司面包', '吐司面包', '荷叶饼', '龙须面', '锅盔', '法棍', '凉面', '炒拉条', '米皮', '擀面皮', '炒饼丝', '炒饼', '炒细面', '麻什', '麻食', '发糕', '蛋黄酥', '小油条', '牛角包', '土豆粉', '细薄韭叶面', '手工面条', '手工面', '面粉', '炒面', '凉皮', '白凉皮', '馒头', '馍', '小笼包', '哨子面', '臊子面', '扯面', '面条', '油条', '包子', '饺子', '米线', '干米线', '湿米线', '河粉'],
    '粮油干货类': ['玉米粒', '淀粉', '核桃仁', '瓜子仁', '花生米', '黄油', '果酱', '沙拉油醋汁', '白糖', '金罗高汤', '辣椒面', '红薯粉条', '小米', '绿豆', '红豆', '大米', '江米', '黄金豆', '大豆油', '菜籽油', '食用油', '油', '粉条', '粉丝', '小木耳', '木耳', '黑木耳', '海带', '银耳'],
    '调料咸菜类': ['六六红火锅料', '火锅料', '精品魔芋', '素鸡', '泡椒', '东古一品鲜酱油', '蒸鱼豉油', '辣鲜露', '鸡汁', '小米辣', '老干妈', '甜面酱', '番茄酱', '蕃茄酱', '卤肉料包', '麻婆豆腐调料', '麻婆豆腐料', '火锅底料', '十三香', '杏仁片', '白芝麻', '黄豆', '虾皮', '红99', '桂皮', '花椒', '白胡椒粉', '白胡椒', '干辣椒', '辣椒段', '辣皮子', '裙带菜咸菜', '酱黄瓜', '咸菜', '调料', '生抽', '老抽', '料酒', '红醋', '蚝油', '耗油', '食盐', '盐', '醋', '酱油', '味精', '鸡精'],
    '饮品乳品类': ['果溢多橙汁', '果溢多桃汁', '果溢多橙子', '果溢多猕猴桃汁', '冰红茶', '冰峰罐装', '冰峰瓶装', '瓶装冰峰', '伊利益消益生菌优酪酸奶', '伊利风味发酵乳复原乳酸奶', '益生菌发酵酸奶', '纯牛奶', '蓝莓酱', '苹果酱', '冰峰', '酸梅汤', '酸奶', '橙汁', '牛奶', '饮料', '果汁', '矿泉水'],
    '冻货类': ['玉米红枣糕', '玉米发糕', '花卷', '火锅杂类丸子蟹排', '火锅油条', '荷叶饼', '小油条', '牛柳条', '蟹排', '冻', '冷冻', '冻玉米', '冻货', '冻品', '火锅丸子', '馄饨', '丸子', '素包', '鸡块', '什锦菜', '猪头肉', '里昂火腿', '王中王火腿'],
    '饼干': ['黑全麦面包', '原味面包', '饼干', '曲奇'],
}
HOTEL_CATEGORY_OVERRIDES = {
    '调料咸菜类': ['六六红火锅料', '火锅料', '精品魔芋', '素鸡', '泡椒', '东古一品鲜酱油', '蒸鱼豉油', '辣鲜露', '鸡汁', '小米辣', '老干妈', '甜面酱', '番茄酱', '蕃茄酱', '卤肉料包', '麻婆豆腐调料', '麻婆豆腐料', '火锅底料', '酱黄瓜', '裙带菜咸菜', '咸菜', '十三香', '杏仁片', '白芝麻', '黄豆', '虾皮', '红99', '桂皮', '花椒', '白胡椒粉', '白胡椒', '干辣椒', '辣椒段', '辣皮子', '红醋', '生抽', '老抽', '料酒', '蚝油', '耗油', '食盐'],
    '豆制品类': ['玉米搅团', '凉粉', '红豆腐丝', '黑豆腐丝', '素毛肚丝', '羊肚丝', '黑豆皮', '鱼酸菜', '酸豆角', '方块面筋', '方块儿面筋', '小面筋', '面筋', '酸菜(黑)', '酸菜（黑）', '豆芽菜', '大豆芽', '小豆芽', '黄豆芽', '豆芽', '豆腐干', '豆腐皮', '豆腐', '豆干', '豆皮', '腐竹', '云丝'],
    '水果类': ['乳瓜', '圣女果', '哈密瓜', '西瓜', '火龙果', '柠檬', '油桃', '小油桃', '桃子', '苹果', '香梨', '小香蕉', '香蕉', '青提', '李子'],
    '蔬菜类': ['水洗铁棍山药', '铁棍山药', '山药', '黄土豆', '贝贝南瓜', '水果玉米', '蒜苔', '蒜薹', '荷兰豆', '广红', '韭黄', '白玉菇', '麦芹', '大芹菜', '毛芹', '红薯', '螺丝椒', '青红线椒', '红线椒', '绿线椒', '红辣椒', '青辣椒', '鲜小米椒', '小米椒', '黄甜椒', '甜豆', '奶白菜', '白豆角', '苦瓜', '油麦菜', '油菜', '香葱', '龙须菜', '莲菜', '连菜', '波菜', '生姜', '姜', '白萝卜', '红心萝卜', '樱桃小萝卜', '萝卜苗', '象牙萝卜', '象牙白萝卜', '花心萝卜', '净笋', '线椒', '黄芽菜', '广东菜心', '菜心', '净蒜', '去皮蒜', '扒皮蒜', '新蒜', '毛蒜', '白葱', '苦菊', '西蓝花', '青笋', '圆生菜', '罗马生菜', '叶生菜', '红萝卜', '园茄子', '圆茄子', '圆茄', '广茄', '长豆角', '扁豆角', '龙豆', '长江豆', '长豇豆', '豇豆', '芥菜', '紫甘蓝', '豆苗', '磨菇', '海鲜菇', '藕片', '西芹'],
    '肉蛋类': ['牛柳片', '五花肉块', '前腿肉块', '猪耳朵', '腊肠', '排骨', '鸡脯肉', '鸡肉丁', '里脊肉片', '后腿肉片', '后腿肉', '五花肉', '五花肉片', '黑鱼片', '肉块'],
    '主食面点类': ['黑吐司面包', '白吐司面包', '吐司面包', '龙须面', '锅盔', '法棍', '凉面', '炒拉条', '米皮', '擀面皮', '炒饼丝', '炒饼', '炒细面', '麻什', '发糕', '蛋黄酥', '牛角包', '土豆粉', '细薄韭叶面', '手工面条', '手工面', '白凉皮', '哨子面', '馍'],
    '粮油干货类': ['玉米粒', '香油', '油条', '淀粉', '核桃仁', '瓜子仁', '花生米', '黄油', '果酱', '沙拉油醋汁', '白糖', '金罗高汤', '辣椒面', '红薯粉条', '粉丝', '小木耳', '黑木耳', '绿豆', '江米', '油'],
    '饮品乳品类': ['果溢多橙汁', '果溢多桃汁', '果溢多橙子', '果溢多猕猴桃汁', '冰红茶', '冰峰罐装', '冰峰瓶装', '瓶装冰峰', '伊利益消益生菌优酪酸奶', '伊利风味发酵乳复原乳酸奶', '益生菌发酵酸奶', '风味发酵酸奶', '纯牛奶', '蓝莓酱', '苹果酱', '冰峰', '酸梅汤'],
    '冻货类': ['玉米红枣糕', '玉米发糕', '花卷', '火锅杂类丸子蟹排', '火锅油条', '荷叶饼', '小油条', '牛柳条', '蟹排', '冻', '冷冻', '冻货', '冻品', '火锅丸子', '馄饨', '丸子', '素包', '鸡块', '什锦菜', '猪头肉', '里昂火腿', '王中王火腿'],
    '饼干': ['黑全麦面包', '原味面包', '黑色小饼干', '恋妮曲奇小饼干', '饼干', '曲奇'],
    '其他类': ['六六红', '水晶粉', '火锅宽粉', '火锅邵皮', '火锅苕皮', '芝麻酱', '贡菜', '针金菇', '金针菇', '青岛纯生', '鞭炮笋', '香铃卷', '鲜玉米', '鲜虾', '榨菜丁', '榨菜丝', '蒜米', '泡椒', '冰糖', '打包袋', '豆瓣酱'],
}
HOTEL_CATEGORY_EXACT_OVERRIDES = {
    '六六红': '其他类',
    '水晶粉': '其他类',
    '火锅宽粉': '其他类',
    '火锅邵皮': '其他类',
    '火锅苕皮': '其他类',
    '芝麻酱': '其他类',
    '贡菜': '其他类',
    '针金菇': '其他类',
    '金针菇': '其他类',
    '青岛纯生': '其他类',
    '鞭炮笋': '其他类',
    '香铃卷': '其他类',
    '鲜玉米': '其他类',
    '鲜虾': '其他类',
    '海带丝': '豆制品类',
    '干米线': '豆制品类',
    '湿米线': '豆制品类',
    '牛皮干': '豆制品类',
    '荷叶饼': '冻货类',
    '香葱饼干': '饼干',
    '蓝莓酱': '饮品乳品类',
    '苹果酱': '饮品乳品类',
    '白糖': '调料咸菜类',
    '奶油沙拉酱': '粮油干货类',
    '干木耳': '粮油干货类',
    '核桃仁': '其他类',
    '瓜子仁': '其他类',
    '腰果': '其他类',
    '包谷珍': '其他类',
    '粗辣子面': '其他类',
}
PUBLIC_PURCHASE_TYPE_ORDER = [
    '蔬菜', '水果', '肉', '蛋', '奶', '面食', '冻货', '豆制品',
    '粮油', '调料', '饮品', '饼干', '其他'
]
HOTEL_NAME_ALIASES = {
    '红九九': '红99',
    '红九十九': '红99',
    '红玖玖': '红99',
    '鸡盘': '鸡蛋',
    '番茄': '西红柿',
    '西红市': '西红柿',
    '西虹市': '西红柿',
    '西红柿子': '西红柿',
    '蕃茄酱': '番茄酱',
    '蒜薹': '蒜苔',
    '蒜台': '蒜苔',
    '蒜苔子': '蒜苔',
    '波菜': '菠菜',
    '菠莱': '菠菜',
    '小青菜': '青菜（把）',
    '青莱': '青菜',
    '青菜': '青菜（把）',
    '青菜（把）': '青菜（把）',
    '黄爪': '黄瓜',
    '针金菇': '金针菇',
    '黄甜辣': '黄甜椒',
    '磨菇': '蘑菇',
    '小香焦': '小香蕉',
    '香焦': '香蕉',
    '圣女果粉贝贝': '圣女果',
    '包菜': '包菜',
    '莲花白': '花白',
    '莲莲花白': '花白',
    '连菜': '莲菜',
    '广红': '红萝卜',
    '胡萝卜': '红萝卜',
    '红萝卜': '红萝卜',
    '红辣椒': '红椒',
    '有机菜花': '菜花',
    '花菜': '菜花',
    '菜花': '菜花',
    '黄心土豆': '黄土豆',
    '毛芹': '麦芹',
    '麦芹': '麦芹',
    '四季豆': '豆王',
    '长江豆': '长豇豆',
    '长豆角': '长豇豆',
    '江豆': '豇豆',
    '平茹': '平菇',
    '平姑': '平菇',
    '岐身': '岐山',
    '细薄非叶面': '细薄韭叶面',
    '细薄韭叶面': '细薄韭叶面',
    '韭叶面': '细薄韭叶面',
    '饼丝': '炒饼丝',
    '麻婆豆腐料': '麻婆豆腐调料',
    '白色凉皮': '白凉皮',
    '鲜海带丝': '海带丝',
    '米线': '干米线',
    '湿米线': '湿米线',
    '鲜米线': '湿米线',
    '口罩': '一次性口罩',
    '抽纸': '餐巾抽纸',
    '立昂火腿': '里昂火腿',
    '精品魔芋': '素毛肚丝',
    '黑豆腐丝': '红豆腐丝',
    '方块儿面筋': '方块面筋',
    '扒皮青笋': '净笋',
    '象牙萝卜': '白萝卜',
    '香茹': '香菇',
    '金针蘑': '金针菇',
    '发孝粉': '发酵粉',
}


def parse_quantity(value):
    text = str(value or '').strip()
    if not text:
        return 0
    try:
        return float(text)
    except ValueError:
        pass
    if text == '半':
        return 0.5
    if text.endswith('半'):
        return parse_quantity(text[:-1]) + 0.5
    if text in CN_NUMBERS:
        return float(CN_NUMBERS[text])
    if text == '十':
        return 10
    if '十' in text:
        left, right = text.split('十', 1)
        tens = CN_NUMBERS.get(left, 1) if left else 1
        ones = CN_NUMBERS.get(right, 0) if right else 0
        return float(tens * 10 + ones)
    total = 0
    for char in text:
        if char in CN_NUMBERS:
            total = total * 10 + CN_NUMBERS[char]
    return float(total)


def find_product(product_name, product_code='', spec='', unit=''):
    product = None
    if product_code:
        product = db.query_one('SELECT * FROM products WHERE code = ?', (product_code,))
    if not product and product_name:
        product = db.query_one('''
            SELECT * FROM products
            WHERE name = ?
              AND (? = '' OR COALESCE(spec, '') = ?)
              AND (? = '' OR unit = ?)
            LIMIT 1
        ''', (product_name, spec, spec, unit, unit))
    if not product and product_name:
        product = db.query_one('SELECT * FROM products WHERE name = ? LIMIT 1', (product_name,))
    if not product and product_name:
        product = db.query_one('SELECT * FROM products WHERE name LIKE ? LIMIT 1', (f'%{product_name}%',))
    return product


def create_hotel_flow_product(conn, name, spec='', unit='件', price=0):
    product_id = generate_id()
    code = generate_no('HT')
    while conn.execute('SELECT 1 FROM products WHERE code = ?', (code,)).fetchone():
        code = generate_no('HT')
    initial_cost = as_float(price) if is_hotel_fresh_item_name(name) else as_float(price)
    conn.execute('''
        INSERT INTO products (id, code, name, spec, unit, cost, stock, alert_line, departments, last_sale_price, created_at)
        VALUES (?, ?, ?, ?, ?, ?, 0, 0, '[]', ?, ?)
    ''', (product_id, code, name, spec or '', unit or '件', initial_cost, as_float(price), now_str()))
    return dict(conn.execute('SELECT * FROM products WHERE id = ?', (product_id,)).fetchone())


def parse_hotel_sales_text(text):
    units_pattern = '|'.join(SALE_UNITS)
    number_pattern = r'\d+(?:\.\d+)?|[零〇一二两俩三四五六七八九十百半]+'
    normalized = re.sub(r'[，,。；;、\n\r\t]+', '，', str(text or ''))
    normalized = re.sub(r'\s+', '', normalized)
    normalized = normalized.replace('岐身', '岐山')
    normalized = re.sub(
        rf'(?P<name>[^，\d零〇一二两三四五六七八九十百半]+)，(?P<num>{number_pattern})(?P<unit>{units_pattern})(?=，|$)',
        r'\g<name>\g<num>\g<unit>',
        normalized
    )

    items = []
    used_spans = []

    prefix_pattern = re.compile(
        rf'(?:(?<=^)|(?<=，))(?P<num>{number_pattern})(?P<unit>{units_pattern})(?P<name>[^，]+?)(?=(?:零|〇|一|二|两|三|四|五|六|七|八|九|十|百|半|，|$))'
    )
    suffix_pattern = re.compile(
        rf'(?P<name>[^，\d零〇一二两三四五六七八九十百半]+?)(?P<num>{number_pattern})(?P<unit>{units_pattern})'
    )

    for pattern, mode in ((prefix_pattern, 'prefix'), (suffix_pattern, 'suffix')):
        for match in pattern.finditer(normalized):
            start, end = match.span()
            if any(not (end <= s or start >= e) for s, e in used_spans):
                continue
            name = match.group('name').strip('，, ')
            quantity = parse_quantity(match.group('num'))
            unit = match.group('unit')
            if name and quantity > 0:
                items.append({'name': name, 'quantity': quantity, 'unit': unit, 'mode': mode})
                used_spans.append((start, end))

    items.sort(key=lambda item: normalized.find(item['name']))
    leftovers = normalized
    for start, end in sorted(used_spans, reverse=True):
        leftovers = leftovers[:start] + '，' + leftovers[end:]
    leftovers = [part for part in leftovers.split('，') if part]
    return items, leftovers


def normalize_hotel_item_name(name):
    name = str(name or '').strip()
    return HOTEL_NAME_ALIASES.get(name, name)


def normalize_hotel_match_name(name):
    return re.sub(r'\s+', '', normalize_hotel_item_name(name or ''))


def normalize_hotel_text_aliases(text):
    normalized = str(text or '')
    for alias in sorted(HOTEL_NAME_ALIASES, key=len, reverse=True):
        if alias in ('米线', '饼丝'):
            continue
        normalized = normalized.replace(alias, HOTEL_NAME_ALIASES[alias])
    normalized = normalized.replace('西红柿酱', '番茄酱')
    normalized = normalized.replace('餐巾餐巾抽纸', '餐巾抽纸')
    normalized = re.sub(r'(?:莲)+花白', '花白', normalized)
    normalized = normalized.replace('青菜（把）（把）', '青菜（把）')
    normalized = re.sub(r'(?:细薄)+韭叶面', '细薄韭叶面', normalized)
    normalized = re.sub(r'(?:红萝卜){2,}', '红萝卜', normalized)
    return normalized


def normalize_hotel_item_quantity_unit(name, quantity, unit):
    item_name = normalize_hotel_item_name(re.sub(r'各$', '', str(name or '').strip()))
    item_quantity = as_float(quantity)
    item_unit = str(unit or '').strip()
    if item_name == '鸡蛋' and item_unit == '提':
        item_quantity *= 5
        item_unit = '盘'
    if item_name == '方块面筋' and item_unit == '个' and item_quantity >= 10 and item_quantity % 10 == 0:
        item_quantity = item_quantity / 10
        item_unit = '包'
    return item_name, item_quantity, item_unit


def ark_smart_parse_hotel_grocery_text(text):
    if not ARK_API_KEY or not ARK_MODEL:
        return [], '方舟 API 未配置'
    ok, budget_error = ai_budget_available(AI_TEXT_FALLBACK_COST_CNY)
    if not ok:
        return [], budget_error

    prompt = f"""
你是酒店食材进销存清单识别助手。请从下面的中文清单片段中提取物品、数量、单位，并规范品名。

要求：
1. 只输出 JSON，不要解释。
2. JSON 格式为 {{"items":[{{"name":"品名","quantity":数字,"unit":"单位"}}]}}。
3. 单位必须尽量使用原文单位，例如斤、袋、件、个、盒、盘、桶、瓶、箱；鸡蛋 1 提 = 5 盘，鸡蛋按盘输出。
4. 品名不要带数量和单位。
5. 常见别名规范：广红/胡萝卜/红萝卜=红萝卜，菜花/花菜/有机菜花=菜花，毛芹=麦芹，黑豆腐丝=红豆腐丝，方块儿面筋=方块面筋，精品魔芋=素毛肚丝，细薄非叶面/韭叶面=细薄韭叶面，麻婆豆腐料=麻婆豆腐调料，白色凉皮=白凉皮，口罩=一次性口罩。
6. 如果无法确定，不要编造，跳过该项。

清单片段：
{text}
""".strip()

    payload = {
        'model': ARK_MODEL,
        'messages': [
            {'role': 'system', 'content': '你只返回严格 JSON。'},
            {'role': 'user', 'content': prompt},
        ],
        'temperature': 0.1,
        'max_tokens': 1200,
    }
    if AI_RESPONSE_FORMAT_JSON:
        payload['response_format'] = {'type': 'json_object'}
    req = urllib.request.Request(
        f'{ARK_BASE_URL}/chat/completions',
        data=json.dumps(payload, ensure_ascii=False).encode('utf-8'),
        headers={
            'Authorization': f'Bearer {ARK_API_KEY}',
            'Content-Type': 'application/json',
        },
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            raw = resp.read().decode('utf-8')
        data = json.loads(raw)
        estimated_cost, prompt_tokens, completion_tokens = estimate_ai_cost_from_usage(data.get('usage'), AI_TEXT_FALLBACK_COST_CNY)
        record_ai_usage(ai_provider_name(), ARK_MODEL, 'text_parse', estimated_cost, prompt_tokens, completion_tokens)
        content = data['choices'][0]['message']['content']
        parsed = json.loads(extract_json_object_text(content))
    except urllib.error.HTTPError as e:
        try:
            detail = e.read().decode('utf-8')[:300]
        except Exception:
            detail = str(e)
        return [], f'方舟 API 调用失败：{detail}'
    except Exception as e:
        return [], f'方舟 API 解析失败：{e}'

    items = []
    for row in parsed.get('items', []):
        name, quantity, unit = normalize_hotel_item_quantity_unit(
            row.get('name'),
            row.get('quantity'),
            row.get('unit')
        )
        if name and unit and quantity > 0:
            items.append({'name': name, 'quantity': quantity, 'unit': unit, 'smart': True})
    return items, ''


def extract_json_object_text(content):
    text = str(content or '').strip()
    if text.startswith('```'):
        text = re.sub(r'^```(?:json)?\s*', '', text, flags=re.I)
        text = re.sub(r'\s*```$', '', text).strip()
    if text.startswith('{') and text.endswith('}'):
        return text
    start = text.find('{')
    end = text.rfind('}')
    if start >= 0 and end > start:
        return text[start:end + 1]
    return text


def default_hotel_unit(name):
    category = hotel_item_category(name)
    if category in ('蔬菜类', '肉蛋类'):
        return '斤'
    return ''


def hotel_item_category(name):
    exact_name = normalize_hotel_item_name(re.sub(r'\s+', '', str(name or '').strip()))
    if exact_name in HOTEL_CATEGORY_EXACT_OVERRIDES:
        return HOTEL_CATEGORY_EXACT_OVERRIDES[exact_name]
    override_matches = [
        (category, keyword)
        for category, keywords in HOTEL_CATEGORY_OVERRIDES.items()
        for keyword in keywords
    ]
    for category, keyword in sorted(override_matches, key=lambda pair: len(pair[1]), reverse=True):
        if keyword in name:
            return category
    for category in HOTEL_CATEGORY_ORDER:
        if any(keyword in name for keyword in HOTEL_CATEGORY_KEYWORDS.get(category, [])):
            return category
    return '其他类'


def purchase_category_label(category):
    if category in ('粮油干货类', '调料咸菜类'):
        return '粮油调料类'
    if category == '冻货类':
        return '冻货'
    return category


PURCHASE_SUPPLIER_GROUPS = {
    '青菜供应商': [
        '菠菜', '麦芹', '香菜', '圆生菜', '青菜（把）', '青菜', '小青菜',
        '广东菜心', '菜心', '香葱', '小葱',
    ],
    '王伟供应商': [
        '豇豆', '长豇豆', '白豆角', '螺丝椒', '蒜苔', '韭菜',
    ],
    '自采': [
        '纯牛奶', '牛奶', '水果玉米', '青笋', '平菇', '黄瓜', '西兰花', '西蓝花',
        '红萝卜', '胡萝卜', '洋葱', '新蒜', '净蒜', '蒜苗', '白萝卜', '象牙萝卜',
        '生姜', '姜',
    ],
}

PURCHASE_SUPPLIER_GROUP_ORDER = {
    '青菜供应商': 1,
    '王伟供应商': 2,
    '自采': 3,
}

PURCHASE_SUPPLIER_ITEM_ORDER = {
    '菠菜': 1, '麦芹': 2, '香菜': 3, '圆生菜': 4, '青菜（把）': 5, '青菜': 5, '小青菜': 5,
    '广东菜心': 6, '菜心': 6, '香葱': 7, '小葱': 8,
    '豇豆': 101, '长豇豆': 101, '螺丝椒': 102, '蒜苔': 103, '韭菜': 104, '白豆角': 105,
    '纯牛奶': 201, '牛奶': 201, '水果玉米': 202, '青笋': 203, '平菇': 204,
    '黄瓜': 205, '西兰花': 206, '西蓝花': 206, '红萝卜': 207, '胡萝卜': 207,
    '洋葱': 208, '新蒜': 209, '净蒜': 210, '蒜苗': 211, '白萝卜': 212, '象牙萝卜': 212,
    '生姜': 213, '姜': 213,
}


def purchase_supplier_group(name):
    text = normalize_hotel_item_name(str(name or ''))
    for group, names in PURCHASE_SUPPLIER_GROUPS.items():
        if text in names:
            return group
    return ''


def purchase_group_label(name):
    supplier_group = purchase_supplier_group(name)
    if supplier_group:
        return supplier_group
    return purchase_category_label(hotel_item_category(name))


def public_purchase_type_label(name):
    text = str(name or '')
    category = hotel_item_category(text)
    if any(keyword in text for keyword in ['鸡蛋', '鸭蛋', '咸鸭蛋', '蛋']):
        return '蛋'
    if any(keyword in text for keyword in ['牛奶', '酸奶', '纯奶', '纯牛奶', '乳', '优酪']):
        return '奶'
    if category == '蔬菜类':
        return '蔬菜'
    if category == '水果类':
        return '水果'
    if category == '肉蛋类':
        return '肉'
    if category == '主食面点类':
        return '面食'
    if category == '冻货类':
        return '冻货'
    if category == '豆制品类':
        return '豆制品'
    if category == '粮油干货类':
        return '粮油'
    if category == '调料咸菜类':
        return '调料'
    if category == '饮品乳品类':
        return '饮品'
    if category == '饼干':
        return '饼干'
    return '其他'


def sort_hotel_items_by_category(items):
    rank = {category: idx for idx, category in enumerate(HOTEL_CATEGORY_ORDER)}
    indexed = list(enumerate(items))
    indexed.sort(key=lambda pair: (rank.get(hotel_item_category(pair[1]['name']), 999), pair[0]))
    return [item for _, item in indexed]


PURCHASE_REVIEW_NAME_GROUPS = {
    '姜': '生姜',
    '生姜': '生姜',
    '菠菜': '菠菜',
    '波菜': '菠菜',
    '青菜': '青菜',
    '小青菜': '青菜',
    '青菜（把）': '青菜',
    '西蓝花': '西兰花',
    '西兰花': '西兰花',
    '蒜苔': '蒜苔',
    '蒜薹': '蒜苔',
    '土豆': '土豆',
    '黄土豆': '土豆',
    '茄子': '茄子',
    '广茄': '茄子',
    '圆茄': '茄子',
    '圆茄子': '茄子',
    '园茄子': '茄子',
    '白萝卜': '白萝卜',
    '象牙萝卜': '白萝卜',
    '象牙白萝卜': '白萝卜',
    '花心萝卜': '白萝卜',
    '莲花白': '花白',
    '莲莲花白': '花白',
    '花白': '花白',
    '长豆角': '长豇豆',
    '长江豆': '长豇豆',
    '长豇豆': '长豇豆',
    '豆芽': '豆芽',
    '大豆芽': '大豆芽',
    '小豆芽': '小豆芽',
    '黄豆芽': '黄豆芽',
    '豆腐': '豆腐',
    '老豆腐': '豆腐',
    '嫩豆腐': '豆腐',
    '豆腐干': '豆腐',
    '豆腐皮': '豆腐',
    '豆皮': '豆腐',
    '油豆皮': '油豆皮',
    '鲜豆皮': '鲜豆皮',
}


PURCHASE_REVIEW_NAME_ORDER = {
    '青菜（把）': 1, '青菜': 2, '小青菜': 3,
    '韭菜': 4, '土豆': 5, '黄土豆': 6,
    '蒜苔': 7, '蒜薹': 8, '黄瓜': 9,
    '贝贝南瓜': 10, '水果玉米': 11,
    '花白': 12, '莲花白': 13, '莲莲花白': 14,
    '奶白菜': 15, '罗马生菜': 16,
    '小葱': 17, '香葱': 18, '香菜': 19,
    '白菜': 20, '大白菜': 21,
    '豇豆': 22, '长豇豆': 23, '长江豆': 24, '长豆角': 25,
    '芹菜': 26, '麦芹': 27, '毛芹': 28,
    '菜花': 29, '花菜': 30, '有机菜花': 31,
    '白萝卜': 32, '红萝卜': 33, '胡萝卜': 34, '广红': 35,
    '红薯': 36,
    '西葫芦': 37, '西红柿': 38,
    '青椒': 39, '平菇': 40, '杏鲍菇': 41, '莲菜': 42,
    '豆皮': 101, '海带丝': 102, '干米线': 103, '湿米线': 104,
    '大豆芽': 104, '豆芽': 105, '黄豆芽': 106, '小豆芽': 107, '酸豆角': 108, '凉粉': 109,
    '圣女果': 201, '桃子': 202, '西瓜': 203, '白心火龙果': 204, '火龙果': 205, '苹果': 206, '香蕉': 207, '哈密瓜': 208, '青提': 209, '油桃': 210,
    '五花肉': 301, '五花肉片': 302, '排骨': 303, '猪耳朵': 304, '鸡蛋': 305,
    '法棍': 401, '细薄韭叶面': 402, '肉包子': 403, '臊子面': 404, '馒头': 405,
    '奶油沙拉酱': 501, '干木耳': 502, '白糖': 503,
    '冰红茶': 601, '瓶装冰峰': 602, '纯牛奶': 603, '益生菌酸奶': 604, '蓝莓酱': 605, '苹果酱': 606,
    '冻玉米': 701, '火锅丸子': 702, '里昂火腿': 703, '馄饨': 704, '花卷': 705,
    '原味面包': 801, '黑全麦面包': 802,
    '豆浆粉': 901, '核桃仁': 902, '瓜子仁': 903, '腰果': 904,
    '六六红': 920, '水晶粉': 921, '火锅宽粉': 922, '火锅邵皮': 923, '火锅苕皮': 924,
    '芝麻酱': 925, '贡菜': 926, '针金菇': 927, '金针菇': 927, '鞭炮笋': 928,
    '香铃卷': 929, '鲜玉米': 930, '鲜虾': 931, '青岛纯生': 932,
    '生姜': 950, '姜': 951,
    '菠菜': 952, '波菜': 953,
    '西兰花': 954, '西蓝花': 955,
}

VEGETABLE_REVIEW_ORDER = [
    '青菜（把）', '青菜', '小青菜', '黄瓜',
    '水果玉米', '韭黄', '菜花', '花菜', '有机菜花', '红萝卜', '胡萝卜', '广红',
    '西红柿', '莲菜', '娃娃菜', '豆苗', '广东菜心',
    '韭菜', '土豆', '黄土豆', '蒜苔', '蒜薹',
    '贝贝南瓜', '南瓜',
    '花白', '莲花白', '莲莲花白', '脆花白',
    '奶白菜', '罗马生菜',
    '小葱', '香葱', '香菜',
    '白菜', '大白菜', '生菜', '叶生菜', '圆生菜',
    '豇豆', '长豇豆', '长江豆', '长豆角',
    '韭菜', '芹菜', '麦芹', '毛芹',
    '白萝卜',
    '红薯', '西葫芦', '青椒',
    '平菇', '杏鲍菇',
    '包菜', '黄芽菜', '娃娃菜',
    '菜心', '奶白菜', '菠菜', '波菜', '芥菜', '苦菊',
    '西兰花', '西蓝花', '紫甘蓝',
    '山药', '水洗铁棍山药', '青笋', '净笋',
    '生姜', '净蒜', '蒜苗', '大葱', '白葱', '洋葱',
    '冬瓜', '苦瓜',
    '茄子', '广茄', '圆茄子',
    '青椒', '红椒', '线椒', '红线椒', '绿线椒', '螺丝椒',
    '白玉菇', '金针菇', '平菇', '蘑菇', '磨菇', '香菇', '杏鲍菇',
    '白豆角', '豆王', '荷兰豆', '豇豆', '长豆角', '长豇豆', '豆苗', '水果玉米',
]

VEGETABLE_REVIEW_RANK = {}
for idx, name in enumerate(VEGETABLE_REVIEW_ORDER):
    VEGETABLE_REVIEW_RANK.setdefault(name, idx)


def purchase_review_name_group(name):
    text = re.sub(r'\s+', '', str(name or '').strip())
    return PURCHASE_REVIEW_NAME_GROUPS.get(text, text)


def purchase_review_name_order(name):
    text = re.sub(r'\s+', '', str(name or '').strip())
    return PURCHASE_REVIEW_NAME_ORDER.get(text, 50)


def vegetable_review_rank(name):
    text = re.sub(r'\s+', '', str(name or '').strip())
    return VEGETABLE_REVIEW_RANK.get(text, 9999)


def sort_purchase_review_items(items):
    rank = {category: idx + 10 for idx, category in enumerate(HOTEL_CATEGORY_ORDER)}
    indexed = list(enumerate(items))
    def sort_key(pair):
        item = pair[1]
        raw_name = item.get('name') or item.get('product_name') or ''
        name = normalize_hotel_item_name(raw_name)
        supplier_group = purchase_supplier_group(name)
        category = hotel_item_category(name)
        return (
            PURCHASE_SUPPLIER_GROUP_ORDER.get(supplier_group, rank.get(category, 999)),
            PURCHASE_SUPPLIER_ITEM_ORDER.get(name, 9999),
            vegetable_review_rank(name) if category == '蔬菜类' else 9999,
            purchase_review_name_order(name),
            purchase_review_name_group(name),
            name,
            item.get('unit') or item.get('product_unit') or '',
            pair[0],
        )
    indexed.sort(key=sort_key)
    return [item for _, item in indexed]


def hotel_sales_order_key(customer, item_name):
    """全季酒店水果要单独生成一张销售单。"""
    customer_text = str(customer or '')
    if '全季' in customer_text and hotel_item_category(item_name) == '水果类':
        return 'fruit'
    if '全季' in customer_text:
        return 'regular'
    return ''


def split_front_desk_fruit_text(customer, text):
    """美居酒店备注“前台水果”后面的明细单独开销售单。"""
    customer_text = str(customer or '')
    raw_text = str(text or '')
    if '美居' not in customer_text or '前台水果' not in raw_text:
        return [(raw_text, '')]

    before, after = raw_text.split('前台水果', 1)
    after = re.sub(r'^[（(][^）)]*[）)]', '', after).strip()
    after = re.sub(r'^[：:，,、；;。\s]+', '', after)

    sections = []
    if before.strip():
        sections.append((before, ''))
    if after.strip():
        sections.append((after, 'front_fruit'))
    return sections or [(raw_text, '')]


def clean_hotel_item_note(note):
    text = str(note or '').strip()
    text = re.sub(r'^[,，、；;。:：\\s]+', '', text)
    text = re.sub(r'[）)]$', '', text)
    text = text.strip()
    return text


def split_parenthetical_note(text):
    notes = []

    def collect(match):
        content = clean_hotel_item_note(match.group(1))
        if content == '把':
            return match.group(0)
        if content:
            notes.append(content)
        return ''

    cleaned = re.sub(r'[（(]([^）)]{1,30})[）)]', collect, str(text or ''))
    return cleaned, '；'.join(notes)


HOTEL_NOTE_LINE_PATTERNS = (
    '分开装', '要好的', '去根的', '有辣椒少', '切好', '去皮',
    '不要太大', '不要太小', '品质好一些', '新鲜的', '大的', '小的'
)
HOTEL_SECTION_HEADINGS = {'凉菜', '炒面', '牛肉面', '手抓', '热菜', '水果', '蔬菜', '面点'}


def apply_hotel_section_note(items, section, seen_sections, start_idx):
    section_text = str(section or '').strip()
    if not section_text or section_text in seen_sections or len(items) <= start_idx:
        return
    first_item = items[start_idx]
    old_note = clean_hotel_item_note(first_item.get('note') or '')
    note_parts = [section_text]
    if old_note:
        note_parts.append(old_note)
    first_item['note'] = '；'.join(note_parts)
    seen_sections.add(section_text)


def normalize_hotel_repeated_units(text, unit_pattern, number_pattern):
    normalized = str(text or '')
    normalized = re.sub(
        rf'(?P<num>{number_pattern})(?P<unit>袋|箱|桶|瓶|包|盒|件|把|捆)子',
        r'\g<num>\g<unit>',
        normalized
    )

    def half_unit_repl(match):
        amount = parse_quantity(match.group('num')) + 0.5
        return f'{amount:g}{match.group("unit")}'

    normalized = re.sub(
        rf'(?P<num>{number_pattern})(?P<unit>{unit_pattern})半',
        half_unit_repl,
        normalized
    )
    note_pattern = '|'.join(re.escape(note) for note in HOTEL_NOTE_LINE_PATTERNS)
    normalized = re.sub(
        rf'(?P<num>{number_pattern})(?P<unit>{unit_pattern})(?P=unit)(?=(?:{note_pattern}))',
        r'\g<num>\g<unit>',
        normalized
    )
    return re.sub(
        rf'(?P<num>{number_pattern})(?P<unit>{unit_pattern})(?P=unit)(?=$|[\s，,、；;。])',
        r'\g<num>\g<unit>',
        normalized
    )


def merge_hotel_wrapped_lines(text, unit_pattern, number_pattern):
    raw = str(text or '').replace('\r\n', '\n').replace('\r', '\n')
    if '\n' not in raw:
        return raw

    note_pattern = '|'.join(re.escape(note) for note in HOTEL_NOTE_LINE_PATTERNS)
    complete_item = re.compile(
        rf'(?:{number_pattern})(?:{unit_pattern})(?:$|[）)]|(?:(?:{note_pattern}))+$)'
    )
    prefix_complete_item = re.compile(
        rf'^(?:{number_pattern})(?:{unit_pattern})[\u4e00-\u9fa5A-Za-z].+$'
    )
    merged = []
    for line in raw.split('\n'):
        current = line.strip()
        if not current:
            continue
        if not merged:
            merged.append(current)
            continue

        previous = merged[-1]
        is_note_line = current in HOTEL_NOTE_LINE_PATTERNS
        is_heading = re.match(r'^(备注|注)[:：]', current) or re.match(r'^[\u4e00-\u9fa5]{1,8}类[:：]?$', current)
        has_explicit_delimiter = re.search(r'[，,、；;。:：]$', previous)

        if is_note_line:
            merged[-1] = previous + current
        elif has_explicit_delimiter or is_heading or complete_item.search(previous) or prefix_complete_item.match(previous.lstrip('，,、；;。')):
            merged.append('，' + current)
        else:
            merged.append(current)

    return ''.join(merged)


def append_hotel_item(items, name, quantity, unit, note=''):
    item_name, item_quantity, item_unit = normalize_hotel_item_quantity_unit(name, quantity, unit)
    item = {
        'name': item_name,
        'quantity': item_quantity,
        'unit': item_unit,
    }
    note_text = clean_hotel_item_note(note)
    if note_text:
        item['note'] = note_text
    if item['name'] and item['quantity'] > 0 and item['unit']:
        items.append(item)


SIMPLE_PURCHASE_NOTE_WORDS = {
    '斤', '个', '件', '包', '袋', '箱', '瓶', '盒', '盘', '桶', '把', '根',
    '条', '颗', '捆', '提', '只', '块', '张', '份', '卷', '板'
}


def purchase_export_note(name, note):
    item_name = normalize_hotel_item_name(name)
    parts = []
    for part in re.split(r'[；;]+', str(note or '')):
        text = clean_hotel_item_note(part)
        if not text:
            continue
        normalized_note = normalize_hotel_item_name(text)
        if text in SIMPLE_PURCHASE_NOTE_WORDS:
            continue
        if normalized_note == item_name:
            continue
        if text in item_name:
            continue
        if text not in parts:
            parts.append(text)
    return '；'.join(parts)


def build_purchase_list_workbook(items, title):
    wb = Workbook()
    ws = wb.active
    ws.title = '采购单'
    ws.append([title, '', '', '', ''])
    ws.append(['食材分类', '食材名称', '采购数量', '计量单位', '备注'])

    thin = Side(style='thin', color='666666')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='F2F2F2')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center

    current_category = None
    for item in sort_purchase_review_items(items):
        name = item.get('name') or item.get('product_name') or ''
        category = purchase_group_label(name)
        show_category = category if category != current_category else ''
        current_category = category
        ws.append([
            show_category,
            name,
            clean_number(item.get('quantity', 0)),
            item.get('unit') or item.get('product_unit') or '',
            purchase_export_note(name, item.get('note') or '')
        ])

    widths = [18, 22, 14, 12, 24]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = border
            cell.alignment = center if cell.column in (1, 3, 4) else left
        if row[0].row == 2:
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = header_fill

    return wb


def short_hotel_customer_name(customer):
    text = str(customer or '').strip()
    if '汉庭' in text:
        return '汉庭'
    if '美居' in text:
        return '浐灞'
    if '全季' in text:
        return '全季'
    return text or '客户'


def hotel_customer_export_rank(customer):
    text = str(customer or '')
    if '汉庭' in text:
        return (1, text)
    if '美居' in text or '浐灞' in text:
        return (2, text)
    if '全季' in text:
        return (3, text)
    return (99, text)


def build_customer_split_purchase_workbook(summary_items, customer_items, customers, title):
    wb = Workbook()
    ws = wb.active
    ws.title = '采购单'

    customer_headers = [short_hotel_customer_name(customer) for customer in customers]
    headers = ['食材分类', '食材名称', '采购数量', '计量单位'] + customer_headers + ['备注']
    ws.append([title] + [''] * (len(headers) - 1))
    ws.append(headers)

    end_col = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)

    thin = Side(style='thin', color='666666')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='F2F2F2')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center

    by_customer = {
        (row['name'], row['unit'], row.get('purchase_group') or '', row['customer']): row['quantity']
        for row in customer_items
    }
    current_category = None
    for item in sort_purchase_review_items(summary_items):
        name = item.get('name') or ''
        unit = item.get('unit') or ''
        purchase_group = item.get('purchase_group') or ''
        category = purchase_group_label(name)
        show_category = category if category != current_category else ''
        current_category = category
        row = [
            show_category,
            name,
            clean_number(item.get('quantity', 0)),
            unit,
        ]
        row.extend(clean_number(by_customer.get((name, unit, purchase_group, customer), 0)) or '' for customer in customers)
        row.append(purchase_export_note(name, item.get('note') or ''))
        ws.append(row)

    widths = [18, 22, 14, 12] + [10] * len(customers) + [24]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.border = border
            cell.alignment = center if cell.column != 2 else left
        if row[0].row == 2:
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = header_fill

    return wb


def parse_hotel_grocery_text(text, use_smart=True, preserve_order=False):
    unit_pattern = '|'.join(re.escape(unit) for unit in sorted(SALE_UNITS, key=len, reverse=True))
    number_pattern = r'\d+(?:\.\d+)?|[零〇一二两俩三四五六七八九十百半]+'
    normalized = normalize_hotel_text_aliases(
        merge_hotel_wrapped_lines(
            normalize_hotel_repeated_units(text, unit_pattern, number_pattern),
            unit_pattern,
            number_pattern
        )
    ).strip()
    normalized = re.sub(r'(?<![一二两俩三四五六七八九十百\\d])一小把小香蕉', '小香蕉1把', normalized)
    normalized = re.sub(r'(?<![一二两俩三四五六七八九十百\\d])一把小香蕉', '小香蕉1把', normalized)
    normalized = re.sub(
        rf'吐司面包黑白(?:名|各)?(?P<num>{number_pattern}|俩)(?P<unit>{unit_pattern})',
        r'黑吐司面包\g<num>\g<unit>，白吐司面包\g<num>\g<unit>',
        normalized
    )
    normalized = re.sub(
        rf'黑白吐司面包(?:名|各)?(?P<num>{number_pattern}|俩)(?P<unit>{unit_pattern})',
        r'黑吐司面包\g<num>\g<unit>，白吐司面包\g<num>\g<unit>',
        normalized
    )
    normalized = re.sub(
        rf'吐司黑白(?:名|各)?(?P<num>{number_pattern}|俩)(?P<unit>{unit_pattern})',
        r'黑吐司面包\g<num>\g<unit>，白吐司面包\g<num>\g<unit>',
        normalized
    )
    normalized = re.sub(
        rf'黄油果酱各(?P<num>{number_pattern})(?P<unit>{unit_pattern})',
        r'黄油\g<num>\g<unit>，果酱\g<num>\g<unit>',
        normalized
    )
    normalized = re.sub(r'^\s*酒店买菜[:：]\s*', '', normalized)
    normalized = re.sub(r'([）)])(?=(?![零〇一二两俩三四五六七八九十百半\d])[\u4e00-\u9fa5A-Za-z])', r'\1,', normalized)
    normalized = re.sub(r'[，、；;。\n\r\t]+', ',', normalized)
    parts = [part.strip() for part in normalized.split(',') if part.strip()]

    items = []
    errors = []
    current_section = ''
    seen_sections = set()
    for part in parts:
        if re.match(r'^(备注|注)[:：]', part):
            continue
        section_match = re.match(r'^(凉菜|炒面|牛肉面|手抓|热菜|水果|蔬菜|面点)[:：]$', part)
        if section_match:
            current_section = section_match.group(1)
            continue
        if re.match(r'^[\u4e00-\u9fa5]{1,8}类[:：]?$', part):
            continue
        clean_part, paren_note = split_parenthetical_note(part)
        compact = re.sub(r'\s+', '', clean_part)
        section_start_idx = len(items)
        glued = re.match(rf'^(?P<base>.+?)(?P<numseq>\d{{3,}})(?P<unit>{unit_pattern})$', compact)
        if glued and glued.group('numseq')[-1] != '0':
            numseq = glued.group('numseq')
            append_hotel_item(
                items,
                glued.group('base') + numseq[:-1],
                float(numseq[-1]),
                glued.group('unit'),
                paren_note
            )
            apply_hotel_section_note(items, current_section, seen_sections, section_start_idx)
            continue

        multi_suffix = list(re.finditer(rf'(?P<name>.+?)(?P<num>{number_pattern})(?P<unit>{unit_pattern})', compact))
        if multi_suffix:
            parsed_items = []
            cursor = 0
            valid = True
            for match in multi_suffix:
                if compact[cursor:match.start()].strip():
                    valid = False
                    break
                item_name = re.sub(r'各$', '', match.group('name').strip())
                if not item_name:
                    valid = False
                    break
                item_name, item_quantity, item_unit = normalize_hotel_item_quantity_unit(
                    item_name,
                    parse_quantity(match.group('num')),
                    match.group('unit')
                )
                parsed_items.append({
                    'name': item_name,
                    'quantity': item_quantity,
                    'unit': item_unit,
                    'note': paren_note,
                })
                cursor = match.end()
            if valid and not compact[cursor:].strip():
                items.extend(parsed_items)
                apply_hotel_section_note(items, current_section, seen_sections, section_start_idx)
                continue

        suffix = re.match(rf'^(?P<name>.+?)(?P<num>{number_pattern})(?P<unit>{unit_pattern})(?P<note>.*)$', compact)
        if suffix:
            note = '；'.join(part for part in [paren_note, clean_hotel_item_note(suffix.group('note'))] if part)
            append_hotel_item(
                items,
                suffix.group('name'),
                parse_quantity(suffix.group('num')),
                suffix.group('unit'),
                note
            )
            apply_hotel_section_note(items, current_section, seen_sections, section_start_idx)
            continue

        prefix = re.match(rf'^(?P<num>{number_pattern})(?P<unit>{unit_pattern})(?P<name>.+)$', compact)
        if prefix:
            append_hotel_item(
                items,
                prefix.group('name'),
                parse_quantity(prefix.group('num')),
                prefix.group('unit'),
                paren_note
            )
            apply_hotel_section_note(items, current_section, seen_sections, section_start_idx)
            continue

        missing_unit_with_note = re.match(
            rf'^(?P<name>.+?)(?P<num>{number_pattern})(?P<note>分开装|要好的|去根的|有辣椒少|切好|去皮|不要太大|不要太小|新鲜的)$',
            compact
        )
        if missing_unit_with_note:
            item_name = normalize_hotel_item_name(re.sub(r'各$', '', missing_unit_with_note.group('name').strip()))
            inferred_unit = default_hotel_unit(item_name)
            if inferred_unit:
                note = '；'.join(part for part in [paren_note, clean_hotel_item_note(missing_unit_with_note.group('note'))] if part)
                append_hotel_item(
                    items,
                    item_name,
                    parse_quantity(missing_unit_with_note.group('num')),
                    inferred_unit,
                    note
                )
                apply_hotel_section_note(items, current_section, seen_sections, section_start_idx)
                continue

        missing_unit = re.match(rf'^(?P<name>.+?)(?P<num>{number_pattern})$', compact)
        if missing_unit:
            item_name = normalize_hotel_item_name(re.sub(r'各$', '', missing_unit.group('name').strip()))
            inferred_unit = default_hotel_unit(item_name)
            if inferred_unit:
                append_hotel_item(
                    items,
                    item_name,
                    parse_quantity(missing_unit.group('num')),
                    inferred_unit,
                    paren_note
                )
                apply_hotel_section_note(items, current_section, seen_sections, section_start_idx)
                continue

        errors.append(part)

    items = [item for item in items if item['name'] and item['quantity'] > 0]
    smart_error = ''
    if use_smart and errors:
        smart_items, smart_error = ark_smart_parse_hotel_grocery_text('，'.join(errors))
        if smart_items:
            items.extend(smart_items)
            errors = []
    sorted_items = items if preserve_order else sort_hotel_items_by_category(items)
    for item in sorted_items:
        if smart_error:
            item.setdefault('smart_error', smart_error)
    return sorted_items, errors


def find_product_with_conn(conn, name, spec='', unit=''):
    row = conn.execute('''
        SELECT * FROM products
        WHERE name = ?
          AND (? = '' OR COALESCE(spec, '') = ?)
          AND (? = '' OR unit = ?)
        LIMIT 1
    ''', (name, spec, spec, unit, unit)).fetchone()
    if not row:
        row = conn.execute('SELECT * FROM products WHERE name = ? LIMIT 1', (name,)).fetchone()
    if not row:
        row = conn.execute('SELECT * FROM products WHERE name LIKE ? LIMIT 1', (f'%{name}%',)).fetchone()
    return dict(row) if row else None


def canonical_customer_name(customer):
    text = str(customer or '').strip()
    if not text:
        return '酒店'
    rows = db.query('SELECT name FROM customers ORDER BY created_at DESC')
    for row in rows:
        name = row.get('name') or ''
        if text == name or text in name or name in text:
            return name
    compact = re.sub(r'[\s（）()]', '', text)
    for row in rows:
        name = row.get('name') or ''
        name_compact = re.sub(r'[\s（）()]', '', name)
        if compact and (compact in name_compact or name_compact in compact):
            return name
    if '美居' in text:
        return '浐灞美居'
    if '汉庭' in text:
        return '西安汉庭酒店（大明宫万达）'
    if '全季' in text:
        return '全季酒店（辛家庙店）'
    return text


def ensure_customer_name_with_conn(conn, customer):
    customer_name = canonical_customer_name(customer)
    if not customer_name:
        customer_name = '酒店'
    existing = conn.execute('SELECT id FROM customers WHERE name = ? LIMIT 1', (customer_name,)).fetchone()
    if existing:
        return customer_name
    code = generate_no('KH')
    while conn.execute('SELECT 1 FROM customers WHERE code = ?', (code,)).fetchone():
        code = generate_no('KH')
    conn.execute('''
        INSERT INTO customers (id, code, name, company, phone, address, note, created_at)
        VALUES (?, ?, ?, ?, '', '', '销售单导入自动新增', ?)
    ''', (generate_id(), code, customer_name, customer_name, now_str()))
    return customer_name


def get_customer_product_price_with_conn(conn, customer, product, unit=''):
    customer_name = canonical_customer_name(customer)
    product_id = product.get('id') if isinstance(product, dict) else ''
    product_name = product.get('name') if isinstance(product, dict) else ''
    unit_text = unit or (product.get('unit') if isinstance(product, dict) else '') or ''
    row = None
    if product_id:
        row = conn.execute('''
            SELECT price FROM customer_product_prices
            WHERE customer = ? AND product_id = ? AND COALESCE(product_unit, '') = ?
            LIMIT 1
        ''', (customer_name, product_id, unit_text)).fetchone()
        if not row:
            row = conn.execute('''
                SELECT price FROM customer_product_prices
                WHERE customer = ? AND product_id = ?
                ORDER BY updated_at DESC
                LIMIT 1
            ''', (customer_name, product_id)).fetchone()
    if not row and product_name:
        row = conn.execute('''
            SELECT price FROM customer_product_prices
            WHERE customer = ? AND product_name = ? AND (? = '' OR COALESCE(product_unit, '') = ?)
            ORDER BY updated_at DESC
            LIMIT 1
        ''', (customer_name, product_name, unit_text, unit_text)).fetchone()
    return as_float(row['price']) if row else 0


def find_customer_item_price_match_with_conn(conn, customer, name, unit=''):
    customer_name = canonical_customer_name(customer)
    raw_name = re.sub(r'\s+', '', str(name or '').strip())
    match_name = normalize_hotel_match_name(raw_name)
    unit_text = str(unit or '').strip()
    rows = conn.execute('''
        SELECT cpp.*, p.name AS product_canonical_name
        FROM customer_product_prices cpp
        LEFT JOIN products p ON p.id = cpp.product_id
        WHERE cpp.customer = ?
          AND (? = '' OR COALESCE(cpp.product_unit, '') = ?)
        ORDER BY cpp.updated_at DESC, cpp.created_at DESC
    ''', (customer_name, unit_text, unit_text)).fetchall()
    exact_match = None
    normalized_match = None
    for row in rows:
        product_name = str(row['product_name'] or '').strip()
        canonical_name = str(row['product_canonical_name'] or '').strip()
        if product_name == raw_name:
            exact_match = row
            break
        if not normalized_match and (
            normalize_hotel_match_name(product_name) == match_name
            or normalize_hotel_match_name(canonical_name) == match_name
        ):
            normalized_match = row
    chosen = exact_match or normalized_match
    return dict(chosen) if chosen else None


def get_customer_product_price(customer, product, unit=''):
    conn = db._get_conn()
    try:
        return get_customer_product_price_with_conn(conn, customer, product, unit)
    finally:
        conn.close()


def fixed_hotel_sale_price(name, unit=''):
    item_name = normalize_hotel_item_name(name)
    unit_text = str(unit or '').strip()
    if item_name == '方块面筋':
        if unit_text == '包':
            return 23
        if unit_text == '个':
            return 2.3
    return 0


def upsert_customer_product_price(conn, customer, product, unit, price, source='', display_name=None):
    price = as_float(price)
    if price <= 0 or not product:
        return False
    customer_name = canonical_customer_name(customer)
    product_id = product.get('id') if isinstance(product, dict) else product['id']
    product_name = product.get('name') if isinstance(product, dict) else product['name']
    unit_text = unit or (product.get('unit') if isinstance(product, dict) else product['unit']) or ''
    existing = conn.execute('''
        SELECT id FROM customer_product_prices
        WHERE customer = ? AND product_id = ? AND COALESCE(product_unit, '') = ?
        LIMIT 1
    ''', (customer_name, product_id, unit_text)).fetchone()
    if existing:
        conn.execute('''
            UPDATE customer_product_prices
            SET product_name = ?, price = ?, source = ?, updated_at = ?
            WHERE id = ?
        ''', (product_name, price, source or '', now_str(), existing['id']))
    else:
        conn.execute('''
            INSERT INTO customer_product_prices
                (id, customer, product_id, product_name, product_unit, price, source, updated_at, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (generate_id(), customer_name, product_id, product_name, unit_text, price, source or '', now_str(), now_str()))
    return True


def prepare_hotel_grocery_sales_rows(conn, items, date_str, customer, company='', phone=''):
    created_products = []
    rows_data = []
    for item in items:
        preferred = find_customer_item_price_match_with_conn(conn, customer, item['name'], item['unit'])
        sale_item_name = item['name']
        product = None
        if preferred and preferred.get('product_id'):
            product_row = conn.execute('SELECT * FROM products WHERE id = ? LIMIT 1', (preferred['product_id'],)).fetchone()
            product = dict(product_row) if product_row else None
        if not product:
            product = (
                find_product_with_conn(conn, sale_item_name, unit=item['unit'])
                or find_product_with_conn(conn, item['name'], unit=item['unit'])
            )
        if not product:
            product = create_hotel_flow_product(conn, sale_item_name, '', item['unit'], 0)
            created_products.append(sale_item_name)
        sale_item_name = product.get('name') or sale_item_name

        sale_price = (
            (as_float(preferred.get('price')) if preferred else 0)
            or
            get_customer_product_price_with_conn(conn, customer, product, item['unit'])
            or fixed_hotel_sale_price(item['name'], item['unit'])
            or as_float(product.get('last_sale_price'))
            or as_float(product.get('cost'))
        )
        rows_data.append({
            'date': date_str,
            'customer': canonical_customer_name(customer or '酒店'),
            'company': company,
            'phone': phone,
            'order_key': item.get('order_key') or hotel_sales_order_key(customer, item['name']),
            'product': product,
            'product_name': sale_item_name,
            'product_unit': item['unit'],
            'quantity': item['quantity'],
            'price': sale_price,
            'note': item.get('note') or '',
        })
    return rows_data, created_products


def sync_hotel_grocery_orders(items, date_str, supplier_name, customer, company='', phone=''):
    """酒店清单直出销售单：不生成采购入库单，不压库存。"""
    if not items:
        return None, {'error': '没有可同步的明细'}, 400

    conn = db._get_conn()
    try:
        rows_data, created_products = prepare_hotel_grocery_sales_rows(
            conn, items, date_str, customer, company, phone
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        return None, {'error': str(e)}, 500
    finally:
        conn.close()

    result, error, status = create_hotel_sales_orders(rows_data)
    if error:
        return None, error, status
    push_sales_to_feishu_async(result.get('saleIds') or [])
    result.update({
        'createdProducts': len(created_products),
        'createdProductNames': sorted(set(created_products)),
        'order': [item['name'] for item in items],
    })
    return result, None, status


def hotel_sale_signature(items):
    signature = []
    for item in items:
        product = item.get('product') or {}
        name = item.get('product_name') or product.get('name') or ''
        unit = item.get('product_unit') or product.get('unit') or ''
        signature.append((name, unit, round(as_float(item.get('quantity')), 4), round(as_float(item.get('price')), 4)))
    return sorted(signature)


def find_existing_hotel_sale(conn, date_str, customer, items):
    target_signature = hotel_sale_signature(items)
    candidates = conn.execute('''
        SELECT s.id
        FROM sales s
        LEFT JOIN sales_items si ON si.sale_id = s.id
        WHERE s.purchase_id = ? AND s.date = ? AND s.customer = ?
        GROUP BY s.id
        HAVING COUNT(si.id) = ?
    ''', (HOTEL_FLOW_PURCHASE_ID, date_str, customer, len(target_signature))).fetchall()

    for candidate in candidates:
        rows = conn.execute('''
            SELECT
                COALESCE(NULLIF(si.product_name, ''), p.name, '') AS product_name,
                COALESCE(NULLIF(si.product_unit, ''), p.unit, '') AS product_unit,
                si.quantity,
                si.price
            FROM sales_items si
            LEFT JOIN products p ON si.product_id = p.id
            WHERE si.sale_id = ?
        ''', (candidate['id'],)).fetchall()
        candidate_signature = sorted(
            (
                row['product_name'] or '',
                row['product_unit'] or '',
                round(as_float(row['quantity']), 4),
                round(as_float(row['price']), 4),
            )
            for row in rows
        )
        if candidate_signature == target_signature:
            return candidate['id']
    return None


def create_hotel_sales_orders(rows_data, conn=None):
    """生成酒店销售单：库存品扣库存，生鲜流转品不影响库存。"""
    if not rows_data:
        return None, {'error': '没有可生成销售单的数据'}, 400

    groups = {}
    for item in rows_data:
        key = (item['date'], item['customer'], item.get('company', ''), item.get('phone', ''), item.get('order_key', ''))
        groups.setdefault(key, []).append(item)

    created = 0
    imported_items = 0
    deducted_items = 0
    duplicate_orders = 0
    sale_ids = []
    owns_conn = conn is None
    if owns_conn:
        conn = db._get_conn()
    try:
        for (date_str, customer, company, phone, _order_key), items in groups.items():
            sid = generate_id()
            sno = generate_no('XS')
            total = sum(item['quantity'] * item['price'] for item in items)
            existing_same = find_existing_hotel_sale(conn, date_str, customer, items)
            if existing_same:
                duplicate_orders += 1
                continue
            conn.execute('''
                INSERT INTO sales (id, no, company, customer, phone, date, total, show_handlers, order_key, purchase_id, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?)
            ''', (sid, sno, company, customer, phone, date_str, total, _order_key or '', HOTEL_FLOW_PURCHASE_ID, now_str()))

            for item in items:
                product = item['product']
                subtotal = item['quantity'] * item['price']
                sale_item_id = generate_id()
                conn.execute('''
                    INSERT INTO sales_items (id, sale_id, product_id, product_name, product_spec, product_unit, quantity, price, subtotal, note)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    sale_item_id, sid, product['id'], item.get('product_name') or product['name'], product['spec'] or '',
                    item.get('product_unit') or product['unit'], item['quantity'], item['price'], subtotal, item.get('note') or ''
                ))
                if as_float(item['price']) > 0:
                    upsert_customer_product_price(
                        conn,
                        customer,
                        product,
                        item.get('product_unit') or product.get('unit') or '',
                        item['price'],
                        source='hotel_sale'
                    )
                if not is_flow_item(product) and as_float(product.get('stock')) > 0:
                    conn.execute("UPDATE products SET stock = stock - ? WHERE id = ?", (item['quantity'], product['id']))
                    conn.execute('''
                        INSERT INTO hotel_stock_deductions (sale_item_id, sale_id, product_id, quantity)
                        VALUES (?, ?, ?, ?)
                    ''', (sale_item_id, sid, product['id'], item['quantity']))
                    deducted_items += 1
                imported_items += 1

            created += 1
            sale_ids.append(sid)

        if owns_conn:
            conn.commit()
        return {
            'created': created,
            'items': imported_items,
            'deductedItems': deducted_items,
            'duplicateOrders': duplicate_orders,
            'saleIds': sale_ids
        }, None, 200
    except Exception as e:
        if owns_conn:
            conn.rollback()
        return None, {'error': str(e)}, 500
    finally:
        if owns_conn:
            conn.close()


def delete_hotel_flow_sales_for_customer_date(conn, customer_name, date_str):
    sales = conn.execute('''
        SELECT id
        FROM sales
        WHERE customer = ? AND date = ? AND purchase_id = ?
    ''', (customer_name, date_str, HOTEL_FLOW_PURCHASE_ID)).fetchall()
    sale_ids = [row['id'] for row in sales]
    restored_items = 0
    for sale_id in sale_ids:
        deductions = conn.execute('''
            SELECT product_id, quantity
            FROM hotel_stock_deductions
            WHERE sale_id = ?
        ''', (sale_id,)).fetchall()
        for item in deductions:
            conn.execute(
                'UPDATE products SET stock = stock + ? WHERE id = ?',
                (item['quantity'], item['product_id'])
            )
            restored_items += 1
        conn.execute('DELETE FROM hotel_stock_deductions WHERE sale_id = ?', (sale_id,))
        conn.execute('DELETE FROM sales_items WHERE sale_id = ?', (sale_id,))
        conn.execute('DELETE FROM sales WHERE id = ?', (sale_id,))
    return {'deletedOrders': len(sale_ids), 'restoredItems': restored_items}


def is_flow_item(row):
    code = str(row.get('code', '') or '')
    return code.startswith('WP') or code.startswith('IMP')


def as_json_array(value):
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return json.dumps(parsed if isinstance(parsed, list) else [], ensure_ascii=False)
        except Exception:
            return '[]'
    if isinstance(value, list):
        return json.dumps(value, ensure_ascii=False)
    return '[]'


def first_value(data, *keys, default=''):
    for key in keys:
        if key in data and data[key] not in (None, ''):
            return data[key]
    return default


def normalize_legacy_backup(data):
    """把旧版 localStorage/data.json 备份转换成新版扁平表结构。"""
    products = data.get('products', [])
    product_map = {p.get('id'): p for p in products}

    result = {
        'products': products,
        'suppliers': data.get('suppliers', []),
        'customers': data.get('customers', []),
        'purchases': data.get('purchases', []),
        'purchaseItems': list(data.get('purchaseItems', [])),
        'outbounds': data.get('outbounds', []),
        'outboundItems': list(data.get('outboundItems', [])),
        'sales': data.get('sales', []),
        'salesItems': list(data.get('salesItems', [])),
        'hotelStockDeductions': list(data.get('hotelStockDeductions', [])),
        'finances': data.get('finances', [])
    }

    if not result['purchaseItems']:
        for order in result['purchases']:
            for index, item in enumerate(order.get('items', []), start=1):
                product = product_map.get(item.get('productId'), {})
                quantity = as_float(item.get('quantity'))
                price = as_float(item.get('price'))
                result['purchaseItems'].append({
                    'id': item.get('id') or f"{order.get('id')}-item-{index}",
                    'purchase_id': order.get('id'),
                    'product_id': item.get('productId') or item.get('product_id'),
                    'product_name': product.get('name', ''),
                    'product_spec': product.get('spec', ''),
                    'product_unit': product.get('unit', ''),
                    'quantity': quantity,
                    'price': price,
                    'subtotal': as_float(item.get('subtotal'), quantity * price)
                })

    if not result['outboundItems']:
        for order in result['outbounds']:
            for index, item in enumerate(order.get('items', []), start=1):
                product = product_map.get(item.get('productId'), {})
                quantity = as_float(item.get('quantity'))
                price = as_float(item.get('price'))
                result['outboundItems'].append({
                    'id': item.get('id') or f"{order.get('id')}-item-{index}",
                    'outbound_id': order.get('id'),
                    'product_id': item.get('productId') or item.get('product_id'),
                    'product_name': product.get('name', ''),
                    'product_spec': product.get('spec', ''),
                    'product_unit': product.get('unit', ''),
                    'quantity': quantity,
                    'price': price,
                    'subtotal': as_float(item.get('subtotal'), quantity * price)
                })

    if not result['salesItems']:
        for order in result['sales']:
            for index, item in enumerate(order.get('items', []), start=1):
                product = product_map.get(item.get('productId'), {})
                quantity = as_float(item.get('quantity'))
                price = as_float(item.get('price'))
                result['salesItems'].append({
                    'id': item.get('id') or f"{order.get('id')}-item-{index}",
                    'sale_id': order.get('id'),
                    'product_id': item.get('productId') or item.get('product_id'),
                    'product_name': product.get('name', ''),
                    'product_spec': product.get('spec', ''),
                    'product_unit': product.get('unit', ''),
                    'quantity': quantity,
                    'price': price,
                    'subtotal': as_float(item.get('subtotal'), quantity * price)
                })

    return result


def get_month_range():
    """获取本月起止日期"""
    today = date.today()
    start = today.replace(day=1)
    # 下月1日减1天 = 本月最后一天
    if today.month == 12:
        next_month = today.replace(year=today.year + 1, month=1, day=1)
    else:
        next_month = today.replace(month=today.month + 1, day=1)
    from datetime import timedelta
    end = next_month - timedelta(days=1)
    return start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d')


# ==================== API路由 ====================

@app.route('/')
def index():
    """备案首页：个人笔记页面。"""
    with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'blog.html'), encoding='utf-8') as f:
        template = f.read()
    return render_template_string(
        template,
        posts=fetch_blog_posts(),
        generated_at=datetime.now().strftime('%Y-%m-%d %H:%M')
    )


@app.route('/admin')
def admin_index():
    """后台管理入口。"""
    return send_file('index.html')


@app.route('/order')
def customer_order_page():
    """客户扫码下单页面"""
    return send_file('customer_order.html')


@app.route('/api/order-qr')
def order_qr():
    """生成客户下单入口二维码"""
    url = public_order_url()
    img = qrcode.make(url)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return send_file(buf, mimetype='image/png', download_name='客户下单二维码.png')


@app.route('/api/public/customers', methods=['GET'])
def public_customers():
    if not valid_public_order_request():
        return jsonify({'error': '下单链接已失效，请联系商家重新获取'}), 401
    rows = db.query('SELECT id, code, name FROM customers ORDER BY code ASC, created_at DESC')
    return jsonify(rows)


@app.route('/api/public/hotel-order', methods=['POST'])
def public_hotel_order():
    if not valid_public_order_request():
        return jsonify({'error': '下单链接已失效，请联系商家重新获取'}), 401

    data = request.json or {}
    customer_id = str(data.get('customerId') or '').strip()
    text = str(data.get('text') or '').strip()
    date_str = normalize_date(data.get('date') or date.today().strftime('%Y-%m-%d'))
    confirmed = data.get('confirmed') is True

    if not customer_id:
        return jsonify({'error': '请选择酒店'}), 400
    if not text:
        return jsonify({'error': '请填写下单清单'}), 400
    if not confirmed:
        return jsonify({'error': '请先识别并确认订单明细后再提交'}), 400

    customer_row = db.query_one('SELECT * FROM customers WHERE id = ?', (customer_id,))
    if not customer_row:
        return jsonify({'error': '酒店不存在，请刷新页面后重试'}), 404

    customer_name = customer_row.get('name') or '酒店'
    company = customer_row.get('company') or '西安禾润佳商贸有限公司'
    phone = customer_row.get('phone') or ''

    items, unparsed, _corrected_text = parse_public_order_items(customer_name, text)
    if not items:
        return jsonify({'error': '没有识别到“物品+数量+单位”的订单内容', 'unparsed': unparsed}), 400

    result, error, status = sync_hotel_grocery_orders(
        items=items,
        date_str=date_str,
        supplier_name='西安禾润佳商贸有限公司',
        customer=customer_name,
        company=company,
        phone=phone,
    )
    if error:
        error['unparsed'] = unparsed
        return jsonify(error), status

    result.update({
        'customer': customer_name,
        'date': date_str,
        'parsed': len(items),
        'smartParsed': sum(1 for item in items if item.get('smart')),
        'unparsed': unparsed,
    })
    return jsonify(result)


@app.route('/api/public/today-order/replace', methods=['POST'])
def public_replace_today_order():
    if not valid_public_order_request():
        return jsonify({'error': '下单链接已失效，请联系商家重新获取'}), 401

    data = request.json or {}
    customer_id = str(data.get('customerId') or '').strip()
    text = str(data.get('text') or '').strip()
    today_str = date.today().strftime('%Y-%m-%d')
    date_str = normalize_date(data.get('date') or today_str)
    confirmed = data.get('confirmed') is True

    if not customer_id:
        return jsonify({'error': '请选择酒店'}), 400
    if date_str != today_str:
        return jsonify({'error': '只能重新编辑当天订单'}), 400
    if not text:
        return jsonify({'error': '请填写下单清单'}), 400
    if not confirmed:
        return jsonify({'error': '请先识别并确认订单明细后再提交'}), 400

    customer_row = db.query_one('SELECT * FROM customers WHERE id = ?', (customer_id,))
    if not customer_row:
        return jsonify({'error': '酒店不存在，请刷新页面后重试'}), 404

    customer_name = customer_row.get('name') or '酒店'
    company = customer_row.get('company') or '西安禾润佳商贸有限公司'
    phone = customer_row.get('phone') or ''
    items, unparsed, _corrected_text = parse_public_order_items(customer_name, text)
    if not items:
        return jsonify({'error': '没有识别到“物品+数量+单位”的订单内容', 'unparsed': unparsed}), 400

    conn = db._get_conn()
    created_products = []
    try:
        rows_data, created_products = prepare_hotel_grocery_sales_rows(
            conn, items, date_str, customer_name, company, phone
        )
        replace_result = delete_hotel_flow_sales_for_customer_date(conn, customer_name, date_str)
        result, error, status = create_hotel_sales_orders(rows_data, conn=conn)
        if error:
            conn.rollback()
            error['unparsed'] = unparsed
            return jsonify(error), status
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e), 'unparsed': unparsed}), 500
    finally:
        conn.close()

    result.update({
        'customer': customer_name,
        'date': date_str,
        'parsed': len(items),
        'smartParsed': sum(1 for item in items if item.get('smart')),
        'unparsed': unparsed,
        'createdProducts': len(created_products),
        'createdProductNames': sorted(set(created_products)),
        'order': [item['name'] for item in items],
        'replacedOrders': replace_result.get('deletedOrders', 0),
        'restoredItems': replace_result.get('restoredItems', 0),
    })
    return jsonify(result)


def parse_public_order_items(customer_name, text):
    corrected_text = normalize_hotel_text_aliases(text)
    items = []
    unparsed = []
    for section_text, order_key in split_front_desk_fruit_text(customer_name, corrected_text):
        section_items, section_unparsed = parse_hotel_grocery_text(section_text, preserve_order=True)
        for item in section_items:
            if order_key:
                item['order_key'] = order_key
            items.append(item)
        unparsed.extend(section_unparsed)
    return items, unparsed, corrected_text


def public_order_item_preview(items):
    rows = []
    for item in items:
        name = item.get('name') or ''
        rows.append({
            'name': name,
            'quantity': clean_number(item.get('quantity')),
            'unit': item.get('unit') or '',
            'category': hotel_item_category(name),
            'purchaseType': public_purchase_type_label(name),
            'orderGroup': item.get('order_key') or '',
            'smart': bool(item.get('smart')),
        })
    return rows


@app.route('/api/public/order-preview', methods=['POST'])
def public_order_preview():
    if not valid_public_order_request():
        return jsonify({'error': '下单链接已失效，请联系商家重新获取'}), 401
    data = request.json or {}
    customer_id = str(data.get('customerId') or '').strip()
    text = str(data.get('text') or '').strip()
    if not customer_id:
        return jsonify({'error': '请选择酒店'}), 400
    if not text:
        return jsonify({'error': '请填写下单清单'}), 400
    customer = db.query_one('SELECT * FROM customers WHERE id = ?', (customer_id,))
    if not customer:
        return jsonify({'error': '酒店不存在，请刷新页面后重试'}), 404

    items, unparsed, corrected_text = parse_public_order_items(customer['name'], text)
    if not items:
        return jsonify({'error': '没有识别到“物品+数量+单位”的订单内容', 'unparsed': unparsed}), 400
    return jsonify({
        'customer': customer['name'],
        'correctedText': corrected_text,
        'items': public_order_item_preview(items),
        'unparsed': unparsed,
        'warnings': ['请核对品名、数量和单位，确认无误后再提交。'],
    })


def get_public_customer():
    customer_id = str(request.args.get('customerId') or '').strip()
    if request.is_json:
        body = request.get_json(silent=True) or {}
        customer_id = customer_id or str(body.get('customerId') or '').strip()
    if not customer_id:
        return None, ({'error': '请选择酒店'}, 400)
    customer = db.query_one('SELECT * FROM customers WHERE id = ?', (customer_id,))
    if not customer:
        return None, ({'error': '酒店不存在，请刷新页面后重试'}, 404)
    return customer, None


def public_history_range(mode, value):
    mode = str(mode or 'day').strip()
    value = str(value or date.today().strftime('%Y-%m-%d')).strip()
    if mode == 'year':
        year = re.sub(r'\D', '', value)[:4] or date.today().strftime('%Y')
        return f'{year}-01-01', f'{year}-12-31', year
    if mode == 'month':
        month = value[:7] if re.match(r'^\d{4}-\d{2}', value) else date.today().strftime('%Y-%m')
        year_num, month_num = (int(part) for part in month.split('-'))
        if month_num == 12:
            next_month = date(year_num + 1, 1, 1)
        else:
            next_month = date(year_num, month_num + 1, 1)
        from datetime import timedelta
        end = next_month - timedelta(days=1)
        return f'{month}-01', end.strftime('%Y-%m-%d'), month
    day = normalize_date(value or date.today().strftime('%Y-%m-%d'))
    return day, day, day


def public_customer_sales(customer_name, start_date, end_date):
    rows = db.query('''
        SELECT *
        FROM sales
        WHERE customer = ?
          AND date >= ?
          AND date <= ?
          AND purchase_id = ?
        ORDER BY date DESC, created_at DESC, no DESC
    ''', (customer_name, start_date, end_date, HOTEL_FLOW_PURCHASE_ID))
    sales = attach_items_to_orders(rows, 'sales_items', 'sale_id')
    for sale in sales:
        for item in sale.get('items') or []:
            name = item.get('product_name') or ''
            item['category'] = hotel_item_category(name)
            item['purchase_type'] = public_purchase_type_label(name)
    return sales


@app.route('/api/public/order-history', methods=['GET'])
def public_order_history():
    if not valid_public_order_request():
        return jsonify({'error': '下单链接已失效，请联系商家重新获取'}), 401
    customer, error = get_public_customer()
    if error:
        payload, status = error
        return jsonify(payload), status
    start_date, end_date, label = public_history_range(request.args.get('mode'), request.args.get('value'))
    sales = public_customer_sales(customer['name'], start_date, end_date)
    total = sum(as_float(sale.get('total')) for sale in sales)
    item_count = sum(len(sale.get('items') or []) for sale in sales)
    return jsonify({
        'customer': customer['name'],
        'label': label,
        'startDate': start_date,
        'endDate': end_date,
        'orderCount': len(sales),
        'itemCount': item_count,
        'total': clean_number(total, 2),
        'sales': sales,
    })


@app.route('/api/public/last-order', methods=['GET'])
def public_last_order():
    if not valid_public_order_request():
        return jsonify({'error': '下单链接已失效，请联系商家重新获取'}), 401
    customer, error = get_public_customer()
    if error:
        payload, status = error
        return jsonify(payload), status
    latest = db.query_one('''
        SELECT MAX(date) AS latest_date
        FROM sales
        WHERE customer = ? AND purchase_id = ?
    ''', (customer['name'], HOTEL_FLOW_PURCHASE_ID))
    latest_date = latest.get('latest_date') if latest else ''
    if not latest_date:
        return jsonify({'error': '该酒店暂无历史下单记录'}), 404
    sales = public_customer_sales(customer['name'], latest_date, latest_date)
    lines = []
    for sale in sorted(sales, key=lambda row: (row.get('created_at') or '', row.get('no') or '')):
        for item in sale.get('items') or []:
            name = item.get('product_name') or ''
            quantity = clean_number(item.get('quantity'))
            unit = item.get('product_unit') or ''
            if name and quantity:
                lines.append(f'{name}{quantity}{unit}')
    return jsonify({
        'customer': customer['name'],
        'date': latest_date,
        'text': '，'.join(lines),
        'sales': sales,
    })


@app.route('/api/public/order-history/export', methods=['GET'])
def public_order_history_export():
    if not valid_public_order_request():
        return jsonify({'error': '下单链接已失效，请联系商家重新获取'}), 401
    customer, error = get_public_customer()
    if error:
        payload, status = error
        return jsonify(payload), status
    start_date, end_date, label = public_history_range(request.args.get('mode'), request.args.get('value'))
    sales = public_customer_sales(customer['name'], start_date, end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = '下单记录'
    ws.append([f"{customer['name']}下单记录 {label}"])
    ws.append(['日期', '销售单号', '品名', '规格', '数量', '单位', '单价', '金额'])
    ws.merge_cells('A1:H1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    thin = Side(style='thin', color='666666')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='F2F2F2')
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for sale in sales:
        for item in sale.get('items') or []:
            ws.append([
                sale.get('date') or '',
                sale.get('no') or '',
                item.get('product_name') or '',
                item.get('product_spec') or '',
                clean_number(item.get('quantity')),
                item.get('product_unit') or '',
                as_float(item.get('price')),
                as_float(item.get('subtotal')),
            ])

    total_row = ws.max_row + 1
    ws.append(['合计', '', '', '', '', '', '', sum(as_float(s.get('total')) for s in sales)])

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if cell.column != 3 else 'left', vertical='center')
            if cell.column in (7, 8):
                cell.number_format = '0.00'
    ws.cell(total_row, 1).font = Font(bold=True)
    ws.cell(total_row, 8).font = Font(bold=True)
    ws.cell(total_row, 8).number_format = '0.00'

    widths = [14, 18, 22, 12, 10, 10, 10, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe_customer = re.sub(r'[\\/:*?"<>|]+', '', customer['name'])
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{safe_customer}下单记录_{label}.xlsx'
    )


# ---------- 物品管理 ----------

@app.route('/api/products', methods=['GET'])
def get_products():
    """获取所有物品"""
    rows = db.query('SELECT * FROM products ORDER BY created_at DESC')
    for row in rows:
        row['departments'] = json.loads(row.get('departments', '[]') or '[]')
        row['stock'] = clean_number(row.get('stock'))
        row['cost'] = clean_number(row.get('cost'))
        row['alert_line'] = clean_number(row.get('alert_line'))
        row['last_sale_price'] = clean_number(row.get('last_sale_price'))
        row['is_flow_item'] = is_flow_item(row)
        row['category'] = purchase_category_label(hotel_item_category(row.get('name') or ''))
        row['purchase_type'] = public_purchase_type_label(row.get('name') or '')
        if row['is_flow_item']:
            row['stock'] = 0
    return jsonify(rows)


@app.route('/api/products', methods=['POST'])
def create_or_update_product():
    """创建或更新物品"""
    data = request.json
    product_id = data.get('id')

    code = data.get('code', '').strip()
    name = data.get('name', '').strip()
    spec = data.get('spec', '').strip()
    unit = data.get('unit', '').strip()
    cost = float(data.get('cost', 0) or 0)
    stock = float(data.get('stock', 0) or 0)
    alert_line = float(data.get('alertLine', 10) or 10)
    departments = json.dumps(data.get('departments', []), ensure_ascii=False)
    last_sale_price = float(data.get('lastSalePrice', 0) or 0)
    if code.startswith('WP') or code.startswith('IMP'):
        stock = 0

    if not code or not name or not unit:
        return jsonify({'error': '编号、名称、单位不能为空'}), 400

    # 检查编号是否重复
    existing = db.query_one('SELECT id FROM products WHERE code = ?', (code,))
    if existing and existing['id'] != product_id:
        return jsonify({'error': '物品编号已存在'}), 400

    if product_id:
        # 更新
        db.execute('''
            UPDATE products SET
                code = ?, name = ?, spec = ?, unit = ?, cost = ?,
                stock = ?, alert_line = ?, departments = ?, last_sale_price = ?
            WHERE id = ?
        ''', (code, name, spec, unit, cost, stock, alert_line, departments, last_sale_price, product_id))
        return jsonify({'id': product_id, 'message': '更新成功'})
    else:
        # 创建
        new_id = generate_id()
        db.execute('''
            INSERT INTO products (id, code, name, spec, unit, cost, stock, alert_line, departments, last_sale_price, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (new_id, code, name, spec, unit, cost, stock, alert_line, departments, last_sale_price, now_str()))
        return jsonify({'id': new_id, 'message': '创建成功'})


@app.route('/api/products/<product_id>', methods=['DELETE'])
def delete_product(product_id):
    """删除物品"""
    # 检查是否被引用
    refs = db.query('SELECT 1 FROM purchase_items WHERE product_id = ? LIMIT 1', (product_id,))
    if refs:
        return jsonify({'error': '该物品已被采购单引用，无法删除'}), 400
    refs = db.query('SELECT 1 FROM outbound_items WHERE product_id = ? LIMIT 1', (product_id,))
    if refs:
        return jsonify({'error': '该物品已被领用单引用，无法删除'}), 400
    refs = db.query('SELECT 1 FROM sales_items WHERE product_id = ? LIMIT 1', (product_id,))
    if refs:
        return jsonify({'error': '该物品已被销售单引用，无法删除'}), 400

    db.execute('DELETE FROM products WHERE id = ?', (product_id,))
    return jsonify({'message': '删除成功'})


# ---------- 供应商管理 ----------

@app.route('/api/suppliers', methods=['GET'])
def get_suppliers():
    rows = db.query('SELECT * FROM suppliers ORDER BY created_at DESC')
    return jsonify(rows)


@app.route('/api/suppliers', methods=['POST'])
def create_or_update_supplier():
    data = request.json
    sid = data.get('id')
    code = data.get('code', '').strip()
    name = data.get('name', '').strip()
    contact = data.get('contact', '').strip()
    phone = data.get('phone', '').strip()
    address = data.get('address', '').strip()
    note = data.get('note', '').strip()

    if not code or not name:
        return jsonify({'error': '编号和名称不能为空'}), 400

    existing = db.query_one('SELECT id FROM suppliers WHERE code = ?', (code,))
    if existing and existing['id'] != sid:
        return jsonify({'error': '供应商编号已存在'}), 400

    if sid:
        db.execute('''
            UPDATE suppliers SET code = ?, name = ?, contact = ?, phone = ?, address = ?, note = ?
            WHERE id = ?
        ''', (code, name, contact, phone, address, note, sid))
        return jsonify({'id': sid, 'message': '更新成功'})
    else:
        new_id = generate_id()
        db.execute('''
            INSERT INTO suppliers (id, code, name, contact, phone, address, note, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (new_id, code, name, contact, phone, address, note, now_str()))
        return jsonify({'id': new_id, 'message': '创建成功'})


@app.route('/api/suppliers/<sid>', methods=['DELETE'])
def delete_supplier(sid):
    db.execute('DELETE FROM suppliers WHERE id = ?', (sid,))
    return jsonify({'message': '删除成功'})


# ---------- 客户管理 ----------

@app.route('/api/customers', methods=['GET'])
def get_customers():
    rows = db.query('SELECT * FROM customers ORDER BY created_at DESC')
    return jsonify(rows)


@app.route('/api/customers', methods=['POST'])
def create_or_update_customer():
    data = request.json
    cid = data.get('id')
    code = data.get('code', '').strip()
    name = data.get('name', '').strip()
    company = data.get('company', '').strip()
    phone = data.get('phone', '').strip()
    address = data.get('address', '').strip()
    note = data.get('note', '').strip()

    if not code or not name:
        return jsonify({'error': '编号和名称不能为空'}), 400

    existing = db.query_one('SELECT id FROM customers WHERE code = ?', (code,))
    if existing and existing['id'] != cid:
        return jsonify({'error': '客户编号已存在'}), 400

    if cid:
        db.execute('''
            UPDATE customers SET code = ?, name = ?, company = ?, phone = ?, address = ?, note = ?
            WHERE id = ?
        ''', (code, name, company, phone, address, note, cid))
        return jsonify({'id': cid, 'message': '更新成功'})
    else:
        new_id = generate_id()
        db.execute('''
            INSERT INTO customers (id, code, name, company, phone, address, note, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (new_id, code, name, company, phone, address, note, now_str()))
        return jsonify({'id': new_id, 'message': '创建成功'})


@app.route('/api/customers/<cid>', methods=['DELETE'])
def delete_customer(cid):
    db.execute('DELETE FROM customers WHERE id = ?', (cid,))
    return jsonify({'message': '删除成功'})


# ---------- 采购入库 ----------

@app.route('/api/purchases', methods=['GET'])
def get_purchases():
    rows = db.query('SELECT * FROM purchases ORDER BY date DESC, created_at DESC')
    return jsonify(attach_items_to_orders(rows, 'purchase_items', 'purchase_id'))


@app.route('/api/purchases', methods=['POST'])
def create_or_update_purchase():
    """创建或更新采购单，同时处理库存和自动生成出库单"""
    data = request.json
    purchase_id = data.get('id')
    supplier_id = data.get('supplierId', '')
    supplier_name = data.get('supplierName', '')
    date_str = data.get('date', '')
    items = data.get('items', [])
    auto_outbound = data.get('autoOutbound', False)
    auto_sales = data.get('autoSales', False)

    if not date_str or not items:
        return jsonify({'error': '日期和物品不能为空'}), 400

    # 计算总金额
    total = sum(item.get('quantity', 0) * item.get('price', 0) for item in items)

    conn = db._get_conn()
    try:
        if purchase_id:
            # 编辑模式：先回退原库存
            old_items = db.query('SELECT product_id, quantity FROM purchase_items WHERE purchase_id = ?', (purchase_id,))
            for old in old_items:
                conn.execute("UPDATE products SET stock = stock - ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')", (old['quantity'], old['product_id']))

            # 删除旧明细
            conn.execute('DELETE FROM purchase_items WHERE purchase_id = ?', (purchase_id,))
            # 更新主表
            conn.execute('''
                UPDATE purchases SET supplier_id = ?, supplier_name = ?, date = ?, total = ?
                WHERE id = ?
            ''', (supplier_id, supplier_name, date_str, total, purchase_id))
            pid = purchase_id
        else:
            # 创建新模式
            pid = generate_id()
            no = generate_no('CG')
            conn.execute('''
                INSERT INTO purchases (id, no, supplier_id, supplier_name, date, total, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (pid, no, supplier_id, supplier_name, date_str, total, now_str()))

        # 插入新明细，更新库存和单价
        for item in items:
            product_id = item.get('productId')
            quantity = float(item.get('quantity', 0))
            price = float(item.get('price', 0))
            subtotal = quantity * price

            # 获取产品信息用于快照
            prod = db.query_one('SELECT name, spec, unit FROM products WHERE id = ?', (product_id,))

            conn.execute('''
                INSERT INTO purchase_items (id, purchase_id, product_id, product_name, product_spec, product_unit, quantity, price, subtotal)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (generate_id(), pid, product_id, prod['name'] if prod else '', prod['spec'] if prod else '', prod['unit'] if prod else '', quantity, price, subtotal))

            # 更新库存
            conn.execute("UPDATE products SET stock = stock + ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')", (quantity, product_id))
            # 如果单价为0，自动更新
            conn.execute('UPDATE products SET cost = ? WHERE id = ? AND cost = 0 AND ? > 0', (price, product_id, price))

        # 自动生成部门出库单
        outbound_result = None
        if auto_outbound:
            dept = data.get('autoOutboundDept', '')
            person = data.get('autoOutboundPerson', '')
            outbound_prices = data.get('outboundPrices', {})

            if dept:
                oid = generate_id()
                ono = generate_no('LY')
                outbound_total = 0
                outbound_items_data = []

                for item in items:
                    product_id = item.get('productId')
                    quantity = float(item.get('quantity', 0))
                    # 使用设置的出库单价，否则使用采购价
                    price_key = f"outboundPrice_{items.index(item)}"
                    price = float(outbound_prices.get(price_key, item.get('price', 0)) or item.get('price', 0))
                    subtotal = quantity * price
                    outbound_total += subtotal

                    prod = db.query_one('SELECT name, spec, unit FROM products WHERE id = ?', (product_id,))
                    outbound_items_data.append({
                        'id': generate_id(), 'oid': oid, 'product_id': product_id,
                        'name': prod['name'] if prod else '', 'spec': prod['spec'] if prod else '', 'unit': prod['unit'] if prod else '',
                        'quantity': quantity, 'price': price, 'subtotal': subtotal
                    })

                conn.execute('''
                    INSERT INTO outbounds (id, no, department, person, date, total, purchase_id, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (oid, ono, dept, person, date_str, outbound_total, pid, now_str()))

                for oi in outbound_items_data:
                    conn.execute('''
                        INSERT INTO outbound_items (id, outbound_id, product_id, product_name, product_spec, product_unit, quantity, price, subtotal)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (oi['id'], oi['oid'], oi['product_id'], oi['name'], oi['spec'], oi['unit'], oi['quantity'], oi['price'], oi['subtotal']))
                    # 扣减库存
                    conn.execute("UPDATE products SET stock = stock - ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')", (oi['quantity'], oi['product_id']))

                outbound_result = {'no': ono, 'dept': dept}

        # 自动生成销售出库单
        sales_result = None
        if auto_sales:
            customer = data.get('autoSalesCustomer', '')
            company = data.get('autoSalesCompany', '')
            phone = data.get('autoSalesPhone', '')
            sales_date = data.get('autoSalesDate', date_str)
            sales_prices = data.get('salesPrices', {})

            if customer:
                sid = generate_id()
                sno = generate_no('XS')
                sales_total = 0
                sales_items_data = []

                for item in items:
                    product_id = item.get('productId')
                    quantity = float(item.get('quantity', 0))
                    price_key = f"salesPrice_{items.index(item)}"
                    price = float(sales_prices.get(price_key, item.get('price', 0)) or item.get('price', 0))
                    subtotal = quantity * price
                    sales_total += subtotal

                    prod = db.query_one('SELECT name, spec, unit FROM products WHERE id = ?', (product_id,))
                    sales_items_data.append({
                        'id': generate_id(), 'sid': sid, 'product_id': product_id,
                        'name': prod['name'] if prod else '', 'spec': prod['spec'] if prod else '', 'unit': prod['unit'] if prod else '',
                        'quantity': quantity, 'price': price, 'subtotal': subtotal
                    })

                conn.execute('''
                    INSERT INTO sales (id, no, company, customer, phone, date, total, show_handlers, purchase_id, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
                ''', (sid, sno, company, customer, phone, sales_date, sales_total, pid, now_str()))

                for si in sales_items_data:
                    conn.execute('''
                        INSERT INTO sales_items (id, sale_id, product_id, product_name, product_spec, product_unit, quantity, price, subtotal, note)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (si['id'], si['sid'], si['product_id'], si['name'], si['spec'], si['unit'], si['quantity'], si['price'], si['subtotal'], si.get('note', '')))

                    # 扣减库存
                    conn.execute("UPDATE products SET stock = stock - ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')", (si['quantity'], si['product_id']))
                    # 更新最后销售价
                    conn.execute('UPDATE products SET last_sale_price = ? WHERE id = ?', (si['price'], si['product_id']))
                    if si['price'] > 0:
                        upsert_customer_product_price(
                            conn,
                            customer,
                            {'id': si['product_id'], 'name': si['name'], 'unit': si['unit'], 'spec': si['spec']},
                            si['unit'],
                            si['price'],
                            source='purchase_auto_sale'
                        )

                sales_result = {'no': sno}

        conn.commit()

        result = {'id': pid, 'message': '保存成功'}
        if outbound_result:
            result['outbound'] = outbound_result
        if sales_result:
            result['sales'] = sales_result
        return jsonify(result)

    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/purchases/<purchase_id>', methods=['DELETE'])
def delete_purchase(purchase_id):
    """删除采购单，回退库存，同时删除关联的出库单和销售单"""
    conn = db._get_conn()
    try:
        # 回退采购入库库存
        items = db.query('SELECT product_id, quantity FROM purchase_items WHERE purchase_id = ?', (purchase_id,))
        for item in items:
            conn.execute("UPDATE products SET stock = stock - ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')", (item['quantity'], item['product_id']))

        # 回退关联的部门出库单库存并删除
        outbounds = db.query('SELECT id FROM outbounds WHERE purchase_id = ?', (purchase_id,))
        for ob in outbounds:
            ob_items = db.query('SELECT product_id, quantity FROM outbound_items WHERE outbound_id = ?', (ob['id'],))
            for obi in ob_items:
                conn.execute("UPDATE products SET stock = stock + ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')", (obi['quantity'], obi['product_id']))
            conn.execute('DELETE FROM outbound_items WHERE outbound_id = ?', (ob['id'],))
            conn.execute('DELETE FROM outbounds WHERE id = ?', (ob['id'],))

        # 回退关联的销售单库存并删除
        sales = db.query('SELECT id FROM sales WHERE purchase_id = ?', (purchase_id,))
        for s in sales:
            s_items = db.query('SELECT product_id, quantity FROM sales_items WHERE sale_id = ?', (s['id'],))
            for si in s_items:
                conn.execute("UPDATE products SET stock = stock + ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')", (si['quantity'], si['product_id']))
            conn.execute('DELETE FROM sales_items WHERE sale_id = ?', (s['id'],))
            conn.execute('DELETE FROM sales WHERE id = ?', (s['id'],))

        conn.execute('DELETE FROM purchase_items WHERE purchase_id = ?', (purchase_id,))
        conn.execute('DELETE FROM purchases WHERE id = ?', (purchase_id,))
        conn.commit()
        return jsonify({'message': '删除成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ---------- 部门领用 ----------

@app.route('/api/outbounds', methods=['GET'])
def get_outbounds():
    dept = request.args.get('dept', '')
    sql = 'SELECT * FROM outbounds'
    params = []
    if dept:
        sql += ' WHERE department = ?'
        params.append(dept)
    sql += ' ORDER BY date DESC, created_at DESC'

    rows = db.query(sql, tuple(params))
    return jsonify(attach_items_to_orders(rows, 'outbound_items', 'outbound_id'))


@app.route('/api/outbounds', methods=['POST'])
def create_or_update_outbound():
    data = request.json
    oid = data.get('id')
    department = data.get('department', '')
    person = data.get('person', '')
    date_str = data.get('date', '')
    items = data.get('items', [])

    if not department or not date_str or not items:
        return jsonify({'error': '部门、日期和物品不能为空'}), 400

    total = sum(item.get('quantity', 0) * item.get('price', 0) for item in items)

    conn = db._get_conn()
    try:
        if oid:
            # 编辑模式：回退原库存
            old_items = db.query('SELECT product_id, quantity FROM outbound_items WHERE outbound_id = ?', (oid,))
            for old in old_items:
                conn.execute("UPDATE products SET stock = stock + ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')", (old['quantity'], old['product_id']))

            conn.execute('DELETE FROM outbound_items WHERE outbound_id = ?', (oid,))
            conn.execute('''
                UPDATE outbounds SET department = ?, person = ?, date = ?, total = ?
                WHERE id = ?
            ''', (department, person, date_str, total, oid))
        else:
            oid = generate_id()
            ono = generate_no('LY')
            conn.execute('''
                INSERT INTO outbounds (id, no, department, person, date, total, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (oid, ono, department, person, date_str, total, now_str()))

        for item in items:
            product_id = item.get('productId')
            quantity = float(item.get('quantity', 0))
            price = float(item.get('price', 0))
            subtotal = quantity * price

            prod = db.query_one('SELECT name, spec, unit FROM products WHERE id = ?', (product_id,))

            conn.execute('''
                INSERT INTO outbound_items (id, outbound_id, product_id, product_name, product_spec, product_unit, quantity, price, subtotal)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (generate_id(), oid, product_id, prod['name'] if prod else '', prod['spec'] if prod else '', prod['unit'] if prod else '', quantity, price, subtotal))

            conn.execute("UPDATE products SET stock = stock - ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')", (quantity, product_id))

        conn.commit()
        return jsonify({'id': oid, 'message': '保存成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/outbounds/<oid>', methods=['DELETE'])
def delete_outbound(oid):
    conn = db._get_conn()
    try:
        items = db.query('SELECT product_id, quantity FROM outbound_items WHERE outbound_id = ?', (oid,))
        for item in items:
            conn.execute("UPDATE products SET stock = stock + ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')", (item['quantity'], item['product_id']))

        conn.execute('DELETE FROM outbounds WHERE id = ?', (oid,))
        conn.commit()
        return jsonify({'message': '删除成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ---------- 销售出库 ----------

@app.route('/api/sales', methods=['GET'])
def get_sales():
    if request.args.get('summary') in ('1', 'true', 'yes'):
        return jsonify(db.query('''
            SELECT
                s.*,
                COUNT(si.id) AS item_count
            FROM sales s
            LEFT JOIN sales_items si ON si.sale_id = s.id
            GROUP BY s.id
            ORDER BY s.date DESC, s.created_at DESC
        '''))
    rows = db.query('SELECT * FROM sales ORDER BY date DESC, created_at DESC')
    return jsonify(attach_items_to_orders(rows, 'sales_items', 'sale_id'))


@app.route('/api/sales/<sid>', methods=['GET'])
def get_sale(sid):
    rows = db.query('SELECT * FROM sales WHERE id = ?', (sid,))
    if not rows:
        return jsonify({'error': '销售单不存在'}), 404
    return jsonify(attach_items_to_orders(rows, 'sales_items', 'sale_id')[0])


def build_sales_purchase_list_workbook(ids=None, sale_date=None):
    ids = [str(item) for item in (ids or []) if item]
    title_date = ''
    if ids:
        placeholders = ','.join('?' for _ in ids)
        where_sql = f's.id IN ({placeholders})'
        params = tuple(ids)
        date_rows = db.query(f'SELECT MIN(date) AS min_date, MAX(date) AS max_date FROM sales s WHERE {where_sql}', params)
        if date_rows and date_rows[0].get('min_date') == date_rows[0].get('max_date'):
            title_date = date_rows[0].get('min_date') or date.today().strftime('%Y-%m-%d')
        else:
            title_date = date.today().strftime('%Y-%m-%d')
    else:
        if not sale_date:
            raise ValueError('缺少销售日期')
        date_str = normalize_date(sale_date)
        where_sql = 's.date = ?'
        params = (date_str,)
        title_date = date_str

    purchase_group_sql = '''
                CASE
                    WHEN (COALESCE(s.customer, '') LIKE '%美居%' OR COALESCE(s.customer, '') LIKE '%浐灞%')
                         AND COALESCE(s.order_key, '') = 'front_fruit'
                    THEN 'front_fruit'
                    WHEN COALESCE(s.customer, '') LIKE '%全季%'
                         AND COALESCE(s.order_key, '') = 'fruit'
                    THEN 'fruit'
                    ELSE ''
                END
    '''

    items = db.query(f'''
        SELECT name, unit, purchase_group, SUM(quantity) AS quantity, GROUP_CONCAT(DISTINCT NULLIF(note, '')) AS note, MIN(first_row) AS first_row
        FROM (
            SELECT
                COALESCE(NULLIF(si.product_name, ''), p.name, '') AS name,
                COALESCE(NULLIF(si.product_unit, ''), p.unit, '') AS unit,
                {purchase_group_sql} AS purchase_group,
                COALESCE(si.quantity, 0) AS quantity,
                COALESCE(si.note, '') AS note,
                si.rowid AS first_row
            FROM sales_items si
            JOIN sales s ON si.sale_id = s.id
            LEFT JOIN products p ON si.product_id = p.id
            WHERE {where_sql}
        ) rows
        GROUP BY name, unit, purchase_group
        HAVING quantity > 0
        ORDER BY first_row
    ''', params)
    if not items:
        raise LookupError('没有可生成采购单的销售明细')

    if ids and len(ids) == 1:
        sale_row = db.query_one('''
            SELECT COALESCE(NULLIF(customer, ''), company, '未填写客户') AS customer
            FROM sales
            WHERE id = ?
        ''', (ids[0],))
        customer_name = sale_row.get('customer') if sale_row else '未填写客户'
        title = f'{customer_name}采购单{title_date}'
        return build_purchase_list_workbook(items, title), title_date

    customer_items = db.query(f'''
        SELECT name, unit, purchase_group, customer, SUM(quantity) AS quantity, GROUP_CONCAT(DISTINCT NULLIF(note, '')) AS note, MIN(first_sale) AS first_sale
        FROM (
            SELECT
                COALESCE(NULLIF(si.product_name, ''), p.name, '') AS name,
                COALESCE(NULLIF(si.product_unit, ''), p.unit, '') AS unit,
                {purchase_group_sql} AS purchase_group,
                COALESCE(NULLIF(s.customer, ''), '未填写客户') AS customer,
                COALESCE(si.quantity, 0) AS quantity,
                COALESCE(si.note, '') AS note,
                s.created_at AS first_sale
            FROM sales_items si
            JOIN sales s ON si.sale_id = s.id
            LEFT JOIN products p ON si.product_id = p.id
            WHERE {where_sql}
        ) rows
        GROUP BY name, unit, purchase_group, customer
        HAVING quantity > 0
        ORDER BY first_sale, customer
    ''', params)
    customers = [
        row['customer']
        for row in db.query(f'''
            SELECT COALESCE(NULLIF(customer, ''), '未填写客户') AS customer, MIN(created_at) AS first_sale
            FROM sales s
            WHERE {where_sql}
            GROUP BY customer
            ORDER BY first_sale, customer
        ''', params)
    ]
    customers.sort(key=hotel_customer_export_rank)
    title = f'酒店合并采购单{title_date}'
    return build_customer_split_purchase_workbook(items, customer_items, customers, title), title_date


@app.route('/api/sales/purchase-list', methods=['GET', 'POST'])
def download_sales_purchase_list():
    """按销售单汇总生成供应商采购填表单，支持按日期或勾选销售单。"""
    if request.method == 'POST':
        data = request.json or {}
        ids = [str(item) for item in data.get('ids') or [] if item]
        if not ids:
            return jsonify({'error': '请先勾选要合并的销售单'}), 400
        sale_date = None
    else:
        ids = []
        sale_date = request.args.get('date') or ''
    try:
        wb, title_date = build_sales_purchase_list_workbook(ids=ids, sale_date=sale_date)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except LookupError as e:
        return jsonify({'error': str(e)}), 404
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'采购单_{title_date}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def safe_sheet_title(title, index=1):
    text = re.sub(r'[:\\/?*\[\]]', '', str(title or '销售单')).strip() or '销售单'
    suffix = f'-{index}' if index > 1 else ''
    return (text[:31 - len(suffix)] + suffix)[:31]


def copy_row_style(ws, source_row, target_row):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, 9):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        target._style = copy(source._style)
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)


def sales_total_formula(total_cell):
    return (
        f'=IF(ROUND({total_cell},2)<0,"无效数值",'
        f'IF(ROUND({total_cell},2)=0,"零元整",'
        f'IF(ROUND({total_cell},2)>=1,TEXT(INT(ROUND({total_cell},2)),"[DBNum2]")&"元","")&'
        f'SUBSTITUTE(SUBSTITUTE(TEXT(RIGHT(DOLLAR({total_cell},2),2),"[DBNum2]0角0分;;整"),"零角","零"),"零分","整")))'
    )


def configure_sales_export_print_settings(ws, paper='half', total_row=25):
    ws.print_area = f'A1:H{total_row}'
    ws.page_margins.left = 0.18
    ws.page_margins.right = 0.18
    ws.page_margins.top = 0.14
    ws.page_margins.bottom = 0.14
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    for row_idx in range(total_row + 1, 26):
        ws.row_dimensions[row_idx].hidden = True


def fill_sales_export_sheet(ws, sale, items=None, sheet_title_index=1, start_index=1, page_text='', paper='half', dynamic_rows=False):
    title = sale.get('company') or '西安禾润佳商贸有限公司'
    customer = sale.get('customer') or ''
    phone = sale.get('phone') or ''
    date_text = normalize_date(sale.get('date'))
    items = items if items is not None else (sale.get('items') or [])
    item_start_row = 5
    base_item_rows = 12 if paper == 'third' else 20
    template_item_rows = max(base_item_rows, len(items)) if dynamic_rows else base_item_rows
    total_row = item_start_row + template_item_rows

    if total_row > 25:
        extra_rows = total_row - 25
        ws.insert_rows(25, extra_rows)
        for row_idx in range(25, 25 + extra_rows):
            copy_row_style(ws, 24, row_idx)

    ws.title = safe_sheet_title(sale.get('sheet_title') or sale.get('no') or sale.get('date') or '销售单', sheet_title_index)
    ws['A1'] = title
    ws['A3'] = f'客户：{customer}'
    ws['E3'] = f'  电话：{phone}'
    ws['G3'] = f'日期：{date_text}{page_text}'

    if total_row != 25 and total_row < 25:
        copy_row_style(ws, 25, total_row)
        for row_idx in range(total_row + 1, 26):
            for col_idx in range(1, 9):
                ws.cell(row_idx, col_idx).value = None

    last_item_row = item_start_row + template_item_rows - 1
    for row_idx in range(item_start_row, last_item_row + 1):
        row_item = items[row_idx - item_start_row] if row_idx - item_start_row < len(items) else None
        values = [
            start_index + row_idx - item_start_row if row_item else None,
            row_item.get('product_name') if row_item else None,
            row_item.get('product_spec') if row_item else None,
            clean_number(row_item.get('quantity')) if row_item else None,
            row_item.get('product_unit') if row_item else None,
            as_float(row_item.get('price')) if row_item else None,
            f'=D{row_idx}*F{row_idx}' if row_item else None,
            purchase_export_note(row_item.get('product_name') or '', row_item.get('note') or '') if row_item else None,
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx).value = value
        ws.cell(row_idx, 6).number_format = '0.00'
        ws.cell(row_idx, 7).number_format = '0.00'

    ws.cell(total_row, 1).value = '合计：'
    for rng in list(ws.merged_cells.ranges):
        is_total_merge = (
            rng.min_row == rng.max_row
            and ((rng.min_col, rng.max_col) in ((2, 5), (6, 7)))
        )
        if is_total_merge and rng.min_row != total_row:
            ws.unmerge_cells(str(rng))
    existing_merges = [str(rng) for rng in ws.merged_cells.ranges]
    if f'B{total_row}:E{total_row}' not in existing_merges:
        ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=5)
    if f'F{total_row}:G{total_row}' not in existing_merges:
        ws.merge_cells(start_row=total_row, start_column=6, end_row=total_row, end_column=7)
    ws.cell(total_row, 2).value = sales_total_formula(f'H{total_row}')
    ws.cell(total_row, 8).value = f'=SUM(G{item_start_row}:G{last_item_row})'
    ws.cell(total_row, 8).number_format = '0.00'

    ws.freeze_panes = 'A4'
    configure_sales_export_print_settings(ws, paper, total_row)


def sales_export_group_key(sale, item):
    customer = sale.get('customer') or '未填写客户'
    name = item.get('product_name') or ''
    category = hotel_item_category(name)
    if '全季' in customer and category == '水果类':
        return (customer, 'fruit', f'{customer}水果')
    if ('美居' in customer or '浐灞' in customer) and sale.get('order_key') == 'front_fruit':
        return (customer, 'front_fruit', f'{customer}前台水果')
    return (customer, 'regular', customer)


def grouped_sales_for_export(sales):
    groups = {}
    for sale in sales:
        sale_items = sale.get('items') or []
        if not sale_items:
            customer = sale.get('customer') or '未填写客户'
            key = (customer, 'regular')
            if key not in groups:
                groups[key] = {
                    **sale,
                    'customer': customer,
                    'items': [],
                    'sheet_title': customer,
                    '_dates': [],
                    '_nos': [],
                }
            continue
        for item in sale_items:
            customer, group_type, sheet_title = sales_export_group_key(sale, item)
            key = (customer, group_type)
            if key not in groups:
                groups[key] = {
                    **sale,
                    'customer': customer,
                    'items': [],
                    'sheet_title': sheet_title,
                    '_dates': [],
                    '_nos': [],
                }
            group = groups[key]
            if sale.get('date') and sale.get('date') not in group['_dates']:
                group['_dates'].append(sale.get('date'))
            if sale.get('no'):
                group['_nos'].append(sale.get('no'))
            group['items'].append(item)
    result = []
    for group in groups.values():
        dates = sorted(group.pop('_dates', []))
        group.pop('_nos', None)
        group['items'] = group.get('items') or []
        if len(dates) == 1:
            group['date'] = dates[0]
        elif len(dates) > 1:
            group['date'] = f'{dates[0]} 至 {dates[-1]}'
        group['no'] = f"{group.get('sheet_title') or group.get('customer') or '客户'}合并销售单"
        result.append(group)
    return result


def build_sales_export_workbook(ids, paper='half', group_by_customer=False):
    paper = str(paper or 'half').strip()
    if paper not in ('half', 'third'):
        paper = 'half'
    if not ids:
        raise ValueError('请先勾选要导出的销售单')
    ids = [str(item) for item in ids if item]
    placeholders = ','.join('?' for _ in ids)
    sales = db.query(f'SELECT * FROM sales WHERE id IN ({placeholders}) ORDER BY date ASC, no ASC', tuple(ids))
    if not sales:
        raise LookupError('没有找到要导出的销售单')
    sales = attach_items_to_orders(sales, 'sales_items', 'sale_id')

    if os.path.exists(SALES_EXPORT_TEMPLATE):
        wb = load_workbook(SALES_EXPORT_TEMPLATE)
    else:
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('A1:H2')
        ws.merge_cells('A3:C3')
        ws.merge_cells('G3:H3')
        ws.merge_cells('B25:E25')
        ws.merge_cells('F25:G25')
        ws.append([])

    template_ws = wb.active
    sheet_index = 1
    rows_per_page = 12 if paper == 'third' else 20
    export_sales = grouped_sales_for_export(sales) if group_by_customer else sales
    for sale in export_sales:
        sale_items = sort_purchase_review_items(sale.get('items') or [])
        chunks = [sale_items[i:i + rows_per_page] for i in range(0, len(sale_items), rows_per_page)] or [[]]
        for page_index, chunk in enumerate(chunks, start=1):
            ws = wb.copy_worksheet(template_ws)
            fill_sales_export_sheet(
                ws,
                sale,
                items=chunk,
                sheet_title_index=sheet_index,
                start_index=(page_index - 1) * rows_per_page + 1,
                page_text='',
                paper=paper,
                dynamic_rows=False,
            )
            sheet_index += 1
    if len(wb.worksheets) > 1:
        wb.remove(template_ws)
    return wb


@app.route('/api/sales/export', methods=['POST'])
def export_sales_xlsx():
    data = request.json or {}
    ids = data.get('ids') or []
    paper = str(data.get('paper') or 'half').strip()
    group_by_customer = bool(data.get('groupByCustomer'))
    try:
        wb = build_sales_export_workbook(ids, paper, group_by_customer=group_by_customer)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except LookupError as e:
        return jsonify({'error': str(e)}), 404

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    paper_name = '三等分' if paper == 'third' else '二等分'
    filename = f'销售单_{paper_name}_{date.today().strftime("%Y-%m-%d")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/sales', methods=['POST'])
def create_or_update_sale():
    data = request.json
    sid = data.get('id')
    company = data.get('company', '')
    customer = data.get('customer', '')
    phone = data.get('phone', '')
    date_str = data.get('date', '')
    items = data.get('items', [])
    show_handlers = data.get('showHandlers', False)
    handler = data.get('handler', '')
    issuer = data.get('issuer', '')

    if not customer or not date_str or not items:
        return jsonify({'error': '客户、日期和物品不能为空'}), 400

    total = sum(item.get('quantity', 0) * item.get('price', 0) for item in items)

    conn = db._get_conn()
    try:
        is_hotel_flow = False
        if sid:
            old_sale_row = conn.execute('SELECT purchase_id FROM sales WHERE id = ?', (sid,)).fetchone()
            old_sale = dict(old_sale_row) if old_sale_row else None
            is_hotel_flow = old_sale and old_sale.get('purchase_id') == HOTEL_FLOW_PURCHASE_ID
            if is_hotel_flow:
                deductions = conn.execute('SELECT product_id, quantity FROM hotel_stock_deductions WHERE sale_id = ?', (sid,)).fetchall()
                for old in deductions:
                    conn.execute("UPDATE products SET stock = stock + ? WHERE id = ?", (old['quantity'], old['product_id']))
                conn.execute('DELETE FROM hotel_stock_deductions WHERE sale_id = ?', (sid,))
            else:
                old_items = conn.execute('SELECT product_id, quantity FROM sales_items WHERE sale_id = ?', (sid,)).fetchall()
                for old in old_items:
                    conn.execute("UPDATE products SET stock = stock + ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')", (old['quantity'], old['product_id']))

            conn.execute('DELETE FROM sales_items WHERE sale_id = ?', (sid,))
            conn.execute('''
                UPDATE sales SET company = ?, customer = ?, phone = ?, date = ?, total = ?,
                    show_handlers = ?, handler = ?, issuer = ?
                WHERE id = ?
            ''', (company, customer, phone, date_str, total, 1 if show_handlers else 0, handler, issuer, sid))
        else:
            sid = generate_id()
            sno = generate_no('XS')
            conn.execute('''
                INSERT INTO sales (id, no, company, customer, phone, date, total, show_handlers, handler, issuer, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (sid, sno, company, customer, phone, date_str, total, 1 if show_handlers else 0, handler, issuer, now_str()))

        for item in items:
            product_id = item.get('productId')
            quantity = float(item.get('quantity', 0))
            price = float(item.get('price', 0))
            subtotal = quantity * price

            prod_row = conn.execute('SELECT id, name, spec, unit, code, stock FROM products WHERE id = ?', (product_id,)).fetchone()
            prod = dict(prod_row) if prod_row else None
            product_unit = (item.get('productUnit') or item.get('product_unit') or '').strip()
            if not product_unit and prod:
                product_unit = prod.get('unit') or ''
            product_name = prod['name'] if prod else (item.get('productName') or item.get('product_name') or '')
            product_spec = prod['spec'] if prod else (item.get('productSpec') or item.get('product_spec') or '')
            sale_item_id = generate_id()

            conn.execute('''
                INSERT INTO sales_items (id, sale_id, product_id, product_name, product_spec, product_unit, quantity, price, subtotal, note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                sale_item_id,
                sid,
                product_id,
                product_name,
                product_spec,
                product_unit,
                quantity,
                price,
                subtotal,
                item.get('note') or ''
            ))

            if is_hotel_flow:
                full_prod = prod
                if full_prod and not is_flow_item(full_prod) and as_float(full_prod.get('stock')) > 0:
                    conn.execute("UPDATE products SET stock = stock - ? WHERE id = ?", (quantity, product_id))
                    conn.execute('''
                        INSERT INTO hotel_stock_deductions (sale_item_id, sale_id, product_id, quantity)
                        VALUES (?, ?, ?, ?)
                    ''', (sale_item_id, sid, product_id, quantity))
            else:
                conn.execute("UPDATE products SET stock = stock - ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')", (quantity, product_id))
                conn.execute('UPDATE products SET last_sale_price = ? WHERE id = ?', (price, product_id))
            if price > 0 and prod:
                upsert_customer_product_price(
                    conn,
                    customer,
                    {'id': product_id, 'name': prod['name'], 'unit': product_unit, 'spec': prod['spec'] if prod else ''},
                    product_unit,
                    price,
                    source='manual_sale'
                )

        conn.commit()
        feishu_error = safe_push_sale_to_feishu(sid)
        response = {'id': sid, 'message': '保存成功'}
        if feishu_error:
            response['feishuSyncError'] = feishu_error
        return jsonify(response)
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/sales/<sid>', methods=['DELETE'])
def delete_sale(sid):
    conn = db._get_conn()
    try:
        sale = db.query_one('SELECT purchase_id FROM sales WHERE id = ?', (sid,))
        is_hotel_flow = sale and sale.get('purchase_id') == HOTEL_FLOW_PURCHASE_ID
        if is_hotel_flow:
            deductions = db.query('SELECT product_id, quantity FROM hotel_stock_deductions WHERE sale_id = ?', (sid,))
            for item in deductions:
                conn.execute("UPDATE products SET stock = stock + ? WHERE id = ?", (item['quantity'], item['product_id']))
            conn.execute('DELETE FROM hotel_stock_deductions WHERE sale_id = ?', (sid,))
        else:
            items = db.query('SELECT product_id, quantity FROM sales_items WHERE sale_id = ?', (sid,))
            for item in items:
                conn.execute("UPDATE products SET stock = stock + ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')", (item['quantity'], item['product_id']))

        conn.execute('DELETE FROM sales WHERE id = ?', (sid,))
        conn.commit()
        return jsonify({'message': '删除成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ---------- 打印任务 ----------

def notify_print_agents():
    with _print_job_condition:
        _print_job_condition.notify_all()


def print_job_public(row):
    return {
        'id': row.get('id'),
        'type': row.get('job_type'),
        'title': row.get('title') or '',
        'targetPrinter': row.get('target_printer') or '',
        'paper': row.get('paper') or '',
        'status': row.get('status') or '',
        'createdAt': row.get('created_at') or '',
        'downloadUrl': f"/api/print-agent/jobs/{row.get('id')}/download?token={PRINT_AGENT_TOKEN}",
    }


def create_print_job(job_type, title, target_printer, paper, payload):
    job_id = generate_id()
    db.execute('''
        INSERT INTO print_jobs (id, job_type, title, target_printer, paper, payload_json, status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, 'pending', ?)
    ''', (
        job_id,
        job_type,
        title,
        target_printer,
        paper,
        json.dumps(payload or {}, ensure_ascii=False),
        now_str(),
    ))
    notify_print_agents()
    return job_id


@app.route('/api/print-jobs', methods=['POST'])
def enqueue_print_job():
    data = request.json or {}
    job_type = str(data.get('type') or '').strip()
    ids = [str(item) for item in (data.get('ids') or []) if item]
    paper = str(data.get('paper') or 'half').strip()
    if paper not in ('half', 'third', 'a4'):
        paper = 'half'
    if job_type not in ('sales', 'purchase'):
        return jsonify({'error': '打印类型不正确'}), 400
    if not ids:
        return jsonify({'error': '请先选择要打印的销售单'}), 400

    rows = db.query(
        f"SELECT no, customer, date FROM sales WHERE id IN ({','.join('?' for _ in ids)}) ORDER BY date ASC, no ASC",
        tuple(ids),
    )
    if not rows:
        return jsonify({'error': '没有找到销售单'}), 404

    first = rows[0]
    customer_part = first.get('customer') or '酒店'
    if len(rows) == 1:
        title = f"{customer_part}{'销售单' if job_type == 'sales' else '采购单'}_{first.get('date')}"
    else:
        title = f"{first.get('date')}_合并{'销售单' if job_type == 'sales' else '采购单'}_{len(rows)}张"
    target = 'sales_dot_matrix' if job_type == 'sales' else 'purchase_laser'
    if job_type == 'purchase':
        paper = 'a4'
    job_id = create_print_job(job_type, title, target, paper, {'ids': ids})
    return jsonify({'id': job_id, 'message': '已发送到台式机打印队列'})


@app.route('/api/print-jobs', methods=['GET'])
def list_print_jobs():
    rows = db.query('''
        SELECT * FROM print_jobs
        ORDER BY created_at DESC
        LIMIT 50
    ''')
    return jsonify([print_job_public(row) for row in rows])


@app.route('/api/print-agent/events', methods=['GET'])
def print_agent_events():
    if not valid_print_agent_request():
        return jsonify({'error': '打印助手令牌不正确'}), 401

    def stream():
        last_sent = ''
        yield 'event: hello\ndata: {"message":"connected"}\n\n'
        while True:
            jobs = db.query('''
                SELECT * FROM print_jobs
                WHERE status IN ('pending', 'failed')
                ORDER BY created_at ASC
                LIMIT 10
            ''')
            payload = [print_job_public(row) for row in jobs]
            data = json.dumps(payload, ensure_ascii=False)
            if data != last_sent:
                yield f'event: jobs\ndata: {data}\n\n'
                last_sent = data
            with _print_job_condition:
                _print_job_condition.wait(timeout=25)
            yield ': keepalive\n\n'

    return Response(stream(), mimetype='text/event-stream', headers={'Cache-Control': 'no-cache'})


@app.route('/api/print-agent/ping', methods=['GET'])
def print_agent_ping():
    if not valid_print_agent_request():
        return jsonify({'error': '打印助手令牌不正确'}), 401
    return jsonify({'ok': True, 'time': now_str()})


@app.route('/api/print-agent/jobs/<job_id>/download', methods=['GET'])
def download_print_job(job_id):
    if not valid_print_agent_request():
        return jsonify({'error': '打印助手令牌不正确'}), 401
    job = db.query_one('SELECT * FROM print_jobs WHERE id = ?', (job_id,))
    if not job:
        return jsonify({'error': '打印任务不存在'}), 404
    payload = json.loads(job.get('payload_json') or '{}')
    ids = payload.get('ids') or []
    try:
        if job.get('job_type') == 'sales':
            wb = build_sales_export_workbook(ids, job.get('paper') or 'half')
            filename = f"{job.get('title') or '销售单'}.xlsx"
        elif job.get('job_type') == 'purchase':
            wb, _ = build_sales_purchase_list_workbook(ids=ids)
            filename = f"{job.get('title') or '采购单'}.xlsx"
        else:
            return jsonify({'error': '未知打印任务类型'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    db.execute("UPDATE print_jobs SET status = 'printing', claimed_at = ? WHERE id = ?", (now_str(), job_id))
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/print-agent/jobs/<job_id>/complete', methods=['POST'])
def complete_print_job(job_id):
    if not valid_print_agent_request():
        return jsonify({'error': '打印助手令牌不正确'}), 401
    data = request.json or {}
    ok = bool(data.get('ok', True))
    status = 'done' if ok else 'failed'
    error = str(data.get('error') or '')[:500]
    db.execute(
        'UPDATE print_jobs SET status = ?, error = ?, completed_at = ? WHERE id = ?',
        (status, error, now_str(), job_id),
    )
    notify_print_agents()
    return jsonify({'message': '已更新'})


# ---------- 财务记账 ----------

@app.route('/api/finances', methods=['GET'])
def get_finances():
    rows = db.query('SELECT * FROM finances ORDER BY date DESC, created_at DESC')
    return jsonify(rows)


@app.route('/api/finances', methods=['POST'])
def create_or_update_finance():
    data = request.json
    fid = data.get('id')
    date_str = data.get('date', '')
    ftype = data.get('type', '')
    category = data.get('category', '')
    amount = float(data.get('amount', 0))
    note = data.get('note', '')

    if not date_str or not ftype:
        return jsonify({'error': '日期和类型不能为空'}), 400

    if fid:
        db.execute('''
            UPDATE finances SET date = ?, type = ?, category = ?, amount = ?, note = ?
            WHERE id = ?
        ''', (date_str, ftype, category, amount, note, fid))
        return jsonify({'id': fid, 'message': '更新成功'})
    else:
        new_id = generate_id()
        db.execute('''
            INSERT INTO finances (id, date, type, category, amount, note, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (new_id, date_str, ftype, category, amount, note, now_str()))
        return jsonify({'id': new_id, 'message': '创建成功'})


@app.route('/api/finances/<fid>', methods=['DELETE'])
def delete_finance(fid):
    db.execute('DELETE FROM finances WHERE id = ?', (fid,))
    return jsonify({'message': '删除成功'})


# ---------- 首页概览 ----------

@app.route('/api/dashboard', methods=['GET'])
def get_dashboard():
    """酒店后台首页：只看酒店销售、下单频率和高频采购物品。"""
    start, end = get_month_range()
    today = today_str()
    flow_params = (HOTEL_FLOW_PURCHASE_ID, start, end)

    month_sales = db.query_one('''
        SELECT COALESCE(SUM(total), 0) AS total
        FROM sales
        WHERE purchase_id = ? AND date >= ? AND date <= ?
    ''', flow_params)['total']

    month_orders = db.query_one('''
        SELECT COUNT(*) AS count
        FROM sales
        WHERE purchase_id = ? AND date >= ? AND date <= ?
    ''', flow_params)['count']

    today_orders = db.query_one('''
        SELECT COUNT(*) AS count
        FROM sales
        WHERE purchase_id = ? AND date = ?
    ''', (HOTEL_FLOW_PURCHASE_ID, today))['count']

    active_customers = db.query_one('''
        SELECT COUNT(DISTINCT customer) AS count
        FROM sales
        WHERE purchase_id = ? AND date >= ? AND date <= ?
    ''', flow_params)['count']

    hotel_customers = db.query('''
        SELECT name
        FROM customers
        WHERE name LIKE '%酒店%' OR name LIKE '%汉庭%' OR name LIKE '%全季%' OR name LIKE '%美居%'
        ORDER BY created_at, name
    ''')
    customer_names = [row['name'] for row in hotel_customers]

    customer_rows = db.query('''
        SELECT customer,
               COUNT(*) AS month_orders,
               COALESCE(SUM(total), 0) AS month_sales,
               MAX(date) AS last_order_date
        FROM sales
        WHERE purchase_id = ? AND date >= ? AND date <= ?
        GROUP BY customer
    ''', flow_params)
    all_time_last_rows = db.query('''
        SELECT customer, MAX(date) AS last_order_date
        FROM sales
        WHERE purchase_id = ?
        GROUP BY customer
    ''', (HOTEL_FLOW_PURCHASE_ID,))
    for row in customer_rows + all_time_last_rows:
        name = row.get('customer') or ''
        if name and name not in customer_names:
            customer_names.append(name)

    current_by_customer = {row['customer']: row for row in customer_rows}
    last_by_customer = {row['customer']: row.get('last_order_date') for row in all_time_last_rows}
    customer_stats = []
    for name in customer_names:
        row = current_by_customer.get(name, {})
        customer_stats.append({
            'customer': name,
            'monthSales': clean_number(row.get('month_sales'), 2),
            'monthOrders': int(row.get('month_orders') or 0),
            'lastOrderDate': row.get('last_order_date') or last_by_customer.get(name) or '',
        })
    customer_stats.sort(key=lambda row: (row['monthSales'], row['monthOrders'], row['lastOrderDate']), reverse=True)

    top_items = db.query(f'''
        SELECT si.product_name AS name,
               si.product_unit AS unit,
               COALESCE(SUM({EFFECTIVE_SALE_QUANTITY_SQL}), 0) AS quantity,
               COALESCE(SUM(si.subtotal), 0) AS revenue
        FROM sales_items si
        JOIN sales s ON si.sale_id = s.id
        LEFT JOIN products p ON si.product_id = p.id
        WHERE s.purchase_id = ? AND s.date >= ? AND s.date <= ?
        GROUP BY si.product_name, si.product_unit
        ORDER BY quantity DESC, revenue DESC
        LIMIT 10
    ''', flow_params)
    for row in top_items:
        row['quantity'] = clean_number(row.get('quantity'))
        row['revenue'] = clean_number(row.get('revenue'), 2)
        row['type'] = public_purchase_type_label(row.get('name'))

    return jsonify({
        'monthSales': clean_number(month_sales, 2),
        'todayOrders': today_orders,
        'monthOrders': month_orders,
        'activeCustomers': active_customers,
        'customerStats': customer_stats,
        'topItems': top_items,
    })


# ---------- 库存查询 ----------

@app.route('/api/inventory', methods=['GET'])
def get_inventory():
    """获取库存列表"""
    search = request.args.get('search', '').lower()
    status = request.args.get('status', '')

    sql = 'SELECT * FROM products WHERE 1=1'
    params = []

    if search:
        sql += ' AND (LOWER(name) LIKE ? OR LOWER(code) LIKE ? OR LOWER(spec) LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])

    if status:
        if status == 'normal':
            sql += f' AND NOT {FLOW_ITEM_SQL} AND stock > alert_line'
        elif status == 'low':
            sql += f' AND NOT {FLOW_ITEM_SQL} AND stock <= alert_line AND stock > 0'
        elif status == 'out':
            sql += f' AND NOT {FLOW_ITEM_SQL} AND stock <= 0'

    sql += ' ORDER BY stock ASC'

    rows = db.query(sql, tuple(params))
    for row in rows:
        row['departments'] = json.loads(row.get('departments', '[]') or '[]')
        row['stock'] = clean_number(row.get('stock'))
        row['cost'] = clean_number(row.get('cost'))
        row['alert_line'] = clean_number(row.get('alert_line'))
        row['is_flow_item'] = is_flow_item(row)
        if row['is_flow_item']:
            row['stock'] = 0
            row['status'] = 'flow'
        elif row['stock'] <= 0:
            row['status'] = 'out'
        elif row['stock'] <= row['alert_line']:
            row['status'] = 'low'
        else:
            row['status'] = 'normal'
    return jsonify(rows)


# ---------- 部门统计 ----------

@app.route('/api/stats/departments', methods=['GET'])
def get_dept_stats():
    """获取部门领用统计"""
    depts = ['食堂', '保洁', '办公', '维修', '酒店']
    totals = db.query('''
        SELECT department,
               COUNT(*) as count,
               COALESCE(SUM(total), 0) as total,
               COALESCE(SUM(item_count), 0) as item_count
        FROM (
            SELECT o.id, o.department, o.total, COALESCE(SUM(oi.quantity), 0) as item_count
            FROM outbounds o
            LEFT JOIN outbound_items oi ON o.id = oi.outbound_id
            GROUP BY o.id
        )
        GROUP BY department
    ''')
    totals_by_dept = {row['department']: row for row in totals}
    stats = [{
        'department': dept,
        'total': totals_by_dept.get(dept, {}).get('total', 0),
        'count': totals_by_dept.get(dept, {}).get('count', 0),
        'itemCount': totals_by_dept.get(dept, {}).get('item_count', 0)
    } for dept in depts]

    # 部门领用明细
    details = db.query('''
        SELECT department,
               COUNT(*) as order_count,
               COALESCE(SUM(item_count), 0) as item_count,
               COALESCE(SUM(total), 0) as total_amount
        FROM (
            SELECT o.id, o.department, o.total, COALESCE(SUM(oi.quantity), 0) as item_count
            FROM outbounds o
            LEFT JOIN outbound_items oi ON o.id = oi.outbound_id
            GROUP BY o.id
        )
        GROUP BY department
        ORDER BY total_amount DESC
    ''')

    return jsonify({'stats': stats, 'details': details})


@app.route('/api/stats/sales', methods=['GET'])
def get_sales_stats():
    """获取销售统计"""
    # 销售单数
    order_count = db.query_one('SELECT COUNT(*) as count FROM sales')['count']

    # 销售总成本：酒店生鲜按原价出，其他按物品进货价计算。
    total_cost_result = db.query_one(f'''
        SELECT COALESCE(SUM({HOTEL_COST_SQL}), 0) as cost
        FROM sales_items si
        JOIN sales s ON si.sale_id = s.id
        LEFT JOIN products p ON si.product_id = p.id
    ''')
    total_cost = total_cost_result['cost'] if total_cost_result else 0

    # 销售总金额
    total_revenue = db.query_one('SELECT COALESCE(SUM(total), 0) as total FROM sales')['total']

    # 总利润
    total_profit = total_revenue - total_cost

    # 平均毛利率
    avg_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0

    # 销售明细
    details = db.query(f'''
        SELECT s.customer, s.no, s.date, COUNT(si.id) as item_count,
               COALESCE(SUM({HOTEL_COST_SQL}), 0) as cost,
               s.total as revenue,
               (s.total - COALESCE(SUM({HOTEL_COST_SQL}), 0)) as profit
        FROM sales s
        LEFT JOIN sales_items si ON s.id = si.sale_id
        LEFT JOIN products p ON si.product_id = p.id
        GROUP BY s.id
        ORDER BY s.date DESC
    ''')

    for d in details:
        revenue = d['revenue'] or 0
        cost = d['cost'] or 0
        d['margin'] = ((revenue - cost) / revenue * 100) if revenue > 0 else 0

    return jsonify({
        'orderCount': order_count,
        'totalCost': total_cost,
        'totalRevenue': total_revenue,
        'totalProfit': total_profit,
        'avgMargin': avg_margin,
        'details': details
    })


def build_customer_profit_report(month=''):
    month = str(month or '').strip()
    where_sql = ''
    params = ()
    if month:
        where_sql = 'WHERE s.date LIKE ?'
        params = (f'{month}%',)

    months = db.query('''
        SELECT DISTINCT substr(date, 1, 7) as month
        FROM sales
        WHERE date IS NOT NULL AND date != ''
        ORDER BY month DESC
    ''')

    details = db.query(f'''
        SELECT s.id, s.customer, s.no, s.date, COUNT(si.id) as item_count,
               COALESCE(SUM({HOTEL_COST_SQL}), 0) as cost,
               s.total as revenue,
               (s.total - COALESCE(SUM({HOTEL_COST_SQL}), 0)) as profit
        FROM sales s
        LEFT JOIN sales_items si ON s.id = si.sale_id
        LEFT JOIN products p ON si.product_id = p.id
        {where_sql}
        GROUP BY s.id
        ORDER BY s.date DESC, s.created_at DESC
    ''', params)

    item_rows = db.query(f'''
        SELECT s.customer, s.no, s.date, si.product_name, si.product_unit,
               ({EFFECTIVE_SALE_QUANTITY_SQL}) as quantity, {HOTEL_COST_SQL} as cost,
               si.subtotal as revenue,
               (si.subtotal - {HOTEL_COST_SQL}) as profit
        FROM sales_items si
        JOIN sales s ON si.sale_id = s.id
        LEFT JOIN products p ON si.product_id = p.id
        {where_sql}
        ORDER BY s.date DESC, s.created_at DESC, si.rowid
    ''', params)

    customers = {}
    for row in details:
        customer = row['customer'] or '未填写客户'
        revenue = row['revenue'] or 0
        cost = row['cost'] or 0
        profit = row['profit'] or (revenue - cost)
        row['margin'] = (profit / revenue * 100) if revenue > 0 else 0

        group = customers.setdefault(customer, {
            'customer': customer,
            'orderCount': 0,
            'itemCount': 0,
            'cost': 0,
            'revenue': 0,
            'profit': 0,
            'margin': 0,
        })
        group['orderCount'] += 1
        group['itemCount'] += row['item_count'] or 0
        group['cost'] += cost
        group['revenue'] += revenue
        group['profit'] += profit

    summaries = list(customers.values())
    for row in summaries:
        row['margin'] = (row['profit'] / row['revenue'] * 100) if row['revenue'] > 0 else 0
    summaries.sort(key=lambda row: row['profit'], reverse=True)

    categories = {}
    items = {}
    for row in item_rows:
        name = row['product_name'] or '未填写品名'
        unit = row['product_unit'] or ''
        category = hotel_item_category(name)
        revenue = row['revenue'] or 0
        cost = row['cost'] or 0
        profit = row['profit'] or (revenue - cost)
        quantity = row['quantity'] or 0

        category_row = categories.setdefault(category, {
            'category': category,
            'itemCount': 0,
            'quantity': 0,
            'cost': 0,
            'revenue': 0,
            'profit': 0,
            'margin': 0,
        })
        category_row['itemCount'] += 1
        category_row['quantity'] += quantity
        category_row['cost'] += cost
        category_row['revenue'] += revenue
        category_row['profit'] += profit

        key = (name, unit)
        item_row = items.setdefault(key, {
            'name': name,
            'unit': unit,
            'category': category,
            'count': 0,
            'quantity': 0,
            'cost': 0,
            'revenue': 0,
            'profit': 0,
            'margin': 0,
        })
        item_row['count'] += 1
        item_row['quantity'] += quantity
        item_row['cost'] += cost
        item_row['revenue'] += revenue
        item_row['profit'] += profit

    category_summaries = list(categories.values())
    rank = {category: idx for idx, category in enumerate(HOTEL_CATEGORY_ORDER)}
    for row in category_summaries:
        row['margin'] = (row['profit'] / row['revenue'] * 100) if row['revenue'] > 0 else 0
    category_summaries.sort(key=lambda row: rank.get(row['category'], 999))

    item_summaries = list(items.values())
    for row in item_summaries:
        row['margin'] = (row['profit'] / row['revenue'] * 100) if row['revenue'] > 0 else 0
    item_summaries.sort(key=lambda row: row['profit'], reverse=True)

    total = {
        'customerCount': len(summaries),
        'orderCount': sum(row['orderCount'] for row in summaries),
        'itemCount': sum(row['itemCount'] for row in summaries),
        'cost': sum(row['cost'] for row in summaries),
        'revenue': sum(row['revenue'] for row in summaries),
        'profit': sum(row['profit'] for row in summaries),
    }
    total['margin'] = (total['profit'] / total['revenue'] * 100) if total['revenue'] > 0 else 0

    return {
        'month': month,
        'months': [row['month'] for row in months if row.get('month')],
        'customers': summaries,
        'details': details,
        'categories': category_summaries,
        'items': item_summaries,
        'total': total
    }


@app.route('/api/stats/customer-profits', methods=['GET'])
def get_customer_profit_stats():
    """按客户汇总销售成本、销售额和利润。"""
    return jsonify(build_customer_profit_report(request.args.get('month', '')))


@app.route('/api/export/customer-profit-report', methods=['GET'])
def export_customer_profit_report():
    """导出客户利润月报。"""
    report = build_customer_profit_report(request.args.get('month', ''))
    output = io.StringIO()
    writer = csv.writer(output)
    month_label = report['month'] or '全部月份'

    writer.writerow([f'客户利润报表：{month_label}'])
    writer.writerow([])
    writer.writerow(['汇总', '客户数', '销售单数', '物品数', '成本', '销售额', '利润', '毛利率'])
    total = report['total']
    writer.writerow([
        '', total['customerCount'], total['orderCount'], total['itemCount'],
        round(total['cost'], 2), round(total['revenue'], 2),
        round(total['profit'], 2), f"{round(total['margin'], 2)}%"
    ])

    writer.writerow([])
    writer.writerow(['客户汇总'])
    writer.writerow(['客户', '销售单数', '物品数', '成本', '销售额', '利润', '毛利率'])
    for row in report['customers']:
        writer.writerow([
            row['customer'], row['orderCount'], row['itemCount'],
            round(row['cost'], 2), round(row['revenue'], 2),
            round(row['profit'], 2), f"{round(row['margin'], 2)}%"
        ])

    writer.writerow([])
    writer.writerow(['品类汇总'])
    writer.writerow(['品类', '明细数', '数量', '花费/成本', '销售额', '利润', '毛利率'])
    for row in report['categories']:
        writer.writerow([
            row['category'], row['itemCount'], round(row['quantity'], 4),
            round(row['cost'], 2), round(row['revenue'], 2),
            round(row['profit'], 2), f"{round(row['margin'], 2)}%"
        ])

    writer.writerow([])
    writer.writerow(['单品利润排行'])
    writer.writerow(['品名', '单位', '品类', '出现次数', '数量', '成本', '销售额', '利润', '毛利率'])
    for row in report['items']:
        writer.writerow([
            row['name'], row['unit'], row['category'], row['count'], round(row['quantity'], 4),
            round(row['cost'], 2), round(row['revenue'], 2),
            round(row['profit'], 2), f"{round(row['margin'], 2)}%"
        ])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'客户利润报表_{month_label}_{date.today().strftime("%Y-%m-%d")}.csv'
    )


# ---------- 账单统计 ----------

@app.route('/api/statement', methods=['GET'])
def get_statement():
    """获取对账单"""
    # 采购支出
    purchases_total = db.query_one('SELECT COALESCE(SUM(total), 0) as total FROM purchases')['total']

    # 领用金额
    outbounds_total = db.query_one('SELECT COALESCE(SUM(total), 0) as total FROM outbounds')['total']

    # 其他支出
    other_expense = db.query_one(
        "SELECT COALESCE(SUM(amount), 0) as total FROM finances WHERE type = 'expense'"
    )['total']

    # 收支明细
    purchase_records = db.query('''
        SELECT date, '采购入库' as type, supplier_name as category, total as amount, no as note
        FROM purchases
    ''')

    outbound_records = db.query('''
        SELECT date, '部门领用' as type, department as category, total as amount, no as note
        FROM outbounds
    ''')

    finance_records = db.query('''
        SELECT date, CASE WHEN type = 'income' THEN '收入' ELSE '支出' END as type,
               category, amount, note
        FROM finances
    ''')

    all_records = purchase_records + outbound_records + finance_records
    all_records.sort(key=lambda x: x['date'], reverse=True)

    return jsonify({
        'totalPurchase': purchases_total,
        'totalOutbound': outbounds_total,
        'totalOtherExpense': other_expense,
        'records': all_records
    })


# ---------- 数据备份 ----------

@app.route('/api/backup', methods=['GET'])
def export_backup():
    """导出所有数据为JSON"""
    data = {
        'version': '2.0',
        'exportTime': now_str(),
        'products': db.query('SELECT * FROM products'),
        'suppliers': db.query('SELECT * FROM suppliers'),
        'customers': db.query('SELECT * FROM customers'),
        'purchases': db.query('SELECT * FROM purchases'),
        'purchaseItems': db.query('SELECT * FROM purchase_items'),
        'outbounds': db.query('SELECT * FROM outbounds'),
        'outboundItems': db.query('SELECT * FROM outbound_items'),
        'sales': db.query('SELECT * FROM sales'),
        'salesItems': db.query('SELECT * FROM sales_items'),
        'hotelStockDeductions': db.query('SELECT * FROM hotel_stock_deductions'),
        'customerProductPrices': db.query('SELECT * FROM customer_product_prices'),
        'finances': db.query('SELECT * FROM finances')
    }
    return jsonify(data)


@app.route('/api/backup', methods=['POST'])
def import_backup():
    """从JSON导入数据"""
    data = normalize_legacy_backup(request.json or {})

    conn = db._get_conn()
    try:
        # 清空现有数据
        tables = ['finances', 'customer_product_prices', 'hotel_stock_deductions', 'sales_items', 'sales', 'outbound_items', 'outbounds',
                  'purchase_items', 'purchases', 'customers', 'suppliers', 'products']
        for table in tables:
            conn.execute(f'DELETE FROM {table}')

        # 导入物品
        for p in data.get('products', []):
            conn.execute('''
                INSERT INTO products (id, code, name, spec, unit, cost, stock, alert_line, departments, last_sale_price, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (p['id'], p['code'], p['name'], p.get('spec', ''), p['unit'], as_float(p.get('cost')),
                  as_float(p.get('stock')), as_float(first_value(p, 'alert_line', 'alertLine', default=10), 10),
                  as_json_array(p.get('departments')), as_float(first_value(p, 'last_sale_price', 'lastSalePrice')),
                  p.get('created_at', now_str())))

        # 导入供应商
        for s in data.get('suppliers', []):
            conn.execute('''
                INSERT INTO suppliers (id, code, name, contact, phone, address, note, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (s['id'], s['code'], s['name'], s.get('contact', ''), s.get('phone', ''),
                  s.get('address', ''), s.get('note', ''), s.get('created_at', now_str())))

        # 导入客户
        for c in data.get('customers', []):
            conn.execute('''
                INSERT INTO customers (id, code, name, company, phone, address, note, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (c['id'], c['code'], c['name'], c.get('company', ''), c.get('phone', ''),
                  c.get('address', ''), c.get('note', ''), c.get('created_at', now_str())))

        # 导入采购单
        for p in data.get('purchases', []):
            conn.execute('''
                INSERT INTO purchases (id, no, supplier_id, supplier_name, date, total, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (p['id'], p['no'], first_value(p, 'supplier_id', 'supplierId'),
                  first_value(p, 'supplier_name', 'supplierName', 'supplier'),
                  normalize_date(p['date']), as_float(p.get('total')), p.get('created_at', now_str())))

        for pi in data.get('purchaseItems', []):
            conn.execute('''
                INSERT INTO purchase_items (id, purchase_id, product_id, product_name, product_spec, product_unit, quantity, price, subtotal)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (pi['id'], pi['purchase_id'], pi['product_id'], pi.get('product_name', ''),
                  pi.get('product_spec', ''), pi.get('product_unit', ''), as_float(pi['quantity']),
                  as_float(pi['price']), as_float(pi['subtotal'])))

        # 导入领用单
        for o in data.get('outbounds', []):
            conn.execute('''
                INSERT INTO outbounds (id, no, department, person, date, total, purchase_id, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (o['id'], o['no'], o['department'], o.get('person', ''), normalize_date(o['date']),
                  as_float(o.get('total')), first_value(o, 'purchase_id', 'purchaseId'),
                  o.get('created_at', now_str())))

        for oi in data.get('outboundItems', []):
            conn.execute('''
                INSERT INTO outbound_items (id, outbound_id, product_id, product_name, product_spec, product_unit, quantity, price, subtotal)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (oi['id'], oi['outbound_id'], oi['product_id'], oi.get('product_name', ''),
                  oi.get('product_spec', ''), oi.get('product_unit', ''), as_float(oi['quantity']),
                  as_float(oi['price']), as_float(oi['subtotal'])))

        # 导入销售单
        for s in data.get('sales', []):
            conn.execute('''
                INSERT INTO sales (id, no, company, customer, phone, date, total, show_handlers, handler, issuer, order_key, purchase_id, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (s['id'], s['no'], s.get('company', ''), s['customer'], s.get('phone', ''),
                  normalize_date(s['date']), as_float(s.get('total')),
                  1 if first_value(s, 'show_handlers', 'showHandlers', default=0) else 0,
                  s.get('handler', ''), s.get('issuer', ''), first_value(s, 'order_key', 'orderKey', default=''),
                  first_value(s, 'purchase_id', 'purchaseId'), s.get('created_at', now_str())))

        for si in data.get('salesItems', []):
            conn.execute('''
                INSERT INTO sales_items (id, sale_id, product_id, product_name, product_spec, product_unit, quantity, price, subtotal, note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (si['id'], si['sale_id'], si['product_id'], si.get('product_name', ''),
                  si.get('product_spec', ''), si.get('product_unit', ''), as_float(si['quantity']),
                  as_float(si['price']), as_float(si['subtotal']), si.get('note', '')))

        for hsd in data.get('hotelStockDeductions', []):
            conn.execute('''
                INSERT INTO hotel_stock_deductions (sale_item_id, sale_id, product_id, quantity)
                VALUES (?, ?, ?, ?)
            ''', (hsd['sale_item_id'], hsd['sale_id'], hsd['product_id'], as_float(hsd['quantity'])))

        for row in data.get('customerProductPrices', []):
            conn.execute('''
                INSERT INTO customer_product_prices
                    (id, customer, product_id, product_name, product_unit, price, source, updated_at, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                row.get('id') or generate_id(), row.get('customer') or '', row.get('product_id') or '',
                row.get('product_name') or '', row.get('product_unit') or '', as_float(row.get('price')),
                row.get('source') or '', row.get('updated_at') or now_str(), row.get('created_at') or now_str()
            ))

        # 导入财务
        for f in data.get('finances', []):
            conn.execute('''
                INSERT INTO finances (id, date, type, category, amount, note, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (f['id'], normalize_date(f['date']), f['type'], f.get('category', ''), as_float(f['amount']),
                  f.get('note', ''), f.get('created_at', now_str())))

        conn.commit()
        return jsonify({'message': '导入成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/clear', methods=['POST'])
def clear_all_data():
    """清空所有数据"""
    conn = db._get_conn()
    try:
        tables = ['finances', 'hotel_stock_deductions', 'sales_items', 'sales', 'outbound_items', 'outbounds',
                  'purchase_items', 'purchases', 'customers', 'suppliers', 'products']
        for table in tables:
            conn.execute(f'DELETE FROM {table}')
        conn.commit()
        return jsonify({'message': '清空成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ---------- Excel导入 ----------

@app.route('/api/import/products-excel', methods=['POST'])
def import_products_excel():
    """从Excel/CSV导入物品"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    try:
        content = file.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(content))
        headers = next(reader, None)
        if not headers:
            return jsonify({'error': '文件为空'}), 400

        # 列名映射
        col_map = {}
        for i, h in enumerate(headers):
            h_lower = h.strip().lower()
            if h_lower in ['编号', 'code', '物品编号']:
                col_map['code'] = i
            elif h_lower in ['名称', 'name', '物品名称']:
                col_map['name'] = i
            elif h_lower in ['规格', 'spec', '规格型号']:
                col_map['spec'] = i
            elif h_lower in ['单位', 'unit']:
                col_map['unit'] = i
            elif h_lower in ['采购单价', 'cost', '单价', '价格']:
                col_map['cost'] = i
            elif h_lower in ['初始库存', 'stock', '库存']:
                col_map['stock'] = i
            elif h_lower in ['预警线', 'alert', 'alert_line', '预警']:
                col_map['alert_line'] = i
            elif h_lower in ['专属部门', 'dept', '部门', 'departments']:
                col_map['departments'] = i

        if 'code' not in col_map or 'name' not in col_map:
            return jsonify({'error': '缺少必要的列（编号、名称）'}), 400

        imported = 0
        skipped = 0

        for row in reader:
            if not row:
                continue
            code = row[col_map.get('code', 0)].strip() if len(row) > col_map.get('code', 0) else ''
            name = row[col_map.get('name', 1)].strip() if len(row) > col_map.get('name', 1) else ''

            if not code or not name:
                continue

            # 检查是否已存在
            existing = db.query_one('SELECT id FROM products WHERE code = ?', (code,))
            if existing:
                skipped += 1
                continue

            spec = row[col_map.get('spec', 2)].strip() if len(row) > col_map.get('spec', 2) else ''
            unit = row[col_map.get('unit', 3)].strip() if len(row) > col_map.get('unit', 3) else '件'
            cost = float(row[col_map.get('cost', 4)] or 0) if len(row) > col_map.get('cost', 4) else 0
            stock = float(row[col_map.get('stock', 5)] or 0) if len(row) > col_map.get('stock', 5) else 0
            alert_line = float(row[col_map.get('alert_line', 6)] or 10) if len(row) > col_map.get('alert_line', 6) else 10

            depts = []
            if 'departments' in col_map and len(row) > col_map['departments']:
                dept_str = row[col_map['departments']].strip()
                if dept_str:
                    depts = [d.strip() for d in dept_str.split(',') if d.strip()]

            new_id = generate_id()
            db.execute('''
                INSERT INTO products (id, code, name, spec, unit, cost, stock, alert_line, departments, last_sale_price, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (new_id, code, name, spec, unit, cost, stock, alert_line, json.dumps(depts, ensure_ascii=False), 0, now_str()))
            imported += 1

        return jsonify({'imported': imported, 'skipped': skipped})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/import/outbounds-excel', methods=['POST'])
def import_outbounds_excel():
    """从Excel/CSV导入领用单"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    try:
        content = file.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(content))
        headers = next(reader, None)
        if not headers:
            return jsonify({'error': '文件为空'}), 400

        col_map = {}
        for i, h in enumerate(headers):
            h_lower = h.strip().lower()
            if h_lower in ['日期', 'date']:
                col_map['date'] = i
            elif h_lower in ['部门', 'department', '领用部门']:
                col_map['dept'] = i
            elif h_lower in ['领用人', 'person']:
                col_map['person'] = i
            elif h_lower in ['物品编号', 'product_code', '编号']:
                col_map['product_code'] = i
            elif h_lower in ['物品名称', 'product_name', '名称']:
                col_map['product_name'] = i
            elif h_lower in ['数量', 'quantity']:
                col_map['quantity'] = i
            elif h_lower in ['单价', 'price']:
                col_map['price'] = i

        if 'date' not in col_map or 'dept' not in col_map:
            return jsonify({'error': '缺少必要的列（日期、部门）'}), 400

        # 读取所有行，按日期+部门+领用人分组
        rows_data = []
        for row in reader:
            if not row:
                continue
            date_str = row[col_map.get('date', 0)].strip() if len(row) > col_map.get('date', 0) else ''
            dept = row[col_map.get('dept', 1)].strip() if len(row) > col_map.get('dept', 1) else ''
            person = row[col_map.get('person', 2)].strip() if len(row) > col_map.get('person', 2) else ''
            product_code = row[col_map.get('product_code', 3)].strip() if 'product_code' in col_map and len(row) > col_map['product_code'] else ''
            product_name = row[col_map.get('product_name', 4)].strip() if 'product_name' in col_map and len(row) > col_map['product_name'] else ''
            quantity = float(row[col_map.get('quantity', 5)] or 0) if 'quantity' in col_map and len(row) > col_map['quantity'] else 0
            price = float(row[col_map.get('price', 6)] or 0) if 'price' in col_map and len(row) > col_map['price'] else 0

            if not date_str or not dept or quantity <= 0:
                continue

            # 查找物品
            product = None
            if product_code:
                product = db.query_one('SELECT * FROM products WHERE code = ?', (product_code,))
            if not product and product_name:
                product = db.query_one('SELECT * FROM products WHERE name = ?', (product_name,))
                if not product:
                    product = db.query_one('SELECT * FROM products WHERE name LIKE ?', (f'%{product_name}%',))

            if not product:
                continue

            rows_data.append({
                'date': date_str, 'dept': dept, 'person': person,
                'product_id': product['id'], 'product_name': product['name'],
                'product_spec': product['spec'], 'product_unit': product['unit'],
                'quantity': quantity, 'price': price or product['cost']
            })

        # 按日期+部门+领用人分组
        groups = {}
        for rd in rows_data:
            key = (rd['date'], rd['dept'], rd['person'])
            if key not in groups:
                groups[key] = []
            groups[key].append(rd)

        created = 0
        for key, items in groups.items():
            date_str, dept, person = key
            total = sum(i['quantity'] * i['price'] for i in items)

            oid = generate_id()
            ono = generate_no('LY')

            conn = db._get_conn()
            try:
                conn.execute('''
                    INSERT INTO outbounds (id, no, department, person, date, total, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (oid, ono, dept, person, date_str, total, now_str()))

                for item in items:
                    conn.execute('''
                        INSERT INTO outbound_items (id, outbound_id, product_id, product_name, product_spec, product_unit, quantity, price, subtotal)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (generate_id(), oid, item['product_id'], item['product_name'],
                          item['product_spec'], item['product_unit'], item['quantity'],
                          item['price'], item['quantity'] * item['price']))

                    conn.execute("UPDATE products SET stock = stock - ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')",
                                 (item['quantity'], item['product_id']))

                conn.commit()
                created += 1
            except Exception as e:
                conn.rollback()
            finally:
                conn.close()

        return jsonify({'created': created})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/import/purchases-excel', methods=['POST'])
def import_purchases_excel():
    """从Excel/CSV导入采购单"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    try:
        raw = file.read()
        try:
            content = raw.decode('utf-8-sig')
        except UnicodeDecodeError:
            content = raw.decode('gbk')

        reader = csv.reader(io.StringIO(content))
        headers = next(reader, None)
        if not headers:
            return jsonify({'error': '文件为空'}), 400

        col_map = {}
        for i, h in enumerate(headers):
            h_lower = h.strip().lower()
            if h_lower in ['日期', 'date']:
                col_map['date'] = i
            elif h_lower in ['供应商', 'supplier', '供应商名称']:
                col_map['supplier'] = i
            elif h_lower in ['物品编号', 'product_code', '编号']:
                col_map['product_code'] = i
            elif h_lower in ['物品名称', 'product_name', '名称']:
                col_map['product_name'] = i
            elif h_lower in ['规格', 'spec', '规格型号']:
                col_map['spec'] = i
            elif h_lower in ['单位', 'unit']:
                col_map['unit'] = i
            elif h_lower in ['数量', 'quantity']:
                col_map['quantity'] = i
            elif h_lower in ['单价', 'price', '采购单价']:
                col_map['price'] = i

        required = ['date', 'supplier', 'product_name', 'quantity']
        missing = [name for name in required if name not in col_map]
        if missing:
            return jsonify({'error': '缺少必要的列（日期、供应商、物品名称、数量）'}), 400

        rows_data = []
        skipped = 0
        for row in reader:
            if not row or all(not str(cell or '').strip() for cell in row):
                continue

            date_str = normalize_date(row[col_map['date']].strip() if len(row) > col_map['date'] else '')
            supplier = row[col_map['supplier']].strip() if len(row) > col_map['supplier'] else ''
            product_code = row[col_map['product_code']].strip() if 'product_code' in col_map and len(row) > col_map['product_code'] else ''
            product_name = row[col_map['product_name']].strip() if len(row) > col_map['product_name'] else ''
            spec = row[col_map['spec']].strip() if 'spec' in col_map and len(row) > col_map['spec'] else ''
            unit = row[col_map['unit']].strip() if 'unit' in col_map and len(row) > col_map['unit'] else '件'
            quantity = as_float(row[col_map['quantity']] if len(row) > col_map['quantity'] else 0)
            price = as_float(row[col_map['price']] if 'price' in col_map and len(row) > col_map['price'] else 0)

            if not date_str or not supplier or not product_name or quantity <= 0:
                skipped += 1
                continue

            rows_data.append({
                'date': date_str,
                'supplier': supplier,
                'product_code': product_code,
                'product_name': product_name,
                'spec': spec,
                'unit': unit or '件',
                'quantity': quantity,
                'price': price
            })

        if not rows_data:
            return jsonify({'error': '没有可导入的数据'}), 400

        groups = {}
        for row in rows_data:
            groups.setdefault((row['date'], row['supplier']), []).append(row)

        created = 0
        created_products = 0
        imported_items = 0
        conn = db._get_conn()
        try:
            for (date_str, supplier_name), items in groups.items():
                total = sum(item['quantity'] * item['price'] for item in items)
                existing_same = conn.execute('''
                    SELECT id FROM purchases
                    WHERE date = ? AND supplier_name = ? AND ABS(total - ?) < 0.001
                    LIMIT 1
                ''', (date_str, supplier_name, total)).fetchone()
                if existing_same:
                    skipped += len(items)
                    continue

                supplier = conn.execute('SELECT * FROM suppliers WHERE name = ? LIMIT 1', (supplier_name,)).fetchone()
                supplier_id = supplier['id'] if supplier else ''

                pid = generate_id()
                pno = generate_no('CG')
                conn.execute('''
                    INSERT INTO purchases (id, no, supplier_id, supplier_name, date, total, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (pid, pno, supplier_id, supplier_name, date_str, total, now_str()))

                for item in items:
                    product = None
                    if item['product_code']:
                        product = conn.execute('SELECT * FROM products WHERE code = ?', (item['product_code'],)).fetchone()
                    if not product:
                        product = conn.execute('''
                            SELECT * FROM products
                            WHERE name = ? AND COALESCE(spec, '') = ? AND unit = ?
                            LIMIT 1
                        ''', (item['product_name'], item['spec'], item['unit'])).fetchone()
                    if not product:
                        product = conn.execute('SELECT * FROM products WHERE name = ? LIMIT 1', (item['product_name'],)).fetchone()

                    if not product:
                        product_id = generate_id()
                        code = item['product_code'] or generate_no('AUTO')
                        while conn.execute('SELECT 1 FROM products WHERE code = ?', (code,)).fetchone():
                            code = generate_no('AUTO')
                        conn.execute('''
                            INSERT INTO products (id, code, name, spec, unit, cost, stock, alert_line, departments, last_sale_price, created_at)
                            VALUES (?, ?, ?, ?, ?, ?, 0, 10, '[]', 0, ?)
                        ''', (product_id, code, item['product_name'], item['spec'], item['unit'], item['price'], now_str()))
                        product = conn.execute('SELECT * FROM products WHERE id = ?', (product_id,)).fetchone()
                        created_products += 1

                    subtotal = item['quantity'] * item['price']
                    conn.execute('''
                        INSERT INTO purchase_items (id, purchase_id, product_id, product_name, product_spec, product_unit, quantity, price, subtotal)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        generate_id(), pid, product['id'], product['name'], product['spec'] or '',
                        product['unit'], item['quantity'], item['price'], subtotal
                    ))
                    conn.execute("UPDATE products SET stock = stock + ? WHERE id = ? AND NOT (code LIKE 'WP%' OR code LIKE 'IMP%')",
                                 (item['quantity'], product['id']))
                    conn.execute('UPDATE products SET cost = ? WHERE id = ? AND ? > 0',
                                 (item['price'], product['id'], item['price']))
                    imported_items += 1

                created += 1

            conn.commit()
            return jsonify({
                'created': created,
                'items': imported_items,
                'createdProducts': created_products,
                'skipped': skipped
            })
        except Exception as e:
            conn.rollback()
            return jsonify({'error': str(e)}), 500
        finally:
            conn.close()

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/import/hotel-sales-excel', methods=['POST'])
def import_hotel_sales_excel():
    """按酒店采购清单生成销售单，销售价优先使用上次销售价。"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    try:
        raw = file.read()
        try:
            content = raw.decode('utf-8-sig')
        except UnicodeDecodeError:
            content = raw.decode('gbk')

        reader = csv.reader(io.StringIO(content))
        headers = next(reader, None)
        if not headers:
            return jsonify({'error': '文件为空'}), 400

        col_map = {}
        for i, h in enumerate(headers):
            h_lower = h.strip().lower()
            if h_lower in ['日期', 'date', '采购日期']:
                col_map['date'] = i
            elif h_lower in ['客户', '客户名称', '酒店', '酒店名称', 'customer']:
                col_map['customer'] = i
            elif h_lower in ['公司抬头', '抬头', 'company']:
                col_map['company'] = i
            elif h_lower in ['电话', '联系电话', 'phone']:
                col_map['phone'] = i
            elif h_lower in ['物品编号', 'product_code', '编号']:
                col_map['product_code'] = i
            elif h_lower in ['物品名称', 'product_name', '名称', '品名']:
                col_map['product_name'] = i
            elif h_lower in ['规格', 'spec', '规格型号']:
                col_map['spec'] = i
            elif h_lower in ['单位', 'unit']:
                col_map['unit'] = i
            elif h_lower in ['数量', 'quantity', '采购数量']:
                col_map['quantity'] = i

        missing = [name for name in ['product_name', 'quantity'] if name not in col_map]
        if missing:
            return jsonify({'error': '缺少必要的列（物品名称、数量）'}), 400

        rows_data = []
        skipped = 0
        missing_products = []
        no_last_price = []
        for row in reader:
            if not row or all(not str(cell or '').strip() for cell in row):
                continue

            date_str = normalize_date(row[col_map['date']].strip()) if 'date' in col_map and len(row) > col_map['date'] else date.today().strftime('%Y-%m-%d')
            customer = row[col_map['customer']].strip() if 'customer' in col_map and len(row) > col_map['customer'] else '酒店'
            company = row[col_map['company']].strip() if 'company' in col_map and len(row) > col_map['company'] else ''
            phone = row[col_map['phone']].strip() if 'phone' in col_map and len(row) > col_map['phone'] else ''
            product_code = row[col_map['product_code']].strip() if 'product_code' in col_map and len(row) > col_map['product_code'] else ''
            product_name = row[col_map['product_name']].strip() if len(row) > col_map['product_name'] else ''
            spec = row[col_map['spec']].strip() if 'spec' in col_map and len(row) > col_map['spec'] else ''
            unit = row[col_map['unit']].strip() if 'unit' in col_map and len(row) > col_map['unit'] else ''
            quantity = as_float(row[col_map['quantity']] if len(row) > col_map['quantity'] else 0)

            if not product_name or quantity <= 0:
                skipped += 1
                continue

            product = None
            if product_code:
                product = db.query_one('SELECT * FROM products WHERE code = ?', (product_code,))
            if not product:
                product = db.query_one('''
                    SELECT * FROM products
                    WHERE name = ?
                      AND (? = '' OR COALESCE(spec, '') = ?)
                      AND (? = '' OR unit = ?)
                    LIMIT 1
                ''', (product_name, spec, spec, unit, unit))
            if not product:
                product = db.query_one('SELECT * FROM products WHERE name = ? LIMIT 1', (product_name,))
            if not product:
                product = db.query_one('SELECT * FROM products WHERE name LIKE ? LIMIT 1', (f'%{product_name}%',))

            if not product:
                missing_products.append(product_name)
                skipped += 1
                continue

            price = get_customer_product_price(customer, product, unit or product.get('unit') or '')
            if price <= 0:
                no_last_price.append(product['name'])
                price = as_float(product.get('last_sale_price')) or as_float(product.get('cost'))

            rows_data.append({
                'date': date_str,
                'customer': customer or '酒店',
                'company': company,
                'phone': phone,
                'order_key': hotel_sales_order_key(customer, product_name),
                'product': product,
                'product_name': product_name or product.get('name'),
                'product_unit': unit or product.get('unit'),
                'quantity': quantity,
                'price': price
            })

        result, error, status = create_hotel_sales_orders(rows_data)
        if error:
            error['missingProducts'] = sorted(set(missing_products))
            return jsonify(error), status

        result.update({
            'skipped': skipped,
            'missingProducts': sorted(set(missing_products)),
            'noLastPrice': sorted(set(no_last_price))
        })
        return jsonify(result)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/import/hotel-sales-text', methods=['POST'])
def import_hotel_sales_text():
    """把酒店自然语言清单生成销售单。"""
    data = request.json or {}
    text = data.get('text', '')
    date_str = normalize_date(data.get('date') or date.today().strftime('%Y-%m-%d'))
    customer = (data.get('customer') or '酒店').strip()
    company = (data.get('company') or '').strip()
    phone = (data.get('phone') or '').strip()

    parsed_items = []
    leftovers = []
    for section_text, order_key in split_front_desk_fruit_text(customer, text):
        section_items, section_leftovers = parse_hotel_sales_text(section_text)
        for item in section_items:
            if order_key:
                item['order_key'] = order_key
            parsed_items.append(item)
        leftovers.extend(section_leftovers)
    if not parsed_items:
        return jsonify({'error': '没有识别到“物品+数量+单位”的清单内容'}), 400

    conn = db._get_conn()
    try:
        rows_data, created_products = prepare_hotel_grocery_sales_rows(
            conn, parsed_items, date_str, customer or '酒店', company, phone
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e), 'unparsed': leftovers}), 500
    finally:
        conn.close()

    result, error, status = create_hotel_sales_orders(rows_data)
    if error:
        error['unparsed'] = leftovers
        return jsonify(error), status
    push_sales_to_feishu_async(result.get('saleIds') or [])

    result.update({
        'parsed': len(parsed_items),
        'skipped': 0,
        'missingProducts': [],
        'createdProducts': len(created_products),
        'createdProductNames': sorted(set(created_products)),
        'unparsed': leftovers
    })
    return jsonify(result)


@app.route('/api/hotel-grocery/sync', methods=['POST'])
def sync_hotel_grocery_text():
    """把酒店订单文字同步生成采购入库单和销售出库单。"""
    data = request.json or {}
    text = data.get('text', '')
    date_str = normalize_date(data.get('date') or date.today().strftime('%Y-%m-%d'))
    supplier_name = (data.get('supplierName') or '西安禾润佳商贸有限公司').strip()
    customer = (data.get('customer') or '西安汉庭酒店（大明宫万达）').strip()
    company = (data.get('company') or supplier_name).strip()
    phone = (data.get('phone') or '').strip()

    items = []
    unparsed = []
    for section_text, order_key in split_front_desk_fruit_text(customer, text):
        section_items, section_unparsed = parse_hotel_grocery_text(section_text, preserve_order=True)
        for item in section_items:
            if order_key:
                item['order_key'] = order_key
            items.append(item)
        unparsed.extend(section_unparsed)
    if not items:
        return jsonify({'error': '没有识别到“物品+数量+单位”的订单内容', 'unparsed': unparsed}), 400

    result, error, status = sync_hotel_grocery_orders(
        items=items,
        date_str=date_str,
        supplier_name=supplier_name,
        customer=customer,
        company=company,
        phone=phone,
    )
    if error:
        error['unparsed'] = unparsed
        return jsonify(error), status

    result.update({
        'parsed': len(items),
        'smartParsed': sum(1 for item in items if item.get('smart')),
        'unparsed': unparsed,
    })
    return jsonify(result)


@app.route('/api/hotel-grocery/purchase-list', methods=['POST'])
def download_hotel_grocery_purchase_list():
    """把酒店文字清单生成供应商采购填表单，不写入数据库。"""
    data = request.json or {}
    text = data.get('text', '')
    date_str = normalize_date(data.get('date') or date.today().strftime('%Y-%m-%d'))
    title = (data.get('title') or f'酒店采购填表单{date_str}').strip()

    items, unparsed = parse_hotel_grocery_text(text)
    if not items:
        return jsonify({'error': '没有识别到“物品+数量+单位”的订单内容', 'unparsed': unparsed}), 400

    wb = build_purchase_list_workbook(items, title)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'采购单_{date_str}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def extract_date_text(value):
    if value is None:
        return ''
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')
    text = str(value).strip()
    match = re.search(r'(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})', text)
    if match:
        return match.group(1).replace('年', '-').replace('月', '-').replace('日', '')
    return text


def extract_customer_from_sales_row(values, fallback=''):
    row_text = ' '.join(str(v) for v in values if v is not None)
    customer = fallback or ''
    for value in values:
        text = str(value or '').strip()
        if not text:
            continue
        if '客户' in text:
            parts = re.split(r'客户\s*[:：]?', text, maxsplit=1)
            candidate = parts[-1].strip() if parts else ''
            candidate = re.split(r'\s+电话|\s+日期|电话[:：]|日期[:：]', candidate)[0].strip()
            if candidate:
                customer = candidate
    match = re.search(r'客户\s*[:：]\s*([^电话日期]+?)(?:\s{2,}|电话|日期|$)', row_text)
    if match:
        customer = match.group(1).strip()
    return customer or fallback or '酒店'


def locate_sales_price_columns(ws, header_row):
    aliases = {
        'name': ('货品名称', '商品名称', '物品名称', '品名', '名称'),
        'spec': ('规格', '规格型号'),
        'quantity': ('数量', '采购数量'),
        'unit': ('单位', '计量单位'),
        'price': ('单价', '销售单价', '售价'),
        'amount': ('金额', '小计', '合计金额'),
        'note': ('备注',),
    }
    col_map = {}
    for col in range(1, ws.max_column + 1):
        text = re.sub(r'\s+', '', str(ws.cell(header_row, col).value or ''))
        if not text:
            continue
        for key, names in aliases.items():
            if text in names and key not in col_map:
                col_map[key] = col
    return col_map


def extract_sales_price_rows_from_workbook(wb, fallback_customer='', source_name='销售单导入'):
    rows = []
    skipped = 0
    for ws in wb.worksheets:
        current_customer = fallback_customer or ''
        for row_idx in range(1, ws.max_row + 1):
            values = [ws.cell(row_idx, col).value for col in range(1, min(ws.max_column, 8) + 1)]
            row_text = ' '.join(str(v) for v in values if v is not None)
            if '客户' in row_text:
                current_customer = extract_customer_from_sales_row(values, current_customer)

            col_map = locate_sales_price_columns(ws, row_idx)
            if not {'name', 'quantity', 'unit', 'price'}.issubset(col_map):
                continue

            for item_row in range(row_idx + 1, ws.max_row + 1):
                first = str(ws.cell(item_row, 1).value or '').strip()
                name = str(ws.cell(item_row, col_map['name']).value or '').strip()
                if first.startswith('合计') or name.startswith('合计'):
                    break
                if not name or name in ('货品名称', '商品名称', '物品名称', '品名'):
                    continue
                quantity = as_float(ws.cell(item_row, col_map['quantity']).value)
                unit = str(ws.cell(item_row, col_map['unit']).value or '').strip()
                price = as_float(ws.cell(item_row, col_map['price']).value)
                amount = as_float(ws.cell(item_row, col_map['amount']).value) if 'amount' in col_map else 0
                if price <= 0 and quantity > 0 and amount > 0:
                    price = amount / quantity
                if not unit or price <= 0:
                    skipped += 1
                    continue
                raw_name = re.sub(r'\s+', '', name)
                normalized_name = normalize_hotel_item_name(raw_name)
                rows.append({
                    'customer': current_customer or fallback_customer or '酒店',
                    'name': normalized_name,
                    'displayName': raw_name,
                    'spec': str(ws.cell(item_row, col_map['spec']).value or '').strip() if 'spec' in col_map else '',
                    'unit': unit,
                    'quantity': quantity,
                    'price': price,
                    'amount': amount,
                    'source': source_name,
                    'sheet': ws.title,
                    'row': item_row,
                })
    return rows, skipped


def import_sales_price_rows_to_db(extracted_rows, skipped=0):
    latest = {}
    for row in extracted_rows:
        key = (canonical_customer_name(row['customer']), row['name'], row['unit'])
        latest[key] = row

    conn = db._get_conn()
    created_prices = 0
    updated_prices = 0
    created_products = 0
    created_customers = set()
    samples = []
    try:
        for row in latest.values():
            customer_name = ensure_customer_name_with_conn(conn, row['customer'])
            created_customers.add(customer_name)
            product = find_product_with_conn(conn, row['name'], row.get('spec') or '', row['unit'])
            if not product:
                product = create_hotel_flow_product(conn, row['name'], row.get('spec') or '', row['unit'], row['price'])
                created_products += 1
            action = 'updated'
            existing = conn.execute('''
                SELECT id FROM customer_product_prices
                WHERE customer = ? AND product_id = ? AND COALESCE(product_unit, '') = ?
                LIMIT 1
            ''', (customer_name, product['id'], row['unit'] or '')).fetchone()
            if not existing:
                action = 'created'
            if upsert_customer_product_price(conn, customer_name, product, row['unit'], row['price'], row['source']):
                if action == 'created':
                    created_prices += 1
                else:
                    updated_prices += 1
                conn.execute('UPDATE products SET last_sale_price = ? WHERE id = ? AND ? > 0', (row['price'], product['id'], row['price']))
                if len(samples) < 12:
                    samples.append({
                        'customer': customer_name,
                        'name': product['name'],
                        'unit': row['unit'],
                        'price': clean_number(row['price']),
                    })
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return {
        'sourceRows': len(extracted_rows),
        'latestPrices': len(latest),
        'createdPrices': created_prices,
        'updatedPrices': updated_prices,
        'createdProducts': created_products,
        'customers': sorted(created_customers),
        'skipped': skipped,
        'samples': samples,
    }


def uploaded_excel_daily_dir(base_dir):
    day = datetime.now().strftime('%Y-%m-%d')
    path = os.path.join(base_dir, day)
    os.makedirs(path, exist_ok=True)
    return path


def unique_uploaded_excel_path(target_dir, filename):
    os.makedirs(target_dir, exist_ok=True)
    safe_name = os.path.basename(filename)
    candidate = os.path.join(target_dir, safe_name)
    if not os.path.exists(candidate):
        return candidate
    stem, ext = os.path.splitext(safe_name)
    suffix = datetime.now().strftime('%H%M%S')
    candidate = os.path.join(target_dir, f'{stem}_{suffix}{ext}')
    counter = 1
    while os.path.exists(candidate):
        candidate = os.path.join(target_dir, f'{stem}_{suffix}_{counter}{ext}')
        counter += 1
    return candidate


def move_uploaded_excel(path, target_base_dir):
    target_dir = uploaded_excel_daily_dir(target_base_dir)
    target = unique_uploaded_excel_path(target_dir, os.path.basename(path))
    shutil.move(path, target)
    return target


def copy_uploaded_excel(path, target_base_dir):
    target_dir = uploaded_excel_daily_dir(target_base_dir)
    target = unique_uploaded_excel_path(target_dir, os.path.basename(path))
    shutil.copy2(path, target)
    return target


def is_processable_uploaded_excel(filename):
    name = os.path.basename(filename)
    if not name or name.startswith('~$') or name.startswith('.'):
        return False
    if name.lower().endswith(('.tmp', '.bak')):
        return False
    return name.lower().endswith(('.xlsx', '.xlsm'))


def uploaded_excel_relpath(path):
    try:
        return os.path.relpath(path, UPLOADED_EXCEL_INCOMING_DIR).replace(os.sep, '/')
    except ValueError:
        return os.path.basename(path)


def uploaded_excel_digest(path):
    stat = os.stat(path)
    digest = hashlib.sha256()
    with open(path, 'rb') as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b''):
            digest.update(chunk)
    return {
        'size': stat.st_size,
        'mtimeNs': getattr(stat, 'st_mtime_ns', int(stat.st_mtime * 1000000000)),
        'sha256': digest.hexdigest(),
    }


def load_uploaded_excel_manifest():
    if not os.path.exists(UPLOADED_EXCEL_MANIFEST_PATH):
        return {}
    try:
        with open(UPLOADED_EXCEL_MANIFEST_PATH, 'r', encoding='utf-8') as handle:
            data = json.load(handle)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_uploaded_excel_manifest(manifest):
    os.makedirs(os.path.dirname(UPLOADED_EXCEL_MANIFEST_PATH), exist_ok=True)
    temp_path = f'{UPLOADED_EXCEL_MANIFEST_PATH}.tmp'
    with open(temp_path, 'w', encoding='utf-8') as handle:
        json.dump(manifest, handle, ensure_ascii=False, indent=2, sort_keys=True)
    os.replace(temp_path, UPLOADED_EXCEL_MANIFEST_PATH)


def collect_uploaded_excel_files():
    files = []
    if not os.path.exists(UPLOADED_EXCEL_INCOMING_DIR):
        return files
    for root, _, filenames in os.walk(UPLOADED_EXCEL_INCOMING_DIR):
        for filename in filenames:
            path = os.path.join(root, filename)
            if os.path.isfile(path) and is_processable_uploaded_excel(filename):
                files.append(path)
    files.sort(key=lambda item: (os.path.getmtime(item), uploaded_excel_relpath(item)))
    return files


def uploaded_excel_status():
    incoming = []
    processed = []
    failed = []
    for label, base_dir, bucket in (
        ('incoming', UPLOADED_EXCEL_INCOMING_DIR, incoming),
        ('processed', UPLOADED_EXCEL_PROCESSED_DIR, processed),
        ('failed', UPLOADED_EXCEL_FAILED_DIR, failed),
    ):
        if not os.path.exists(base_dir):
            continue
        for root, _, files in os.walk(base_dir):
            for filename in files:
                if not is_processable_uploaded_excel(filename):
                    continue
                path = os.path.join(root, filename)
                try:
                    stat = os.stat(path)
                except OSError:
                    continue
                bucket.append({
                    'name': filename,
                    'path': path,
                    'size': stat.st_size,
                    'mtime': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                    'status': label,
                })
    incoming.sort(key=lambda row: row['mtime'], reverse=True)
    processed.sort(key=lambda row: row['mtime'], reverse=True)
    failed.sort(key=lambda row: row['mtime'], reverse=True)
    return {
        'incomingDir': UPLOADED_EXCEL_INCOMING_DIR,
        'processedDir': UPLOADED_EXCEL_PROCESSED_DIR,
        'failedDir': UPLOADED_EXCEL_FAILED_DIR,
        'incoming': incoming[:50],
        'processed': processed[:50],
        'failed': failed[:50],
    }


def process_uploaded_sales_price_excels():
    os.makedirs(UPLOADED_EXCEL_INCOMING_DIR, exist_ok=True)
    os.makedirs(UPLOADED_EXCEL_PROCESSED_DIR, exist_ok=True)
    os.makedirs(UPLOADED_EXCEL_FAILED_DIR, exist_ok=True)
    manifest = load_uploaded_excel_manifest()
    all_files = collect_uploaded_excel_files()
    files = []
    skipped_unchanged = 0
    file_signatures = {}
    for path in all_files:
        relpath = uploaded_excel_relpath(path)
        try:
            signature = uploaded_excel_digest(path)
        except OSError:
            continue
        file_signatures[relpath] = signature
        previous = manifest.get(relpath) or {}
        if previous.get('signature') == signature:
            skipped_unchanged += 1
            continue
        files.append(path)
    backup_path = ''
    if files:
        backup_path = f'{db.db_path}.bak_uploaded_excels_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        shutil.copy2(db.db_path, backup_path)

    summary = {
        'totalIncoming': len(all_files),
        'scanned': len(files),
        'skippedUnchanged': skipped_unchanged,
        'backup': backup_path,
        'processed': 0,
        'failed': 0,
        'sourceRows': 0,
        'latestPrices': 0,
        'createdPrices': 0,
        'updatedPrices': 0,
        'createdProducts': 0,
        'skipped': 0,
        'customers': set(),
        'files': [],
    }

    for path in files:
        filename = os.path.basename(path)
        relpath = uploaded_excel_relpath(path)
        signature = file_signatures.get(relpath)
        file_result = {'name': filename, 'path': relpath, 'ok': False}
        try:
            wb = load_workbook(path, data_only=True)
            fallback_customer = os.path.splitext(filename)[0]
            extracted_rows, skipped = extract_sales_price_rows_from_workbook(
                wb,
                fallback_customer=fallback_customer,
                source_name=f'WinSCP上传:{filename}'
            )
            if not extracted_rows:
                raise ValueError('没有识别到销售单明细，请确认表头包含：货品名称、数量、单位、单价')
            result = import_sales_price_rows_to_db(extracted_rows, skipped)
            copied_to = copy_uploaded_excel(path, UPLOADED_EXCEL_PROCESSED_DIR)
            file_result.update(result)
            file_result.update({'ok': True, 'copiedTo': copied_to})
            manifest[relpath] = {
                'signature': signature,
                'status': 'processed',
                'processedAt': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'archive': copied_to,
                'sourceRows': result.get('sourceRows', 0),
                'latestPrices': result.get('latestPrices', 0),
                'customers': result.get('customers') or [],
            }
            summary['processed'] += 1
            for key in ('sourceRows', 'latestPrices', 'createdPrices', 'updatedPrices', 'createdProducts', 'skipped'):
                summary[key] += result.get(key, 0)
            summary['customers'].update(result.get('customers') or [])
        except Exception as e:
            try:
                copied_to = copy_uploaded_excel(path, UPLOADED_EXCEL_FAILED_DIR)
            except Exception:
                copied_to = ''
            file_result.update({'ok': False, 'error': str(e), 'copiedTo': copied_to})
            manifest[relpath] = {
                'signature': signature,
                'status': 'failed',
                'processedAt': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'archive': copied_to,
                'error': str(e),
            }
            summary['failed'] += 1
        summary['files'].append(file_result)
        save_uploaded_excel_manifest(manifest)

    if files:
        save_uploaded_excel_manifest(manifest)
    summary['customers'] = sorted(summary['customers'])
    return summary


@app.route('/api/import/sales-prices-excel', methods=['POST'])
def import_sales_prices_excel():
    """从销售单 Excel 提取各客户商品售价，更新客户专属售价。"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    fallback_customer = (request.form.get('customer') or '').strip()
    filename = file.filename or '销售单'
    try:
        if not filename.lower().endswith(('.xlsx', '.xlsm')):
            return jsonify({'error': '请上传 .xlsx 或 .xlsm 格式的销售单'}), 400
        wb = load_workbook(file, data_only=True)
        extracted_rows, skipped = extract_sales_price_rows_from_workbook(
            wb,
            fallback_customer=fallback_customer,
            source_name=f'销售单导入:{filename}'
        )
        if not extracted_rows:
            return jsonify({'error': '没有识别到销售单明细，请确认表头包含：货品名称、数量、单位、单价'}), 400

        try:
            result = import_sales_price_rows_to_db(extracted_rows, skipped)
        except Exception as e:
            return jsonify({'error': str(e)}), 500

        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/uploaded-excels/status', methods=['GET'])
def api_uploaded_excel_status():
    return jsonify(uploaded_excel_status())


@app.route('/api/uploaded-excels/process', methods=['POST'])
def api_process_uploaded_sales_price_excels():
    try:
        return jsonify(process_uploaded_sales_price_excels())
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/feishu/status', methods=['GET'])
def feishu_status():
    return jsonify({
        'enabled': FEISHU_SYNC_ENABLED,
        'configured': feishu_configured(),
        'intervalSeconds': FEISHU_SYNC_INTERVAL_SECONDS,
        'tableId': FEISHU_BITABLE_TABLE_ID if feishu_configured() else '',
    })


@app.route('/api/feishu/sync-confirmed', methods=['POST'])
def api_sync_confirmed_feishu_records():
    try:
        return jsonify(sync_confirmed_feishu_records())
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/import/hotel-orders-xlsx', methods=['POST'])
def import_hotel_orders_xlsx():
    """导入酒店订单 Excel，每个客户/日期明细块生成一张销售单。"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb['订单明细'] if '订单明细' in wb.sheetnames else wb[wb.sheetnames[0]]

        orders = []
        current = None
        skipped = 0
        for row_idx in range(1, ws.max_row + 1):
            values = [ws.cell(row_idx, col).value for col in range(1, 9)]
            first = str(values[0] or '').strip()
            row_text = ' '.join(str(v) for v in values if v is not None)

            if '客户' in first and '日期' in row_text:
                if current and current['items']:
                    orders.append(current)
                customer = first.split('：', 1)[1].strip() if '：' in first else first.replace('客户:', '').replace('客户', '').strip()
                date_text = ''
                for value in values:
                    if value is not None and '日期' in str(value):
                        date_text = extract_date_text(value)
                        break
                current = {
                    'order_key': f'xlsx-{row_idx}',
                    'customer': customer or '酒店',
                    'date': normalize_date(date_text or date.today().strftime('%Y-%m-%d')),
                    'items': []
                }
                continue

            if not current or first in ('', '序号') or first.startswith('合计'):
                continue

            if not isinstance(values[0], (int, float)):
                continue

            name = str(values[1] or '').strip()
            if not name:
                continue
            spec = str(values[2] or '').strip()
            if as_float(values[3]) <= 0 and str(values[3] or '').strip() and as_float(values[4]) > 0:
                quantity = as_float(values[4])
                unit = str(values[3] or '').strip() or '件'
                price = as_float(values[5])
                amount = as_float(values[6])
            else:
                quantity = as_float(values[3])
                unit = str(values[4] or '').strip() or '件'
                price = as_float(values[5])
                amount = as_float(values[6])
            if quantity <= 0:
                skipped += 1
                continue
            if price <= 0 and amount > 0:
                price = amount / quantity
            note = str(values[7] or '').strip()

            current['items'].append({
                'name': name,
                'spec': spec,
                'quantity': quantity,
                'unit': unit,
                'price': price,
                'note': note
            })

        if current and current['items']:
            orders.append(current)
        if not orders:
            return jsonify({'error': '没有找到可导入的订单明细'}), 400

        rows_data = []
        created_products = 0
        no_price = []
        conn = db._get_conn()
        try:
            for order in orders:
                for item in order['items']:
                    product = None
                    if item['name']:
                        product = conn.execute('''
                            SELECT * FROM products
                            WHERE name = ?
                              AND (? = '' OR COALESCE(spec, '') = ?)
                              AND (? = '' OR unit = ?)
                            LIMIT 1
                        ''', (item['name'], item['spec'], item['spec'], item['unit'], item['unit'])).fetchone()
                    if not product:
                        product = conn.execute('SELECT * FROM products WHERE name = ? LIMIT 1', (item['name'],)).fetchone()
                    if not product:
                        product = conn.execute('SELECT * FROM products WHERE name LIKE ? LIMIT 1', (f"%{item['name']}%",)).fetchone()
                    if product:
                        product = dict(product)
                    else:
                        product = create_hotel_flow_product(conn, item['name'], item['spec'], item['unit'], item['price'])
                        created_products += 1

                    price = (
                        item['price'] if item['price'] > 0
                        else get_customer_product_price_with_conn(conn, order['customer'], product, item['unit'])
                        or as_float(product.get('last_sale_price'))
                        or as_float(product.get('cost'))
                    )
                    if price <= 0:
                        no_price.append(item['name'])

                    rows_data.append({
                        'date': order['date'],
                        'customer': order['customer'],
                        'company': '',
                        'phone': '',
                        'order_key': f"{order['order_key']}-{hotel_sales_order_key(order['customer'], item['name'])}",
                        'product': product,
                        'product_name': item['name'],
                        'product_unit': item['unit'],
                        'quantity': item['quantity'],
                        'price': price,
                        'note': item.get('note') or ''
                    })
            conn.commit()
        except Exception as e:
            conn.rollback()
            return jsonify({'error': str(e)}), 500
        finally:
            conn.close()

        result, error, status = create_hotel_sales_orders(rows_data)
        if error:
            return jsonify(error), status
        result.update({
            'orders': len(orders),
            'createdProducts': created_products,
            'skipped': skipped,
            'noPrice': sorted(set(no_price))
        })
        return jsonify(result)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ---------- Excel导出 ----------

@app.route('/api/export/inventory', methods=['GET'])
def export_inventory_excel():
    """导出库存表为CSV"""
    rows = db.query('SELECT * FROM products ORDER BY code')

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['编号', '名称', '规格', '单位', '采购单价', '库存', '预警线', '库存总值', '状态', '专属部门'])

    for p in rows:
        status = '正常'
        stock = 0 if is_flow_item(p) else p['stock']
        inventory_value = 0 if is_flow_item(p) else round(stock * p['cost'], 2)
        if is_flow_item(p):
            status = '生鲜流转'
        elif stock <= 0:
            status = '缺货'
        elif stock <= p['alert_line']:
            status = '低库存'

        depts = json.loads(p.get('departments', '[]') or '[]')
        dept_str = ','.join(depts) if depts else '通用'

        writer.writerow([
            p['code'], p['name'], p['spec'] or '', p['unit'],
            p['cost'], stock, p['alert_line'],
            inventory_value, status, dept_str
        ])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'库存表_{date.today().strftime("%Y-%m-%d")}.csv'
    )


@app.route('/api/export/statement', methods=['GET'])
def export_statement_excel():
    """导出对账单为CSV"""
    # 采购
    purchases = db.query('SELECT date, no, supplier_name, total FROM purchases ORDER BY date DESC')
    # 领用
    outbounds = db.query('SELECT date, no, department, total FROM outbounds ORDER BY date DESC')
    # 财务
    finances = db.query('SELECT date, type, category, amount, note FROM finances ORDER BY date DESC')

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['日期', '类型', '单号/类别', '收支', '金额', '备注'])

    total_purchase = 0
    total_outbound = 0

    for p in purchases:
        writer.writerow([p['date'], '采购入库', p['no'], '支出', p['total'], p['supplier_name']])
        total_purchase += p['total']

    for o in outbounds:
        writer.writerow([o['date'], '部门领用', o['no'], '支出', o['total'], o['department']])
        total_outbound += o['total']

    for f in finances:
        direction = '收入' if f['type'] == 'income' else '支出'
        writer.writerow([f['date'], '财务记账', f['category'], direction, f['amount'], f['note']])

    writer.writerow([])
    writer.writerow(['', '', '', '累计采购支出', total_purchase, ''])
    writer.writerow(['', '', '', '累计部门领用', total_outbound, ''])
    writer.writerow(['', '', '', '净额', total_purchase + total_outbound, ''])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'对账单_{date.today().strftime("%Y-%m-%d")}.csv'
    )


@app.route('/api/export/sales-stats', methods=['GET'])
def export_sales_stats_excel():
    """导出销售统计为CSV"""
    mode = request.args.get('mode', 'order')

    output = io.StringIO()
    writer = csv.writer(output)

    if mode == 'order':
        writer.writerow(['客户', '单号', '日期', '物品数', '成本', '销售额', '利润', '毛利率'])
        rows = db.query(f'''
            SELECT s.customer, s.no, s.date, COUNT(si.id) as item_count,
                   COALESCE(SUM({HOTEL_COST_SQL}), 0) as cost,
                   s.total as revenue
            FROM sales s
            LEFT JOIN sales_items si ON s.id = si.sale_id
            LEFT JOIN products p ON si.product_id = p.id
            GROUP BY s.id
            ORDER BY s.date DESC
        ''')
        for r in rows:
            profit = (r['revenue'] or 0) - (r['cost'] or 0)
            margin = (profit / r['revenue'] * 100) if r['revenue'] else 0
            writer.writerow([
                r['customer'], r['no'], r['date'], r['item_count'],
                round(r['cost'] or 0, 2), round(r['revenue'] or 0, 2),
                round(profit, 2), f'{margin:.1f}%'
            ])
    else:
        writer.writerow(['日期', '销售单数', '成本', '销售额', '利润'])
        rows = db.query(f'''
            SELECT date, COUNT(*) as order_count,
                   COALESCE(SUM(cost), 0) as cost,
                   COALESCE(SUM(revenue), 0) as revenue
            FROM (
                SELECT s.id, s.date, s.total as revenue,
                       COALESCE(SUM({HOTEL_COST_SQL}), 0) as cost
                FROM sales s
                LEFT JOIN sales_items si ON s.id = si.sale_id
                LEFT JOIN products p ON si.product_id = p.id
                GROUP BY s.id
            )
            GROUP BY date
            ORDER BY date DESC
        ''')
        for r in rows:
            profit = (r['revenue'] or 0) - (r['cost'] or 0)
            writer.writerow([
                r['date'], r['order_count'],
                round(r['cost'] or 0, 2), round(r['revenue'] or 0, 2), round(profit, 2)
            ])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'销售统计_{date.today().strftime("%Y-%m-%d")}.csv'
    )


# ---------- 启动 ----------

start_feishu_sync_thread()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', '5011'))
    host = os.environ.get('HOST', '0.0.0.0')
    local_url = f'http://localhost:{port}'
    print('=' * 50)
    print('酒店订单管理系统')
    print('=' * 50)
    print(f'数据库: {DB_PATH}')
    print(f'访问地址: {local_url}')
    print('按 Ctrl+C 停止服务')
    print('=' * 50)

    # 自动打开浏览器
    if os.environ.get('NO_BROWSER', '') != '1':
        webbrowser.open(local_url)

    app.run(host=host, port=port, debug=False)
